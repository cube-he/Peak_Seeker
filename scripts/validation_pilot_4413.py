# -*- coding: utf-8 -*-
"""
4413 Validation Pilot: build _verified.xlsx from image ground truth.
Ground truth source: 4413_003.jpg (page 4) + 4413_004.jpg (page 5).
Authority: images > corrected > mimo-v2-omni.
"""
import os
import sys
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

sys.stdout.reconfigure(encoding='utf-8')

BASE = r'C:\Users\Administrator\Documents\VolunteerHelper\data\13_征集志愿\普通高考\少数民族\4413_征集志愿_第一次_2025_本科批'
MIMO = os.path.join(BASE, '4413_征集志愿_第一次_2025_mimo-v2-omni.xlsx')
CORR = os.path.join(BASE, '4413_征集志愿_第一次_2025_mimo-v2-omni_corrected.xlsx')
OUT = os.path.join(BASE, '4413_征集志愿_第一次_2025_mimo-v2-omni_verified.xlsx')

COLS = ['科类', '招生类型', '院校代码', '院校名称', '办学性质', '院校地址', '院校备注',
        '专业组代码', '再选科目要求', '专业组计划数', '专业代码', '专业名称', '专业备注',
        '专业计划数', '收费标准', '页码']

# -------- Ground truth rows (from image 003 + 004) --------
# Each dict: column -> value. All 14 rows in logical order.
GT = [
    # --- 历史类 省属高校少数民族预科 (page 4, image 003) ---
    {'科类': '历史类', '招生类型': '省属高校少数民族预科', '院校代码': 5137, '院校名称': '成都大学', '办学性质': '公办',
     '院校地址': '四川省成都市', '院校备注': None, '专业组代码': 108, '再选科目要求': '不限', '专业组计划数': 1,
     '专业代码': 'Y5', '专业名称': '少数民族预科(旅游管理)', '专业备注': '预科培养地点为西昌学院。',
     '专业计划数': 1, '收费标准': 5500, '页码': 4},
    {'科类': '历史类', '招生类型': '省属高校少数民族预科', '院校代码': 5137, '院校名称': '成都大学', '办学性质': '公办',
     '院校地址': '四川省成都市', '院校备注': None, '专业组代码': 109, '再选科目要求': '化学', '专业组计划数': 2,
     '专业代码': 'Y8', '专业名称': '少数民族预科(护理学)', '专业备注': '预科培养地点为西昌学院。',
     '专业计划数': 2, '收费标准': 5500, '页码': 4},
    # --- 物理类 区域教育均衡发展专项计划 (page 4, image 003) ---
    {'科类': '物理类', '招生类型': '区域教育均衡发展专项计划', '院校代码': 5108, '院校名称': '四川轻化工大学', '办学性质': '公办',
     '院校地址': '四川省自贡市', '院校备注': None, '专业组代码': 812, '再选科目要求': '化学', '专业组计划数': 1,
     '专业代码': '8T', '专业名称': '食品科学与工程(雅安市)', '专业备注': None,
     '专业计划数': 1, '收费标准': 5980, '页码': 4},
    {'科类': '物理类', '招生类型': '区域教育均衡发展专项计划', '院校代码': 5119, '院校名称': '成都医学院', '办学性质': '公办',
     '院校地址': '四川省成都市', '院校备注': '色盲、色弱考生不予录取。新都校区就读。', '专业组代码': 103, '再选科目要求': '化学', '专业组计划数': 2,
     '专业代码': 'J8', '专业名称': '护理学(凉山州)', '专业备注': None,
     '专业计划数': 1, '收费标准': 6670, '页码': 4},
    {'科类': '物理类', '招生类型': '区域教育均衡发展专项计划', '院校代码': 5119, '院校名称': '成都医学院', '办学性质': '公办',
     '院校地址': '四川省成都市', '院校备注': '色盲、色弱考生不予录取。新都校区就读。', '专业组代码': 103, '再选科目要求': '化学', '专业组计划数': 2,
     '专业代码': 'JD', '专业名称': '智能医学工程(凉山州)', '专业备注': None,
     '专业计划数': 1, '收费标准': 6670, '页码': 4},
    {'科类': '物理类', '招生类型': '区域教育均衡发展专项计划', '院校代码': 5123, '院校名称': '绵阳师范学院', '办学性质': '公办',
     '院校地址': '四川省绵阳市', '院校备注': None, '专业组代码': 108, '再选科目要求': '不限', '专业组计划数': 1,
     '专业代码': 'QH', '专业名称': '物流管理(雅安市)', '专业备注': None,
     '专业计划数': 1, '收费标准': 5520, '页码': 4},
    # --- 物理类 普通类 (page 5, image 004) ---
    {'科类': '物理类', '招生类型': '普通类', '院校代码': 5125, '院校名称': '宜宾学院', '办学性质': '公办',
     '院校地址': '四川省宜宾市', '院校备注': None, '专业组代码': 107, '再选科目要求': '化学', '专业组计划数': 1,
     '专业代码': '5D', '专业名称': '制药工程(雅安市)(江北校区A区)', '专业备注': None,
     '专业计划数': 1, '收费标准': 5980, '页码': 5},
    {'科类': '物理类', '招生类型': '普通类', '院校代码': 5130, '院校名称': '阿坝师范学院', '办学性质': '公办',
     '院校地址': '四川省阿坝州', '院校备注': None, '专业组代码': 104, '再选科目要求': '不限', '专业组计划数': 1,
     '专业代码': '2G', '专业名称': '学前教育(雅安市)(师范)', '专业备注': None,
     '专业计划数': 1, '收费标准': 4800, '页码': 5},
    {'科类': '物理类', '招生类型': '普通类', '院校代码': 5137, '院校名称': '成都大学', '办学性质': '公办',
     '院校地址': '四川省成都市', '院校备注': None, '专业组代码': 105, '再选科目要求': '化学', '专业组计划数': 1,
     '专业代码': '37', '专业名称': '环境工程(凉山州)', '专业备注': None,
     '专业计划数': 1, '收费标准': 5980, '页码': 5},
    {'科类': '物理类', '招生类型': '普通类', '院校代码': 5142, '院校名称': '成都工业学院', '办学性质': '公办',
     '院校地址': '四川省成都市', '院校备注': None, '专业组代码': 146, '再选科目要求': '化学', '专业组计划数': 1,
     '专业代码': '47', '专业名称': '汽车服务工程(乐山市)(宜宾校区)', '专业备注': '入学后不允许跨校区转专业。',
     '专业计划数': 1, '收费标准': 5200, '页码': 5},
    {'科类': '物理类', '招生类型': '普通类', '院校代码': 5142, '院校名称': '成都工业学院', '办学性质': '公办',
     '院校地址': '四川省成都市', '院校备注': None, '专业组代码': 147, '再选科目要求': '化学', '专业组计划数': 3,
     '专业代码': '48', '专业名称': '汽车服务工程(雅安市)(宜宾校区)', '专业备注': '入学后不允许跨校区转专业。',
     '专业计划数': 3, '收费标准': 5200, '页码': 5},
    # --- 物理类 省属高校少数民族预科 (page 5, image 004) ---
    # Image shows "专业组109（再选科目：不限）"
    {'科类': '物理类', '招生类型': '省属高校少数民族预科', '院校代码': 5119, '院校名称': '成都医学院', '办学性质': '公办',
     '院校地址': '四川省成都市', '院校备注': None, '专业组代码': 109, '再选科目要求': '不限', '专业组计划数': 2,
     '专业代码': 'Y9', '专业名称': '少数民族预科(应用心理学)(预科培养地点为四川民族学院。)',
     '专业备注': '色盲、色弱考生不予录取。', '专业计划数': 2, '收费标准': 5980, '页码': 5},
    {'科类': '物理类', '招生类型': '省属高校少数民族预科', '院校代码': 5119, '院校名称': '成都医学院', '办学性质': '公办',
     '院校地址': '四川省成都市', '院校备注': None, '专业组代码': 110, '再选科目要求': '化学', '专业组计划数': 2,
     '专业代码': 'Y8', '专业名称': '少数民族预科(护理学)(预科培养地点为四川民族学院。)',
     '专业备注': '色盲、色弱考生不予录取。', '专业计划数': 2, '收费标准': 6670, '页码': 5},
    {'科类': '物理类', '招生类型': '省属高校少数民族预科', '院校代码': 5142, '院校名称': '成都工业学院', '办学性质': '公办',
     '院校地址': '四川省成都市', '院校备注': None, '专业组代码': 130, '再选科目要求': '化学', '专业组计划数': 34,
     '专业代码': 'Y1', '专业名称': '少数民族预科(预科培养地点为阿坝师范学院。)',
     '专业备注': '宜宾校区就读。', '专业计划数': 34, '收费标准': 5500, '页码': 5},
]

# ---------- Load the two source xlsx ----------
def load_rows(path):
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    header = [c.value for c in ws[1]]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in r):
            continue
        rows.append(dict(zip(header, r)))
    return rows

mimo_rows = load_rows(MIMO)
corr_rows = load_rows(CORR)

# Build row key for matching: (科类, 招生类型, 院校代码, 专业组代码, 专业代码)
# Force str() on every key component so int (GT) and str (xlsx) match.
def key(r):
    def s(v):
        return str(v) if v is not None else None
    return (s(r.get('科类')), s(r.get('招生类型')), s(r.get('院校代码')),
            s(r.get('专业组代码')), s(r.get('专业代码')))

mimo_map = {key(r): r for r in mimo_rows}
corr_map = {key(r): r for r in corr_rows}

# ---------- Build verified workbook ----------
wb = Workbook()

# ====== Sheet 1: 真值数据 ======
ws1 = wb.active
ws1.title = '真值数据'

yellow = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
red = PatternFill(start_color='FF6666', end_color='FF6666', fill_type='solid')
gray = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')
header_font = Font(bold=True)

for ci, col in enumerate(COLS, start=1):
    c = ws1.cell(row=1, column=ci, value=col)
    c.font = header_font
    c.fill = gray

diff_engines = []   # for Sheet 3
change_log = []     # for Sheet 2

for ri, gt in enumerate(GT, start=2):
    k = key(gt)
    m = mimo_map.get(k)
    cr = corr_map.get(k)
    src_img = '4413_003.jpg' if gt['页码'] == 4 else '4413_004.jpg'

    for ci, col in enumerate(COLS, start=1):
        gt_val = gt.get(col)
        cell = ws1.cell(row=ri, column=ci, value=gt_val)

        m_val = m.get(col) if m else None
        c_val = cr.get(col) if cr else None

        import math
        def norm(v):
            if v is None:
                return None
            if isinstance(v, float) and math.isnan(v):
                return None
            # string-normalize for comparison so '5137' == 5137
            return str(v).strip()

        m_val_n = norm(m_val)
        c_val_n = norm(c_val)
        gt_val_n = norm(gt_val)

        row_missing_in_mimo = (m is None)

        # Engine-vs-engine diff (Sheet 3)
        if m_val_n != c_val_n:
            diff_engines.append({
                '行号': ri, '列名': col,
                'mimo-v2-omni值': m_val if m is not None else '(整行缺失)',
                'corrected值': c_val if cr is not None else '(整行缺失)',
                '最终采用': gt_val,
                '备注': '两版引擎不一致' if (m and cr) else '仅 corrected 有',
            })

        # Compare GT against best available engine value (corrected preferred)
        # Fill yellow only for cells that differ from BOTH engines OR for rows mimo missing
        needs_mark = False
        reason = ''
        if row_missing_in_mimo:
            # Entire row is a correction (mimo omitted 2 历史类 rows).
            # Mark every non-null cell yellow to flag it was added.
            needs_mark = True
            reason = 'mimo-v2-omni 漏此行，corrected 补入；与图片一致'
        else:
            # Both engines have the row. Any disagreement with GT?
            # If GT == corrected == mimo: no fill.
            # Else mark yellow.
            if gt_val_n != c_val_n or gt_val_n != m_val_n:
                needs_mark = True
                if gt_val_n == c_val_n and gt_val_n != m_val_n:
                    reason = '采用 corrected 值（与图片一致）'
                elif gt_val_n == m_val_n and gt_val_n != c_val_n:
                    reason = '采用 mimo 值（与图片一致）'
                else:
                    reason = '两版均与图片不符，按图片修正'

        if needs_mark:
            cell.fill = yellow
            change_log.append({
                '行号': ri,
                '院校代码_专业代码': f"{gt.get('院校代码')}_{gt.get('专业代码')}",
                '列名': col,
                'mimo-v2-omni原值': m_val if m is not None else '(整行缺失)',
                'corrected原值': c_val if cr is not None else '(整行缺失)',
                '最终真值': gt_val,
                '图片来源': src_img,
                '修改依据': reason,
                '置信度': '高',
            })

# Column widths
widths = {
    '科类': 8, '招生类型': 22, '院校代码': 10, '院校名称': 16, '办学性质': 10,
    '院校地址': 14, '院校备注': 30, '专业组代码': 10, '再选科目要求': 12,
    '专业组计划数': 12, '专业代码': 10, '专业名称': 38, '专业备注': 28,
    '专业计划数': 10, '收费标准': 10, '页码': 6,
}
for ci, col in enumerate(COLS, start=1):
    ws1.column_dimensions[ws1.cell(row=1, column=ci).column_letter].width = widths.get(col, 12)

# ====== Sheet 2: 修改记录 ======
ws2 = wb.create_sheet('修改记录')
cols2 = ['行号', '院校代码_专业代码', '列名', 'mimo-v2-omni原值', 'corrected原值',
         '最终真值', '图片来源', '修改依据', '置信度']
for ci, col in enumerate(cols2, start=1):
    c = ws2.cell(row=1, column=ci, value=col)
    c.font = header_font
    c.fill = gray
for ri, rec in enumerate(change_log, start=2):
    for ci, col in enumerate(cols2, start=1):
        ws2.cell(row=ri, column=ci, value=rec.get(col))
for ci, col in enumerate(cols2, start=1):
    ws2.column_dimensions[ws2.cell(row=1, column=ci).column_letter].width = 22

# ====== Sheet 3: 引擎差异 ======
ws3 = wb.create_sheet('引擎差异')
cols3 = ['行号', '列名', 'mimo-v2-omni值', 'corrected值', '最终采用', '备注']
for ci, col in enumerate(cols3, start=1):
    c = ws3.cell(row=1, column=ci, value=col)
    c.font = header_font
    c.fill = gray
for ri, rec in enumerate(diff_engines, start=2):
    for ci, col in enumerate(cols3, start=1):
        ws3.cell(row=ri, column=ci, value=rec.get(col))
for ci, col in enumerate(cols3, start=1):
    ws3.column_dimensions[ws3.cell(row=1, column=ci).column_letter].width = 28

wb.save(OUT)

print(f'Saved: {OUT}')
print(f'GT rows: {len(GT)}')
print(f'Change log entries: {len(change_log)}')
print(f'Engine-diff entries: {len(diff_engines)}')
print(f'mimo rows: {len(mimo_rows)}; corrected rows: {len(corr_rows)}')

# Breakdown
yellow_cnt = len(change_log)
red_cnt = 0  # no unclear image cells in this folder
print(f'Yellow fills: {yellow_cnt}')
print(f'Red fills: {red_cnt}')
