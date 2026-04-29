#!/usr/bin/env python3
"""校验 BID=3339 征集志愿数据"""
import sys
import io
import re
import openpyxl
from openpyxl.styles import Font, PatternFill

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# Load source
src_path = 'data/13_征集志愿/普通高考/专科批次/3339_2023_理科_专科批_征集志愿_第一次/3339_2023_理科_专科批_征集志愿_第一次_mimo-v2-omni.xlsx'
src = openpyxl.load_workbook(src_path)
ws_src = src.active

# Read all OCR data
ocr_data = []
for row in range(2, ws_src.max_row + 1):
    vals = [ws_src.cell(row, c).value for c in range(1, ws_src.max_column + 1)]
    ocr_data.append(vals)

print(f"Loaded {len(ocr_data)} OCR rows")

# Create output
wb = openpyxl.Workbook()
ws = wb.active
ws.title = '征集志愿'

# Styles
font_default = Font(name='微软雅黑', size=10)
font_bold = Font(name='微软雅黑', size=10, bold=True)
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
green_fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
red_fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')

# Headers
headers = ['科类', '院校代码', '院校名称', '院校地址', '专业代码', '专业名称',
           '专业备注', '专业计划数', '收费标准', '院校备注', '页码', '校正备注']
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = font_bold


def split_major_name(name):
    """Split brackets from major name into notes."""
    if not name or '(' not in name:
        return name, None
    idx = name.index('(')
    pure_name = name[:idx]
    brackets = name[idx:]
    parts = re.findall(r'\(([^)]+)\)', brackets)
    note = '；'.join(parts) if parts else brackets
    return pure_name, note


out_row = 2

def write_row(data, fill=None, notes=''):
    """Write a row. data = [科类..页码] (11 cols)"""
    global out_row
    d = list(data)

    # Ensure code is string
    if d[1] is not None:
        d[1] = str(d[1])

    # Split major name brackets
    orig_name = d[5] if d[5] else ''
    orig_note = d[6] if d[6] else ''

    pure_name, bracket_note = split_major_name(str(orig_name))

    # Combine notes
    final_note = str(orig_note) if orig_note else ''
    if bracket_note:
        if final_note:
            final_note = final_note + '；' + bracket_note
        else:
            final_note = bracket_note

    out_data = d[:5] + [pure_name, final_note if final_note else None] + d[7:]

    for col, val in enumerate(out_data, 1):
        cell = ws.cell(row=out_row, column=col, value=val)
        cell.font = font_default
        if fill:
            cell.fill = fill

    # Correction notes col 12
    if notes:
        cell = ws.cell(row=out_row, column=12, value=notes)
        cell.font = font_default
        if fill:
            cell.fill = fill

    out_row += 1


def write_ocr_rows(start_idx, end_idx, fixes=None):
    """Write OCR rows [start_idx, end_idx), applying fixes dict."""
    if fixes is None:
        fixes = {}
    for i in range(start_idx, end_idx):
        d = list(ocr_data[i])
        corrections = ''
        fill = None

        # Apply fixes
        if i in fixes:
            for fix in fixes[i]:
                field, old_val, new_val, desc = fix
                if field == 4:  # 专业代码
                    d[4] = new_val
                elif field == 6:  # 专业备注
                    if d[6] and old_val in str(d[6]):
                        d[6] = str(d[6]).replace(old_val, new_val)
                elif field == 1:  # 院校代码
                    d[1] = new_val
                elif field == 2:  # 院校名称
                    d[2] = new_val
                if corrections:
                    corrections += '；'
                corrections += desc
            fill = yellow_fill

        write_row(d, fill=fill, notes=corrections)


def write_missing(rows, note_prefix):
    """Write missing rows with green fill."""
    for d in rows:
        write_row(d, fill=green_fill, notes=note_prefix)


# ========== PAGE 1 ==========
# OCR rows 0-10 (R2-R12): schools 1371,1374,2152,2175,2177,2178,2180
write_ocr_rows(0, 11)

# ========== PAGE 2 (missing 2185) ==========
# Add missing school 2185 (8 specialties)
addr_2185 = '大连市甘井子区红旗西路600号'
note_2185 = '认同并全部执行四川省少数民族加分项目和分值。'
write_missing([
    ['理科', '2185', '大连软件职业学院', addr_2185, '01', '软件技术', None, 1, '14800', note_2185, 2],
    ['理科', '2185', '大连软件职业学院', addr_2185, '02', '大数据技术', None, 2, '14800', note_2185, 2],
    ['理科', '2185', '大连软件职业学院', addr_2185, '03', '数字媒体技术', None, 3, '12800', note_2185, 2],
    ['理科', '2185', '大连软件职业学院', addr_2185, '04', '计算机网络技术', None, 2, '12800', note_2185, 2],
    ['理科', '2185', '大连软件职业学院', addr_2185, '05', '电子商务', None, 2, '12800', note_2185, 2],
    ['理科', '2185', '大连软件职业学院', addr_2185, '06', '现代物流管理', None, 1, '9800', note_2185, 2],
    ['理科', '2185', '大连软件职业学院', addr_2185, '07', '市场营销', None, 2, '9800', note_2185, 2],
    ['理科', '2185', '大连软件职业学院', addr_2185, '08', '酒店管理与数字化运营', None, 1, '9800', note_2185, 2],
], '缺漏：OCR遗漏整所院校2185')

# OCR rows 11-25 (R13-R27): 2188,2189,2190,2553,2554
# Fix 2189 codes: 42->4Z, 02->6Z
write_ocr_rows(11, 26, fixes={
    14: [(4, '42', '4Z', '专业代码42→4Z')],
    15: [(4, '02', '6Z', '专业代码02→6Z')],
})

# ========== PAGE 3 (missing 2554/1B) ==========
write_missing([
    ['理科', '2554', '辽宁广告职业学院', '辽宁省沈阳市于洪区造化街道', '1B', '电子商务', None, 3, '9800', None, 3],
], '缺漏：OCR遗漏2554跨页专业1B')

# OCR rows 26-37 (R28-R39): 2555,2351,2353,2370,2398,3171,3273
write_ocr_rows(26, 38)

# ========== PAGE 4 ==========
# OCR rows 38-51 (R40-R52): 3878,3910,3929,3958,3980,3464,3881
write_ocr_rows(38, 51)

# ========== PAGE 5 (missing 3523) ==========
addr_3523 = '福建省南安市教育科技园区学园中路'
note_3523 = '认同并全部执行四川省少数民族加分项目和分值。'
write_missing([
    ['理科', '3523', '泉州工程职业技术学院', addr_3523, '0B', '建设工程管理', None, 1, '8500', note_3523, 5],
    ['理科', '3523', '泉州工程职业技术学院', addr_3523, '0F', '软件技术', None, 2, '9800', note_3523, 5],
], '缺漏：OCR遗漏整所院校3523')

# OCR rows 51-67 (R53-R69): 3524,3555,3556,3557,3561,3567
write_ocr_rows(51, 68)

# ========== PAGE 6 (missing 3567 continued) ==========
write_missing([
    ['理科', '3567', '厦门华天涉外职业技术学院', '厦门市翔安区新店镇洪钟大道5088-5200号', '05', '建设工程管理', None, 2, '11900', None, 6],
    ['理科', '3567', '厦门华天涉外职业技术学院', '厦门市翔安区新店镇洪钟大道5088-5200号', '15', '市场营销', None, 1, '12800', None, 6],
], '缺漏：OCR遗漏3567跨页专业')

# OCR rows 68-88 (R70-R89): 3572,3576,3578,3579,3581
write_ocr_rows(68, 88)

# ========== PAGE 7 ==========
# OCR rows 88-101 (R90-R102): 3583,3588,3626,3822,3957,3659,3674,3683
write_ocr_rows(88, 101)

# ========== PAGE 8 (missing 3683/09) ==========
write_missing([
    ['理科', '3683', '共青科技职业学院', '江西省九江市共青城市共青大道1号', '09', '智能制造装备技术', None, 2, '10320', None, 8],
], '缺漏：OCR遗漏3683跨页专业09')

# OCR rows 101-116 (R103-R116): 3780,3790,3819,3909,3961,3970
write_ocr_rows(101, 116)

# ========== PAGE 9 (missing 3988) ==========
addr_3988 = '山东省枣庄市新城祁连山路2169号'
note_3988 = '认同并全部执行四川省少数民族加分项目和分值。'
write_missing([
    ['理科', '3988', '枣庄职业学院', addr_3988, '4A', '计算机应用技术', '中外合作办学；与英国胡弗汉顿大学合作', 2, '14000', note_3988, 9],
    ['理科', '3988', '枣庄职业学院', addr_3988, '4B', '汽车制造与试验技术', '中外合作办学；与德国比勒菲尔德技术大学合作', 3, '15000', note_3988, 9],
], '缺漏：OCR遗漏整所院校3988')

# OCR rows 116-127 (R117-R128): 3994,4253,4254,4258,4269,4952,4982
write_ocr_rows(116, 127)

# ========== PAGE 10 (missing 4982 continued) ==========
write_missing([
    ['理科', '4982', '湖南外国语职业学院', '湖南省长沙市望城区丁字湾街道京阳大道', '03', '应用法语', None, 2, '18800', None, 10],
    ['理科', '4982', '湖南外国语职业学院', '湖南省长沙市望城区丁字湾街道京阳大道', '04', '应用西班牙语', None, 2, '18800', None, 10],
], '缺漏：OCR遗漏4982跨页专业')

# OCR rows 127-142 (R129-R143): 4466,4472,4476,4479,4499,4780
write_ocr_rows(127, 142)

# ========== PAGE 11 ==========
# OCR rows 142-157 (R144-R158): 4783,4798,4915,4967,4650,4652,4656
write_ocr_rows(142, 157)

# ========== PAGE 12 ==========
# OCR rows 157-166 (R159-R167): 4661,4692,5032,5033,5036,5067,5146,5152
write_ocr_rows(157, 166)

# ========== PAGE 13 (missing 5153) ==========
write_missing([
    ['理科', '5153', '达州职业技术学院', '四川省达州市通川区韩家坝犀牛大道书山路1号', '42', '学前教育', '师范；徐家坝校区', 7, '4800', None, 13],
], '缺漏：OCR遗漏整所院校5153')

# OCR rows 166-175 (R168-R176): 5154,5158,5168,5169,5172,5175,5176,5183,5184
# Fix 5169 code: 20->2C
write_ocr_rows(166, 175, fixes={
    169: [(4, '20', '2C', '专业代码20→2C')],
})

# Missing 5186 (6 specialties, pages 13-14)
addr_5186 = '天府新区视高经济开发区花海大道大学路1号'
write_missing([
    ['理科', '5186', '四川科技职业学院', addr_5186, '03', '软件技术', None, 3, '15000', None, 14],
    ['理科', '5186', '四川科技职业学院', addr_5186, '12', '新能源汽车技术', None, 5, '15000', None, 14],
    ['理科', '5186', '四川科技职业学院', addr_5186, '3A', '药物制剂技术', None, 5, '15000', None, 14],
    ['理科', '5186', '四川科技职业学院', addr_5186, '55', '大数据技术', None, 5, '15000', None, 14],
    ['理科', '5186', '四川科技职业学院', addr_5186, '5A', '中医康复技术', None, 3, '16000', None, 14],
    ['理科', '5186', '四川科技职业学院', addr_5186, '73', '工程造价', None, 2, '14000', None, 14],
], '缺漏：OCR遗漏整所院校5186')

# ========== PAGE 14 ==========
# OCR rows 175-208 (R177-R209): 5192, 5193
write_ocr_rows(175, 208)

# ========== PAGE 15 ==========
# OCR rows 208-237 (R210-R238): 5193 continued, 5197, 5198, 5652, 5660, 5783
write_ocr_rows(208, 237)

# ========== PAGE 16 (missing 5783 continued) ==========
addr_5783 = '四川省天府新区剑南大道南延线航空大道中段168号'
write_missing([
    ['理科', '5783', '天府新区航空旅游职业学院', addr_5783, '04', '民航运输服务', None, 2, '16200', None, 16],
    ['理科', '5783', '天府新区航空旅游职业学院', addr_5783, '05', '航空物流管理', None, 3, '16200', None, 16],
    ['理科', '5783', '天府新区航空旅游职业学院', addr_5783, '06', '生物产品检验检疫', None, 2, '13800', None, 16],
    ['理科', '5783', '天府新区航空旅游职业学院', addr_5783, '07', '社区康复', None, 1, '13800', None, 16],
    ['理科', '5783', '天府新区航空旅游职业学院', addr_5783, '08', '健康管理', None, 1, '13800', None, 16],
    ['理科', '5783', '天府新区航空旅游职业学院', addr_5783, '09', '高速铁路客运服务', None, 1, '13200', None, 16],
    ['理科', '5783', '天府新区航空旅游职业学院', addr_5783, '0A', '城市轨道交通运营管理', None, 1, '13200', None, 16],
    ['理科', '5783', '天府新区航空旅游职业学院', addr_5783, '0B', '旅游管理', None, 4, '12000', None, 16],
    ['理科', '5783', '天府新区航空旅游职业学院', addr_5783, '0C', '会展策划与管理', None, 4, '12000', None, 16],
    ['理科', '5783', '天府新区航空旅游职业学院', addr_5783, '0D', '飞机机电设备维修', None, 3, '16200', None, 16],
    ['理科', '5783', '天府新区航空旅游职业学院', addr_5783, '0E', '飞机电子设备维修', None, 4, '16200', None, 16],
    ['理科', '5783', '天府新区航空旅游职业学院', addr_5783, '0F', '无人机应用技术', None, 1, '13200', None, 16],
    ['理科', '5783', '天府新区航空旅游职业学院', addr_5783, '0G', '电子商务', None, 2, '13200', None, 16],
    ['理科', '5783', '天府新区航空旅游职业学院', addr_5783, '0H', '软件技术', None, 2, '13200', None, 16],
    ['理科', '5783', '天府新区航空旅游职业学院', addr_5783, '0J', '计算机应用技术', None, 2, '13200', None, 16],
], '缺漏：OCR遗漏5783跨页专业')

# OCR rows 237-251 (R239-R252): 5792,5794,5795,5888
write_ocr_rows(237, 251)

# ========== PAGE 17 (missing 5256) ==========
write_missing([
    ['理科', '5256', '安顺职业技术学院', '贵州省安顺市西秀区工业园区两六路与二环路交叉口', '06', '助产', None, 1, '3500', '认同并全部执行四川省少数民族加分项目和分值。', 17],
], '缺漏：OCR遗漏整所院校5256')

# OCR rows 251-262 (R253-R263): 5257,5290,5913,5937,5361,5382,5389,5697
# Fix: 5290 code 00->0G + 备注; 5382 code 22->Z2; 5697 missing code/name
write_ocr_rows(251, 262, fixes={
    252: [(4, '00', '0G', '专业代码00→0G'), (6, '朱校区', '东校区', '专业备注朱校区→东校区')],
    257: [(4, '22', 'Z2', '专业代码22→Z2')],
    260: [(1, None, '5697', '补充院校代码5697'), (2, None, '博尔塔拉职业技术学院', '补充院校名称')],
    261: [(1, None, '5697', '补充院校代码5697'), (2, None, '博尔塔拉职业技术学院', '补充院校名称')],
})

# ========== PAGE 18 ==========
# OCR rows 262-267 (R264-R268): 6557,6558,6562
write_ocr_rows(262, 267)

# Save
out_path = 'data/13_征集志愿/普通高考/专科批次/3339_2023_理科_专科批_征集志愿_第一次/3339_2023_理科_专科批_征集志愿_第一次_已校验.xlsx'
wb.save(out_path)
print(f"Done! Total data rows: {out_row - 2}")
print(f"Output: {out_path}")

# Count stats
green_count = 0
yellow_count = 0
for r in range(2, out_row):
    for c in range(1, 13):
        cell = ws.cell(r, c)
        if cell.fill == green_fill:
            green_count += 1
            break
        elif cell.fill == yellow_fill:
            yellow_count += 1
            break

print(f"Green (missing): {green_count} rows")
print(f"Yellow (corrected): {yellow_count} rows")
print(f"Unchanged: {out_row - 2 - green_count - yellow_count} rows")
