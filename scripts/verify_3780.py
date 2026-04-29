# -*- coding: utf-8 -*-
"""Verify batch 3780 OCR data against original images."""

import re
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Paths
src_path = 'data/13_征集志愿/普通高考/专项计划/3780_2024_文理综合_本科一批_征集志愿_第一次/3780_2024_文理综合_本科一批_征集志愿_第一次_mimo-v2-omni.xlsx'
out_path = 'data/13_征集志愿/普通高考/专项计划/3780_2024_文理综合_本科一批_征集志愿_第一次/3780_2024_文理综合_本科一批_征集志愿_第一次_已校验.xlsx'

# Load OCR data
wb_src = load_workbook(src_path)
ws_src = wb_src.active

ocr_rows = []
for row in ws_src.iter_rows(min_row=2, max_row=ws_src.max_row, values_only=True):
    ocr_rows.append(list(row))

print(f"Total OCR rows: {len(ocr_rows)}")

# Create output workbook
wb = Workbook()
ws = wb.active
ws.title = '征集志愿'

# Styles
font_normal = Font(name='微软雅黑', size=10)
font_bold = Font(name='微软雅黑', size=10, bold=True)
fill_yellow = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
fill_green = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
fill_red = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')

# Headers
headers = ['科类', '院校代码', '院校名称', '院校地址', '专业代码', '专业名称',
           '专业备注', '专业计划数', '收费标准', '院校备注', '页码', '校正备注']
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = font_bold


def split_major_name(name):
    """Split major name at first bracket, extract all bracket contents to notes."""
    if not name:
        return name, None
    clean = name.strip()
    # Remove [V] prefix
    if clean.startswith('[V]'):
        clean = clean[3:]

    # Find first ( or full-width bracket
    idx_half = clean.find('(')
    idx_full = clean.find('\uff08')  # （

    if idx_half == -1 and idx_full == -1:
        return clean, None

    if idx_half == -1:
        idx = idx_full
    elif idx_full == -1:
        idx = idx_half
    else:
        idx = min(idx_half, idx_full)

    major = clean[:idx].strip()
    rest = clean[idx:]

    # Extract all bracket contents (both half and full width)
    notes = []
    pattern = r'[\uff08(](.*?)[\uff09)]'
    matches = re.findall(pattern, rest)
    notes = [m.strip() for m in matches if m.strip()]

    note_str = '\uff1b'.join(notes) if notes else None  # Use full-width semicolon
    return major, note_str


# Process each OCR row with corrections
output_data = []  # Each entry: (row_data_list, fill_or_None, correction_text_or_None)

for i, row in enumerate(ocr_rows):
    kl, yxdm, yxmc, yxdz, zydm, zymc, zybz, zyjh, sfbz, yxbz, ym = row
    corrections_list = []
    fill = None

    # --- Convert to strings for safety ---
    if zymc is not None:
        zymc = str(zymc)
    if zybz is not None:
        zybz = str(zybz)
    if yxbz is not None:
        yxbz = str(yxbz)

    # --- Clean [V] from major name ---
    if zymc and '[V]' in zymc:
        zymc = zymc.replace('[V]', '')
        corrections_list.append('[V]标记已清除')
        fill = fill_yellow

    # --- Clean [V] from 专业备注 ---
    if zybz and '[V]' in zybz:
        zybz = zybz.replace('[V]', '')
        corrections_list.append('专业备注[V]标记已清除')
        fill = fill_yellow

    # --- Specific field corrections ---

    # Row 9 (index 9): 0035 河海大学 专业代码 2K -> ZK
    if i == 9 and str(zydm) == '2K':
        zydm = 'ZK'
        corrections_list.append('专业代码2K->ZK')
        fill = fill_yellow

    # Row 26 (index 26): 1207 天津师范大学 院校备注缺失 (跨页)
    if i == 26 and str(yxdm) == '1207' and yxbz is None:
        yxbz = '认同并执行四川省少数民族加分项目和分值。'
        corrections_list.append('跨页院校备注补充')
        fill = fill_yellow

    # Row 61 (index 61): B1 历史学 院校信息缺失 (0363 西北民族大学)
    if i == 61 and yxdm is None and str(zydm) == 'B1':
        yxdm = '0363'
        yxmc = '西北民族大学'
        yxdz = '甘肃省兰州市城关区西北新村1号'
        corrections_list.append('跨页院校信息补充(0363西北民族大学)')
        fill = fill_yellow

    # Row 70 (index 70): 0013 中国传媒大学 院校备注缺失
    if i == 70 and str(yxdm) == '0013' and yxbz is None:
        yxbz = '只承认教育部规定的全国性加分政策。'
        corrections_list.append('跨页院校备注补充')
        fill = fill_yellow

    # Row 117 (index 117): 0055 电子科技大学 备注缺失
    if i == 117 and str(yxdm) == '0055' and yxbz is None:
        yxbz = '认同并执行四川省少数民族加分项目和分值。'
        corrections_list.append('跨页院校备注补充')
        fill = fill_yellow

    # Row 118 (index 118): 0089 西南大学(荣昌校区) 备注缺失
    if i == 118 and str(yxdm) == '0089' and yxbz is None:
        yxbz = '认同并执行四川省少数民族加分项目和分值。'
        corrections_list.append('跨页院校备注补充')
        fill = fill_yellow

    # Row 119 (index 119): 0057 西南财经大学 理科 备注缺失
    if i == 119 and str(yxdm) == '0057' and kl == '理科' and yxbz is None:
        yxbz = '认同并执行四川省少数民族加分项目和分值。'
        corrections_list.append('跨页院校备注补充')
        fill = fill_yellow

    # Row 120-121 (index 120-121): 0059 西安电子科技大学 理科 备注缺失
    if i in [120, 121] and str(yxdm) == '0059' and kl == '理科' and yxbz is None:
        yxbz = '认同并执行四川省少数民族加分项目和分值。南校区就读。'
        corrections_list.append('跨页院校备注补充')
        fill = fill_yellow

    # Row 122-123 (index 122-123): 0060 陕西师范大学 理科 备注缺失
    if i in [122, 123] and str(yxdm) == '0060' and kl == '理科' and yxbz is None:
        yxbz = '认同并执行四川省少数民族加分项目和分值。'
        corrections_list.append('跨页院校备注补充')
        fill = fill_yellow

    # Row 124 (index 124): 0061 兰州大学 备注缺失
    if i == 124 and str(yxdm) == '0061' and yxbz is None:
        yxbz = '只承认教育部规定的全国性加分政策。'
        corrections_list.append('跨页院校备注补充')
        fill = fill_yellow

    # Row 153 (index 153): 0355 中国民用航空飞行学院 院校代码缺失, 专业代码06->0G
    if i == 153 and yxdm is None and yxmc and '飞行' in str(yxmc):
        yxdm = '0355'
        corrections_list.append('跨页院校代码补充(0355)')
        fill = fill_yellow
    if i == 153 and str(zydm) == '06':
        zydm = '0G'
        corrections_list.append('专业代码06->0G')
        fill = fill_yellow

    # 3305 rows (index 203-205): 院校备注缺失
    if i in [203, 204, 205] and str(yxdm) == '3305' and yxbz is None:
        yxbz = '认同并执行四川省少数民族加分项目和分值。第一年在定海校区，一年后在新城校区就读。'
        corrections_list.append('跨页院校备注补充')
        fill = fill_yellow

    # --- Split bracket contents from major name ---
    if zymc and ('(' in zymc or '\uff08' in zymc):
        zymc_new, note_new = split_major_name(zymc)
        if note_new and zymc_new != zymc:
            # Merge with existing zybz
            if zybz:
                zybz = note_new + '\uff1b' + zybz
            else:
                zybz = note_new
            zymc = zymc_new
            corrections_list.append('专业名称括号拆分')
            if fill is None:
                fill = fill_yellow

    correction_text = '\uff1b'.join(corrections_list) if corrections_list else None
    output_data.append(([kl, yxdm, yxmc, yxdz, zydm, zymc, zybz, zyjh, sfbz, yxbz, ym],
                        fill, correction_text))


# --- Define missing records to insert ---
# Format: (after_ocr_index, list_of_tuples)
# Each tuple: (row_data, fill, correction_note)

missing_records = [
    # 0061 兰州大学: after index 124, insert 27日语 and 72农林经济管理
    (124, [
        (['理科', '0061', '兰州大学', '甘肃省兰州市天水南路222号', '27', '日语',
          None, 2, '5500', '只承认教育部规定的全国性加分政策。', 15],
         fill_green, '缺漏：跨页丢失(页14-15)'),
        (['理科', '0061', '兰州大学', '甘肃省兰州市天水南路222号', '72', '农林经济管理',
          None, 1, '5500', '只承认教育部规定的全国性加分政策。', 15],
         fill_green, '缺漏：跨页丢失(页14-15)'),
    ]),
    # 2106 辽宁工程技术大学: after index 177, insert 4 records
    (177, [
        (['理科', '2106', '辽宁工程技术大学', '辽宁省阜新市中华路47号', '04', '地质工程',
          '阜新校区中华路校园', 3, '5200', '认同并执行四川省少数民族加分项目和分值。', 19],
         fill_green, '缺漏：整所院校跨页丢失(页18-19)'),
        (['理科', '2106', '辽宁工程技术大学', '辽宁省阜新市中华路47号', '06', '安全工程',
          '葫芦岛校区龙湾校园', 2, '5200', '认同并执行四川省少数民族加分项目和分值。', 19],
         fill_green, '缺漏：整所院校跨页丢失(页18-19)'),
        (['理科', '2106', '辽宁工程技术大学', '辽宁省阜新市中华路47号', '09', '测绘工程',
          '阜新校区玉龙校园', 3, '5200', '认同并执行四川省少数民族加分项目和分值。', 19],
         fill_green, '缺漏：整所院校跨页丢失(页18-19)'),
        (['理科', '2106', '辽宁工程技术大学', '辽宁省阜新市中华路47号', '30', '智能建造',
          '阜新校区玉龙校园', 2, '5200', '认同并执行四川省少数民族加分项目和分值。', 19],
         fill_green, '缺漏：整所院校跨页丢失(页18-19)'),
    ]),
    # 3305 浙江海洋大学: after index 205, insert 25 港口航道
    (205, [
        (['理科', '3305', '浙江海洋大学', '浙江省舟山市定海区', '25', '港口航道与海岸工程',
          None, 5, '5500',
          '认同并执行四川省少数民族加分项目和分值。第一年在定海校区，一年后在新城校区就读。', 21],
         fill_green, '缺漏：跨页丢失(页20-21)'),
    ]),
    # 3705 青岛理工大学: after index 217, insert 9 records
    (217, [
        (['理科', '3705', '青岛理工大学', '青岛市黄岛区嘉陵江东路777号', '0F', '土木工程',
          None, 1, '6600', '认同并执行四川省少数民族加分项目和分值。黄岛校区就读。', 22],
         fill_green, '缺漏：整所院校跨页丢失(页21-22)'),
        (['理科', '3705', '青岛理工大学', '青岛市黄岛区嘉陵江东路777号', '0K', '材料成型及控制工程',
          None, 2, '6325', '认同并执行四川省少数民族加分项目和分值。黄岛校区就读。', 22],
         fill_green, '缺漏：整所院校跨页丢失(页21-22)'),
        (['理科', '3705', '青岛理工大学', '青岛市黄岛区嘉陵江东路777号', '0L', '给排水科学与工程',
          None, 2, '6600', '认同并执行四川省少数民族加分项目和分值。黄岛校区就读。', 22],
         fill_green, '缺漏：整所院校跨页丢失(页21-22)'),
        (['理科', '3705', '青岛理工大学', '青岛市黄岛区嘉陵江东路777号', '0M', '建筑环境与能源应用工程',
          None, 1, '6600', '认同并执行四川省少数民族加分项目和分值。黄岛校区就读。', 22],
         fill_green, '缺漏：整所院校跨页丢失(页21-22)'),
        (['理科', '3705', '青岛理工大学', '青岛市黄岛区嘉陵江东路777号', '0P', '环境工程',
          None, 2, '6600', '认同并执行四川省少数民族加分项目和分值。黄岛校区就读。', 22],
         fill_green, '缺漏：整所院校跨页丢失(页21-22)'),
        (['理科', '3705', '青岛理工大学', '青岛市黄岛区嘉陵江东路777号', '0Q', '环境科学',
          None, 2, '6325', '认同并执行四川省少数民族加分项目和分值。黄岛校区就读。', 22],
         fill_green, '缺漏：整所院校跨页丢失(页21-22)'),
        (['理科', '3705', '青岛理工大学', '青岛市黄岛区嘉陵江东路777号', '0S', '安全工程',
          None, 3, '6600', '认同并执行四川省少数民族加分项目和分值。黄岛校区就读。', 22],
         fill_green, '缺漏：整所院校跨页丢失(页21-22)'),
        (['理科', '3705', '青岛理工大学', '青岛市黄岛区嘉陵江东路777号', '14', '工程管理',
          None, 1, '6600', '认同并执行四川省少数民族加分项目和分值。黄岛校区就读。', 22],
         fill_green, '缺漏：整所院校跨页丢失(页21-22)'),
        (['理科', '3705', '青岛理工大学', '青岛市黄岛区嘉陵江东路777号', '16', '工业工程',
          None, 3, '5500', '认同并执行四川省少数民族加分项目和分值。黄岛校区就读。', 22],
         fill_green, '缺漏：整所院校跨页丢失(页21-22)'),
    ]),
    # 5102 成都理工大学: after index 236, insert 8 records
    (236, [
        (['理科', '5102', '成都理工大学', '成都市成华区二仙桥东三路1号', '33', '市场营销',
          None, 4, '5760', None, 23],
         fill_green, '缺漏：跨页丢失(页22-23)'),
        (['理科', '5102', '成都理工大学', '成都市成华区二仙桥东三路1号', '36', '新闻传播学类',
          '包含专业:广播电视学、广告学；色盲不录。', 6, '5760', None, 23],
         fill_green, '缺漏：跨页丢失(页22-23)'),
        (['理科', '5102', '成都理工大学', '成都市成华区二仙桥东三路1号', '44', '建筑类',
          '包含专业:建筑学、风景园林；色弱、色盲不录。', 7, '6240', None, 23],
         fill_green, '缺漏：跨页丢失(页22-23)'),
        (['理科', '5102', '成都理工大学', '成都市成华区二仙桥东三路1号', '45', '旅游管理',
          None, 1, '5760', None, 23],
         fill_green, '缺漏：跨页丢失(页22-23)'),
        (['理科', '5102', '成都理工大学', '成都市成华区二仙桥东三路1号', '46', '人文地理与城乡规划',
          '色弱、色盲不录。', 6, '6240', None, 23],
         fill_green, '缺漏：跨页丢失(页22-23)'),
        (['理科', '5102', '成都理工大学', '成都市成华区二仙桥东三路1号', '47', '地理信息科学',
          '色弱、色盲不录。', 4, '6240', None, 23],
         fill_green, '缺漏：跨页丢失(页22-23)'),
        (['理科', '5102', '成都理工大学', '成都市成华区二仙桥东三路1号', '49', '环境科学与工程类',
          '包含专业:环境工程、环境生态工程；色弱、色盲不录。', 4, '6240', None, 23],
         fill_green, '缺漏：跨页丢失(页22-23)'),
        (['理科', '5102', '成都理工大学', '成都市成华区二仙桥东三路1号', '6H', '资源勘查工程',
          '色弱、色盲不录。', 10, '6240', None, 23],
         fill_green, '缺漏：跨页丢失(页22-23)'),
    ]),
    # 5109 西华大学: after index 260, insert 59
    (260, [
        (['理科', '5109', '西华大学', '成都市金牛区土桥金周路999号', '59', '材料科学与工程',
          None, 5, '5980', None, 24],
         fill_green, '缺漏：跨页丢失(页23-24)'),
    ]),
    # 5120 四川师范大学 理科: after index 286, insert 9 records
    (286, [
        (['理科', '5120', '四川师范大学', '成都市锦江区静安路5号', '16', '经济学',
          '成龙校区', 1, '5760', None, 25],
         fill_green, '缺漏：跨页丢失(页24-25)'),
        (['理科', '5120', '四川师范大学', '成都市锦江区静安路5号', '17', '金融工程',
          '成龙校区；数字金融方向培养', 2, '5760', None, 25],
         fill_green, '缺漏：跨页丢失(页24-25)'),
        (['理科', '5120', '四川师范大学', '成都市锦江区静安路5号', '1E', '食品质量与安全',
          '成龙校区', 1, '6240', None, 25],
         fill_green, '缺漏：跨页丢失(页24-25)'),
        (['理科', '5120', '四川师范大学', '成都市锦江区静安路5号', '1F', '教育技术学',
          '成龙校区', 2, '5760', None, 25],
         fill_green, '缺漏：跨页丢失(页24-25)'),
        (['理科', '5120', '四川师范大学', '成都市锦江区静安路5号', '1L', '地理信息科学',
          '成龙校区', 2, '6240', None, 25],
         fill_green, '缺漏：跨页丢失(页24-25)'),
        (['理科', '5120', '四川师范大学', '成都市锦江区静安路5号', '1M', '自然地理与资源环境',
          '成龙校区', 1, '6240', None, 25],
         fill_green, '缺漏：跨页丢失(页24-25)'),
        (['理科', '5120', '四川师范大学', '成都市锦江区静安路5号', '1S', '安全工程',
          '成龙校区', 1, '6240', None, 25],
         fill_green, '缺漏：跨页丢失(页24-25)'),
        (['理科', '5120', '四川师范大学', '成都市锦江区静安路5号', '1T', '工程造价',
          '成龙校区', 2, '5760', None, 25],
         fill_green, '缺漏：跨页丢失(页24-25)'),
        (['理科', '5120', '四川师范大学', '成都市锦江区静安路5号', '1V', '市场营销',
          '成龙校区', 1, '5760', None, 25],
         fill_green, '缺漏：跨页丢失(页24-25)'),
    ]),
    # After last OCR row (index 292): append missing records from page 25
    (292, [
        (['理科', '0373', '北方民族大学', '宁夏银川市西夏区文昌北街204号', '20', '土木类',
          '包含专业:土木工程、道路桥梁与渡河工程', 1, '4800',
          '认同并执行四川省少数民族加分项目和分值。', 25],
         fill_green, '缺漏：理科②执行本科二批控制线整节遗漏(页25)'),
        (['理科', '6503', '塔里木大学', '新疆阿拉尔市', 'BP', '机械设计制造及其自动化',
          None, 1, '3500', '认同并执行四川省少数民族加分项目和分值。', 25],
         fill_green, '缺漏：理科②执行本科二批控制线整节遗漏(页25)'),
        (['理科', '3308', '温州医科大学', '温州市瓯海区茶山高教园区', 'BW', '公共事业管理',
          '第一学年在滨海校区就读，第二学年起将根据学校规划确定。', 1, '5300',
          '认同并执行四川省少数民族加分项目和分值。', 25],
         fill_green, '缺漏：理科②执行本科二批控制线整节遗漏(页25)'),
        (['理科', '4436', '佛山大学', '广东省佛山市南海区狮山镇广云路33号', 'H3', '园艺',
          '仙溪校区；不招色盲色弱。', 1, '4568',
          '认同并执行四川省少数民族加分项目和分值。', 25],
         fill_green, '缺漏：理科②执行本科二批控制线整节遗漏(页25)'),
    ]),
]

# Sort by index descending to insert from back to front
missing_records.sort(key=lambda x: x[0], reverse=True)

for after_idx, records in missing_records:
    insert_pos = after_idx + 1
    for rec in reversed(records):
        output_data.insert(insert_pos, rec)

# Write all output data to xlsx
for i, (data, fill, corr) in enumerate(output_data):
    row_num = i + 2
    for col, val in enumerate(data, 1):
        cell = ws.cell(row=row_num, column=col, value=val)
        cell.font = font_normal
        if fill:
            cell.fill = fill
    if corr:
        cell = ws.cell(row=row_num, column=12, value=corr)
        cell.font = font_normal
        if fill:
            cell.fill = fill

wb.save(out_path)
print(f"Output saved: {out_path}")
print(f"Total rows (excl header): {len(output_data)}")

# Count stats
n_corrected = sum(1 for _, f, _ in output_data if f == fill_yellow)
n_missing = sum(1 for _, f, _ in output_data if f == fill_green)
n_total = len(output_data)
print(f"Stats: {n_total} total, {n_corrected} corrected (yellow), {n_missing} added (green), 0 excess (red)")
