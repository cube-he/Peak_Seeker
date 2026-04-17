# -*- coding: utf-8 -*-
"""P1.1 数据基线冻结：扫描源目录产出 manifest.

Usage:
    python -m scripts.data_integration.p1_baseline

Outputs:
    docs/superpowers/specs/2026-04-17-data-integration-master/baselines/YYYY-MM-DD-baseline.json
"""
from __future__ import annotations

import hashlib
import json
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterator, Optional

sys.stdout.reconfigure(encoding="utf-8")

REPO_ROOT = Path(__file__).resolve().parents[2]

TARGET_DIRS = [
    REPO_ROOT / "data" / "03_专家版主表" / "output",
    REPO_ROOT / "data" / "01_核心录取数据",
    REPO_ROOT / "data" / "13_征集志愿" / "普通高考",
]

BASELINE_DIR = (
    REPO_ROOT
    / "docs" / "superpowers" / "specs" / "2026-04-17-data-integration-master" / "baselines"
)


def scan_file(path: Path) -> dict:
    """Compute metadata for a single file."""
    path = Path(path)
    stat = path.stat()
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return {
        "path": str(path).replace("\\", "/"),
        "size_bytes": stat.st_size,
        "sha256": h.hexdigest(),
        "mtime": datetime.fromtimestamp(stat.st_mtime, tz=timezone.utc).isoformat(),
        "records": count_records(path),
    }


def count_records(path: Path) -> Optional[int]:
    """Return record count for known formats, else None."""
    path = Path(path)
    suffix = path.suffix.lower()
    try:
        if suffix == ".json":
            data = json.loads(path.read_text(encoding="utf-8"))
            if isinstance(data, list):
                return len(data)
            if isinstance(data, dict):
                return 1
            return None
        if suffix in {".xlsx", ".xls"}:
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True, data_only=True)
            total = sum(max(0, ws.max_row - 1) for ws in wb.worksheets)
            wb.close()
            return total
        if suffix == ".csv":
            with open(path, "r", encoding="utf-8", errors="ignore") as f:
                return max(0, sum(1 for _ in f) - 1)
    except Exception:
        return None
    return None


def scan_directory(root: Path) -> Iterator[dict]:
    """Yield metadata for every file under root (recursive).

    Skips Office lock files (~$...) and any file that raises PermissionError.
    """
    root = Path(root)
    if not root.exists():
        return
    for p in sorted(root.rglob("*")):
        if p.is_file():
            # Skip Office temporary lock files — they are not real data files
            if p.name.startswith("~$"):
                continue
            try:
                yield scan_file(p)
            except PermissionError:
                # File locked by another process; record as unavailable
                yield {
                    "path": str(p).replace("\\", "/"),
                    "size_bytes": 0,
                    "sha256": None,
                    "mtime": None,
                    "records": None,
                    "error": "PermissionError",
                }


def build_manifest() -> dict:
    manifest: dict = {
        "baseline_date": datetime.now(tz=timezone.utc).strftime("%Y-%m-%d"),
        "generated_at": datetime.now(tz=timezone.utc).isoformat(),
        "repo_root": str(REPO_ROOT).replace("\\", "/"),
        "sources": {},
    }
    for d in TARGET_DIRS:
        key = str(d.relative_to(REPO_ROOT)).replace("\\", "/")
        files = list(scan_directory(d))
        manifest["sources"][key] = {
            "exists": d.exists(),
            "file_count": len(files),
            "total_size_bytes": sum(f["size_bytes"] for f in files),
            "total_records": sum((f["records"] or 0) for f in files),
            "files": files,
        }
    return manifest


def main() -> int:
    manifest = build_manifest()
    BASELINE_DIR.mkdir(parents=True, exist_ok=True)
    out = BASELINE_DIR / f"{manifest['baseline_date']}-baseline.json"
    out.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Baseline written: {out}")
    for src_key, info in manifest["sources"].items():
        print(f"  {src_key}: {info['file_count']} files, {info['total_records']} records")
    return 0


if __name__ == "__main__":
    sys.exit(main())
