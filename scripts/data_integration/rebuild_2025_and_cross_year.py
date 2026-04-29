# -*- coding: utf-8 -*-
"""
两步脚本：
  Step 1: 重新合并 2025 年份分表 → 征集志愿_2025_合并.xlsx (30列)
  Step 2: 以 2025 为基底跨年整合 → 征集志愿_跨年整合_2025基底.xlsx (两Sheet)

院校代码：4位字符串（前导零保留），匹配时统一去前导零比较。
"""
from __future__ import annotations

import glob
import sys
import io
import datetime
from pathlib import Path
from collections import defaultdict

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

BASE = Path(r"C:\Users\Administrator\Documents\VolunteerHelper\data\13_征集志愿\普通高考")

# ============================================================================
# STEP 1: 重新合并 2025 年份分表
# ============================================================================

print("=" * 70)
print("STEP 1: 重新合并 2025 年份分表")
print("=" * 70)

# 统一30列输出结构
OUTPUT_COLS_2025 = [
    "年份", "科目", "录取批次", "招生类型", "降分政策",
    "院校代码", "院校名称", "办学性质", "院校地址", "院校备注",
    "调档线", "专业组代码", "再选科目要求", "专业代码", "专业名称",
    "专业备注", "收费标准",
    "第1次专业组计划数", "第1次专业计划数", "第1次来源网页",
    "第2次专业组计划数", "第2次专业计划数", "第2次来源网页",
    "第3次专业组计划数", "第3次专业计划数", "第3次来源网页",
    "页码", "征集次数", "批次ID", "校正备注",
]

MERGE_KEY_COLS = ["科目", "录取批次", "招生类型", "院校代码", "专业组代码", "专业代码"]

SEMICOLON_MERGE_COLS = ["征集次数", "页码", "批次ID", "校正备注"]

NAME_MAP = {"科类": "科目", "招生类别": "招生类型"}


def find_source_files() -> list[Path]:
    """找到所有2025年已校验xlsx，排除旧版3335/3336和原始备份。"""
    pattern = str(BASE / "**" / "*2025*已校验.xlsx")
    files = sorted(glob.glob(pattern, recursive=True))
    files = [
        Path(f) for f in files
        if "3335" not in f and "3336" not in f and "原始备份" not in f
    ]
    return files


def extract_batch_id(filepath: Path) -> str:
    """从目录名提取批次ID（首段数字）。"""
    dirname = filepath.parent.name
    return dirname.split("_")[0]


def ensure_4digit_school_code(val):
    """院校代码保持4位字符串，前导零保留。"""
    if pd.isna(val):
        return val
    s = str(val).strip()
    if s.endswith(".0"):
        s = s[:-2]
    # 纯数字则补足4位
    if s.isdigit():
        return s.zfill(4)
    return s


def normalize_columns_2025(df: pd.DataFrame, batch_id: str) -> pd.DataFrame:
    """映射各文件的不同列结构到统一30列。"""
    df = df.rename(columns=NAME_MAP)

    if "年份" not in df.columns:
        df["年份"] = 2025

    if "批次ID" not in df.columns:
        df["批次ID"] = batch_id

    # 来源网页 → 按征集次数分列
    if "来源网页" in df.columns and "征集次数" in df.columns:
        for idx_row in df.index:
            round_num = df.at[idx_row, "征集次数"]
            url = df.at[idx_row, "来源网页"]
            if pd.notna(round_num) and pd.notna(url):
                try:
                    rn = int(float(round_num))
                    col_name = f"第{rn}次来源网页"
                    if col_name in OUTPUT_COLS_2025:
                        df.at[idx_row, col_name] = url
                except (ValueError, TypeError):
                    pass
        df = df.drop(columns=["来源网页"], errors="ignore")

    # 专业组计划数 → 按征集次数映射到 "第N次专业组计划数"
    if "专业组计划数" in df.columns and "征集次数" in df.columns:
        for idx_row in df.index:
            round_num = df.at[idx_row, "征集次数"]
            plan = df.at[idx_row, "专业组计划数"]
            if pd.notna(round_num) and pd.notna(plan):
                try:
                    rn = int(float(round_num))
                    col_name = f"第{rn}次专业组计划数"
                    if col_name in OUTPUT_COLS_2025:
                        df.at[idx_row, col_name] = plan
                except (ValueError, TypeError):
                    pass
        df = df.drop(columns=["专业组计划数"], errors="ignore")

    # 专业计划数 → 按征集次数映射到 "第N次专业计划数"
    if "专业计划数" in df.columns and "征集次数" in df.columns:
        for idx_row in df.index:
            round_num = df.at[idx_row, "征集次数"]
            plan = df.at[idx_row, "专业计划数"]
            if pd.notna(round_num) and pd.notna(plan):
                try:
                    rn = int(float(round_num))
                    col_name = f"第{rn}次专业计划数"
                    if col_name in OUTPUT_COLS_2025:
                        df.at[idx_row, col_name] = plan
                except (ValueError, TypeError):
                    pass
        df = df.drop(columns=["专业计划数"], errors="ignore")

    # 院校代码保持4位字符串
    if "院校代码" in df.columns:
        df["院校代码"] = df["院校代码"].apply(ensure_4digit_school_code)

    # 确保所有输出列都存在
    for col in OUTPUT_COLS_2025:
        if col not in df.columns:
            df[col] = None

    return df[OUTPUT_COLS_2025].copy()


def merge_rows_2025(group: pd.DataFrame) -> pd.Series:
    """同一key的多行合并：核心数据取首次，分号列合并，来源网页按次分列。"""
    group = group.copy()

    def parse_first_round(val):
        if pd.isna(val):
            return 999
        try:
            return int(float(str(val).split(";")[0]))
        except (ValueError, TypeError):
            return 999

    group["_sort_key"] = group["征集次数"].apply(parse_first_round)
    group = group.sort_values("_sort_key")

    result = group.iloc[0].copy()

    for col in SEMICOLON_MERGE_COLS:
        vals = []
        for v in group[col]:
            if pd.notna(v):
                s = str(v).strip()
                if s and s not in vals:
                    vals.append(s)
        result[col] = ";".join(vals) if vals else None

    # 来源网页和计划数列合并（取各行非空值）
    for i in range(1, 4):
        for prefix in [f"第{i}次来源网页", f"第{i}次专业组计划数", f"第{i}次专业计划数"]:
            vals = group[prefix].dropna().unique()
            if len(vals) > 0:
                result[prefix] = vals[0]

    result = result.drop("_sort_key", errors="ignore")
    return result


# 1.1 扫描源文件
files = find_source_files()
print(f"\n找到 {len(files)} 个源文件:")
for f in files:
    bid = extract_batch_id(f)
    print(f"  [{bid}] {f.name}")

# 1.2 读取并归一化
all_dfs = []
total_input_rows = 0
for f in files:
    bid = extract_batch_id(f)
    df = pd.read_excel(f, dtype=str)
    n = len(df)
    total_input_rows += n
    print(f"  [{bid}] 读取 {n} 行, 列数={len(df.columns)}")
    df_norm = normalize_columns_2025(df, bid)
    all_dfs.append(df_norm)

# 1.3 拼接 + 去重合并
combined = pd.concat(all_dfs, ignore_index=True)
print(f"\n合并前总行数: {total_input_rows}")
print(f"拼接后总行数: {len(combined)}")


def make_merge_key(row):
    parts = []
    for col in MERGE_KEY_COLS:
        v = row[col]
        parts.append(str(v).strip() if pd.notna(v) else "")
    return "|".join(parts)


combined["_merge_key"] = combined.apply(make_merge_key, axis=1)

dup_counts = combined["_merge_key"].value_counts()
n_dup_keys = (dup_counts > 1).sum()
n_dup_rows = dup_counts[dup_counts > 1].sum()
print(f"重复key数: {n_dup_keys} (涉及 {n_dup_rows} 行)")

merged_rows_list = []
for key, group in combined.groupby("_merge_key", sort=False):
    if len(group) == 1:
        merged_rows_list.append(group.iloc[0].drop("_merge_key"))
    else:
        merged_rows_list.append(merge_rows_2025(group).drop("_merge_key"))

result_2025 = pd.DataFrame(merged_rows_list, columns=OUTPUT_COLS_2025)
print(f"合并后总行数: {len(result_2025)}")
print(f"去重减少: {total_input_rows - len(result_2025)} 行")

# 1.4 再次确保院校代码为4位
result_2025["院校代码"] = result_2025["院校代码"].apply(ensure_4digit_school_code)

# 1.5 输出
OUTPUT_2025 = BASE / "征集志愿_2025_合并.xlsx"
print(f"\n输出到: {OUTPUT_2025}")
result_2025.to_excel(OUTPUT_2025, index=False, sheet_name="征集志愿", engine="openpyxl")

# 格式化
wb = load_workbook(OUTPUT_2025)
ws = wb.active
font_normal = Font(name="微软雅黑", size=10)
font_bold = Font(name="微软雅黑", size=10, bold=True)

for cell in ws[1]:
    cell.font = font_bold
    cell.alignment = Alignment(horizontal="center", vertical="center")

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
    for cell in row:
        cell.font = font_normal

ws.freeze_panes = "A2"
wb.save(OUTPUT_2025)
print("格式化完成: 微软雅黑10pt, 首行冻结+加粗")

# 1.6 统计
print("\n--- 2025 合并统计 ---")
print(f"源文件数: {len(files)}")
print(f"合并前: {total_input_rows} 行")
print(f"合并后: {len(result_2025)} 行")
print(f"去重减少: {total_input_rows - len(result_2025)} 行")
if "批次ID" in result_2025.columns:
    bid_stats = result_2025["批次ID"].value_counts()
    print(f"\n按批次ID分布:")
    for bid, cnt in sorted(bid_stats.items(), key=lambda x: str(x[0])):
        print(f"  {bid}: {cnt} 行")
if "征集次数" in result_2025.columns:
    round_stats = result_2025["征集次数"].value_counts()
    print(f"\n征集次数分布:")
    for rnd, cnt in sorted(round_stats.items(), key=lambda x: str(x[0])):
        print(f"  {rnd}: {cnt} 行")


# ============================================================================
# STEP 2: 以 2025 为基底跨年整合
# ============================================================================

print("\n\n" + "=" * 70)
print("STEP 2: 以 2025 为基底跨年整合")
print("=" * 70)

# 2.1 读取数据
print("\nLoading source files...")
df25 = result_2025.copy()  # 直接用刚生成的
df24 = pd.read_excel(BASE / "征集志愿_2024_合并.xlsx", dtype=str)
df23 = pd.read_excel(BASE / "征集志愿_2023_合并.xlsx", dtype=str)
mapping_batch = pd.read_excel(BASE / "批次类别跨年映射表.xlsx", sheet_name="批次映射")
mapping_subject = pd.read_excel(BASE / "批次类别跨年映射表.xlsx", sheet_name="科目映射")

# 确保 2023/2024 院校代码也是4位
df24["院校代码"] = df24["院校代码"].apply(ensure_4digit_school_code)
df23["院校代码"] = df23["院校代码"].apply(ensure_4digit_school_code)

print(f"  2025: {len(df25)} rows, 2024: {len(df24)} rows, 2023: {len(df23)} rows")

# 2.2 科目映射
SUBJECT_MAP = {}
for _, row in mapping_subject.iterrows():
    SUBJECT_MAP[str(row["2023/2024科目"]).strip()] = str(row["2025科目"]).strip()


def normalize_subject(s):
    if pd.isna(s):
        return s
    return SUBJECT_MAP.get(str(s).strip(), str(s).strip())


# 2.3 批次映射
def build_batch_lookup(mapping_df, year):
    col_batch = f"{year}录取批次"
    col_type = f"{year}招生类型"
    lookup = {}
    for _, row in mapping_df.iterrows():
        batch = row[col_batch]
        btype = row[col_type]
        code = row["统一批次代码"]
        if pd.notna(batch) and pd.notna(btype) and pd.notna(code):
            key = (str(batch).strip(), str(btype).strip())
            if key not in lookup:
                lookup[key] = str(code).strip()
    return lookup


batch_lookup_2025 = build_batch_lookup(mapping_batch, 2025)
batch_lookup_2024 = build_batch_lookup(mapping_batch, 2024)
batch_lookup_2023 = build_batch_lookup(mapping_batch, 2023)

# 补充映射
_supp_2025 = [
    ("本科批A段(国家专项)", "国家专项计划征集志愿", "BK_A_GJ"),
    ("本科提前批B段", "其他", "BK_TQ_A"),
    ("本科批B段", "本科层次职业教育人才培养改革试点", "BK_B"),
    ("本科批B段", "非西藏生源定向西藏就业", "BK_B"),
]
for b, t, code in _supp_2025:
    batch_lookup_2025.setdefault((b, t), code)

_supp_2024 = [
    ("本科一批", "一类模式预科", "BK_1LM"),
    ("本科提前批", "司法", "BK_TQ_A"),
    ("本科提前批", "国家专项", "BK_A_GJ"),
]
for b, t, code in _supp_2024:
    batch_lookup_2024.setdefault((b, t), code)

_supp_2023 = [
    ("高水平运动队", "高水平运动队", "BK_B"),
]
for b, t, code in _supp_2023:
    batch_lookup_2023.setdefault((b, t), code)

# 引号变体补丁
_keyword_to_code = {
    "少数民族语言授课为主": {
        "高职": "ZK_1LM", "专科": "ZK_1LM",
        "预科": "BK_YK", "本科批(预科)": "BK_YK",
    },
    "加授少数民族语文": {
        "高职": "ZK_2LM", "专科": "ZK_2LM",
    },
}


def _patch_quote_variants(lookup, df, year_label):
    combos = df.groupby(["录取批次", "招生类型"]).size().reset_index(name="cnt")
    patched = 0
    for _, r in combos.iterrows():
        batch = str(r["录取批次"]).strip() if pd.notna(r["录取批次"]) else ""
        btype = str(r["招生类型"]).strip() if pd.notna(r["招生类型"]) else ""
        key = (batch, btype)
        if key in lookup:
            continue
        for keyword, batch_map in _keyword_to_code.items():
            if keyword in btype:
                for batch_kw, code in batch_map.items():
                    if batch_kw in batch:
                        lookup[key] = code
                        patched += 1
                        break
                break
    if patched:
        print(f"  Patched {patched} quote-variant entries for {year_label}")


_patch_quote_variants(batch_lookup_2025, df25, "2025")
_patch_quote_variants(batch_lookup_2024, df24, "2024")
_patch_quote_variants(batch_lookup_2023, df23, "2023")

print(f"  Batch lookup sizes: 2025={len(batch_lookup_2025)}, 2024={len(batch_lookup_2024)}, 2023={len(batch_lookup_2023)}")


# 2.4 标准化辅助函数
def strip_leading_zeros(code):
    """去前导零，用于匹配比较。"""
    if pd.isna(code):
        return ""
    s = str(code).strip()
    if s.endswith(".0"):
        s = s[:-2]
    stripped = s.lstrip("0")
    return stripped if stripped else "0"


def normalize_major_code(code):
    if pd.isna(code):
        return ""
    s = str(code).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s


def normalize_major_name(name):
    if pd.isna(name):
        return ""
    return str(name).strip()


def get_batch_code(batch, btype, lookup):
    if pd.isna(batch) or pd.isna(btype):
        return None
    key = (str(batch).strip(), str(btype).strip())
    return lookup.get(key)


# 2.5 添加内部匹配key列
def prepare_df(df, batch_lookup, year_label):
    df = df.copy()
    df["_subject"] = df["科目"].apply(normalize_subject)
    df["_school"] = df["院校代码"].apply(strip_leading_zeros)
    df["_major_code"] = df["专业代码"].apply(normalize_major_code)
    df["_major_name"] = df["专业名称"].apply(normalize_major_name)
    df["_batch_code"] = df.apply(
        lambda r: get_batch_code(r["录取批次"], r["招生类型"], batch_lookup), axis=1
    )
    df["_matched"] = False
    return df


print("\nPreparing data...")
df25p = prepare_df(df25, batch_lookup_2025, "2025")
df24p = prepare_df(df24, batch_lookup_2024, "2024")
df23p = prepare_df(df23, batch_lookup_2023, "2023")

# 检查未映射批次
for label, dfp in [("2025", df25p), ("2024", df24p), ("2023", df23p)]:
    no_code = dfp["_batch_code"].isna().sum()
    print(f"  {label}: {no_code}/{len(dfp)} rows without batch code")
    if no_code > 0:
        missing = dfp[dfp["_batch_code"].isna()][["录取批次", "招生类型"]].drop_duplicates()
        for _, r in missing.iterrows():
            print(f"    -> ({r['录取批次']}, {r['招生类型']})")


# 2.6 建索引
def build_indexes(dfp):
    idx_l1 = {}
    idx_l2 = {}
    idx_l3 = {}
    for i, row in dfp.iterrows():
        subj = row["_subject"]
        bc = row["_batch_code"]
        sch = row["_school"]
        mc = row["_major_code"]
        mn = row["_major_name"]

        if subj and sch and mc and bc:
            idx_l1.setdefault((subj, bc, sch, mc), []).append(i)
        if subj and sch and mn and bc:
            idx_l2.setdefault((subj, bc, sch, mn), []).append(i)
        if subj and sch and mn:
            idx_l3.setdefault((subj, sch, mn), []).append(i)
    return idx_l1, idx_l2, idx_l3


print("\nBuilding indexes...")
idx24_l1, idx24_l2, idx24_l3 = build_indexes(df24p)
idx23_l1, idx23_l2, idx23_l3 = build_indexes(df23p)


# 2.7 分层匹配
def find_match(row25, dfp_hist, idx_l1, idx_l2, idx_l3):
    subj = row25["_subject"]
    bc = row25["_batch_code"]
    sch = row25["_school"]
    mc = row25["_major_code"]
    mn = row25["_major_name"]

    if subj and bc and sch and mc:
        for idx in idx_l1.get((subj, bc, sch, mc), []):
            if not dfp_hist.at[idx, "_matched"]:
                dfp_hist.at[idx, "_matched"] = True
                return idx, "L1"

    if subj and bc and sch and mn:
        for idx in idx_l2.get((subj, bc, sch, mn), []):
            if not dfp_hist.at[idx, "_matched"]:
                dfp_hist.at[idx, "_matched"] = True
                return idx, "L2"

    if subj and sch and mn:
        for idx in idx_l3.get((subj, sch, mn), []):
            if not dfp_hist.at[idx, "_matched"]:
                dfp_hist.at[idx, "_matched"] = True
                return idx, "L3"

    return None, None


print("\nMatching 2025 → 2024...")
match_2024 = []
for i, row25 in df25p.iterrows():
    hist_idx, level = find_match(row25, df24p, idx24_l1, idx24_l2, idx24_l3)
    match_2024.append((hist_idx, level))

print("Matching 2025 → 2023...")
match_2023 = []
for i, row25 in df25p.iterrows():
    hist_idx, level = find_match(row25, df23p, idx23_l1, idx23_l2, idx23_l3)
    match_2023.append((hist_idx, level))


# 2.8 构建输出 DataFrame
print("\nBuilding output...")

out = pd.DataFrame()

# 2025 原始字段
out["科目"] = df25["科目"].values
out["录取批次"] = df25["录取批次"].values
out["招生类型"] = df25["招生类型"].values
out["降分政策"] = df25["降分政策"].values
out["院校代码"] = df25["院校代码"].values  # 已是4位
out["院校名称"] = df25["院校名称"].values
out["办学性质"] = df25["办学性质"].values
out["院校地址"] = df25["院校地址"].values
out["院校备注"] = df25["院校备注"].values
out["专业组代码"] = df25["专业组代码"].values
out["再选科目要求"] = df25["再选科目要求"].values
out["专业代码"] = df25["专业代码"].values
out["专业名称"] = df25["专业名称"].values
out["专业备注"] = df25["专业备注"].values
out["收费标准"] = df25["收费标准"].values

# 2025 征集数据
out["2025_第1次组计划数"] = df25["第1次专业组计划数"].values
out["2025_第1次计划数"] = df25["第1次专业计划数"].values
out["2025_第1次来源网页"] = df25["第1次来源网页"].values
out["2025_第2次组计划数"] = df25["第2次专业组计划数"].values
out["2025_第2次计划数"] = df25["第2次专业计划数"].values
out["2025_第2次来源网页"] = df25["第2次来源网页"].values
out["2025_第3次组计划数"] = df25["第3次专业组计划数"].values
out["2025_第3次计划数"] = df25["第3次专业计划数"].values
out["2025_第3次来源网页"] = df25["第3次来源网页"].values

# 2024 历史
hist_2024_cols = {
    "2024_专业代码": "专业代码",
    "2024_录取批次": "录取批次",
    "2024_招生类型": "招生类型",
    "2024_第1次计划数": "第1次专业计划数",
    "2024_第2次计划数": "第2次专业计划数",
    "2024_第3次计划数": "第3次专业计划数",
    "2024_第4次计划数": "第4次专业计划数",
}
for out_col, src_col in hist_2024_cols.items():
    vals = []
    for hist_idx, level in match_2024:
        if hist_idx is not None:
            vals.append(df24.iloc[hist_idx].get(src_col, np.nan))
        else:
            vals.append(np.nan)
    out[out_col] = vals

# 2023 历史
hist_2023_cols = {
    "2023_专业代码": "专业代码",
    "2023_录取批次": "录取批次",
    "2023_招生类型": "招生类型",
    "2023_第1次计划数": "第1次专业计划数",
    "2023_第2次计划数": "第2次专业计划数",
    "2023_第3次计划数": "第3次专业计划数",
    "2023_第4次计划数": "第4次专业计划数",
}
for out_col, src_col in hist_2023_cols.items():
    vals = []
    for hist_idx, level in match_2023:
        if hist_idx is not None:
            vals.append(df23.iloc[hist_idx].get(src_col, np.nan))
        else:
            vals.append(np.nan)
    out[out_col] = vals

# 匹配层级 + 元信息
out["匹配层级_2024"] = [level for _, level in match_2024]
out["匹配层级_2023"] = [level for _, level in match_2023]
out["2025_页码"] = df25["页码"].values
out["2025_征集次数"] = df25["征集次数"].values
out["2025_批次ID"] = df25["批次ID"].values
out["2025_校正备注"] = df25["校正备注"].values

# 再次确保所有院校代码为4位字符串
out["院校代码"] = out["院校代码"].apply(ensure_4digit_school_code)

# 2.9 未匹配 Sheet
unmatched_2024 = df24[~df24p["_matched"]].copy()
unmatched_2024.insert(0, "来源年份", 2024)
unmatched_2023 = df23[~df23p["_matched"]].copy()
unmatched_2023.insert(0, "来源年份", 2023)

common_cols = [
    "来源年份", "科目", "录取批次", "招生类型", "降分政策",
    "院校代码", "院校名称", "办学性质", "院校地址", "院校备注",
    "专业组代码", "再选科目要求", "专业代码", "专业名称", "专业备注", "收费标准",
]
plan_cols = ["第1次专业计划数", "第2次专业计划数", "第3次专业计划数", "第4次专业计划数"]

unmatched_records = []
for _, r in unmatched_2024.iterrows():
    rec = {"来源年份": 2024}
    for c in common_cols[1:]:
        rec[c] = r.get(c, np.nan)
    for c in plan_cols:
        rec[c] = r.get(c, np.nan)
    rec["页码"] = r.get("页码", np.nan)
    rec["征集次数"] = r.get("征集次数", np.nan)
    rec["批次ID"] = r.get("批次ID", np.nan)
    unmatched_records.append(rec)

for _, r in unmatched_2023.iterrows():
    rec = {"来源年份": 2023}
    for c in common_cols[1:]:
        rec[c] = r.get(c, np.nan)
    for c in plan_cols:
        rec[c] = r.get(c, np.nan)
    rec["页码"] = r.get("页码", np.nan)
    rec["征集次数"] = r.get("征集次数", np.nan)
    rec["批次ID"] = r.get("批次ID", np.nan)
    unmatched_records.append(rec)

df_unmatched = pd.DataFrame(unmatched_records) if unmatched_records else pd.DataFrame()
# 未匹配的院校代码也保持4位
if len(df_unmatched) > 0 and "院校代码" in df_unmatched.columns:
    df_unmatched["院校代码"] = df_unmatched["院校代码"].apply(ensure_4digit_school_code)


# 2.10 统计
matched_24_count = sum(1 for idx, _ in match_2024 if idx is not None)
matched_23_count = sum(1 for idx, _ in match_2023 if idx is not None)
unmatched_24_count = len(unmatched_2024)
unmatched_23_count = len(unmatched_2023)

level_stats_24 = {}
for _, level in match_2024:
    if level:
        level_stats_24[level] = level_stats_24.get(level, 0) + 1

level_stats_23 = {}
for _, level in match_2023:
    if level:
        level_stats_23[level] = level_stats_23.get(level, 0) + 1

print("\n" + "=" * 60)
print("跨年整合统计")
print("=" * 60)
print(f"2025 基底行数: {len(df25)}")
print(f"\n2024 匹配:")
print(f"  总记录: {len(df24)}")
print(f"  已匹配: {matched_24_count} ({matched_24_count/len(df24)*100:.1f}%)")
for lv in ["L1", "L2", "L3"]:
    cnt = level_stats_24.get(lv, 0)
    print(f"    {lv}: {cnt}")
print(f"  未匹配: {unmatched_24_count} ({unmatched_24_count/len(df24)*100:.1f}%)")

print(f"\n2023 匹配:")
print(f"  总记录: {len(df23)}")
print(f"  已匹配: {matched_23_count} ({matched_23_count/len(df23)*100:.1f}%)")
for lv in ["L1", "L2", "L3"]:
    cnt = level_stats_23.get(lv, 0)
    print(f"    {lv}: {cnt}")
print(f"  未匹配: {unmatched_23_count} ({unmatched_23_count/len(df23)*100:.1f}%)")

print(f"\n输出 '跨年整合' sheet: {len(out)} rows x {len(out.columns)} cols")
print(f"输出列: {list(out.columns)}")
print(f"输出 '未匹配_2023_2024' sheet: {len(df_unmatched)} rows")


# 2.11 写入 Excel
OUTPUT_CROSS = BASE / "征集志愿_跨年整合_2025基底.xlsx"
print(f"\nWriting to {OUTPUT_CROSS}...")

with pd.ExcelWriter(OUTPUT_CROSS, engine="openpyxl") as writer:
    out.to_excel(writer, sheet_name="跨年整合", index=False)
    if len(df_unmatched) > 0:
        df_unmatched.to_excel(writer, sheet_name="未匹配_2023_2024", index=False)
    else:
        pd.DataFrame({"说明": ["所有历史记录均已匹配"]}).to_excel(
            writer, sheet_name="未匹配_2023_2024", index=False
        )

# 格式化
wb = load_workbook(OUTPUT_CROSS)
font_normal = Font(name="微软雅黑", size=10)
font_header = Font(name="微软雅黑", size=10, bold=True)

for ws_name in wb.sheetnames:
    ws = wb[ws_name]
    ws.freeze_panes = "A2"

    for row in ws.iter_rows():
        for cell in row:
            cell.font = font_normal
            cell.alignment = Alignment(vertical="center", wrap_text=False)

    for cell in ws[1]:
        cell.font = font_header

    for col_idx, col_cells in enumerate(ws.columns, 1):
        max_len = 0
        for cell in col_cells:
            if cell.value:
                val_len = len(str(cell.value))
                if val_len > max_len:
                    max_len = val_len
        width = min(max(max_len + 2, 8), 40)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

wb.save(OUTPUT_CROSS)

print(f"\n完成! 两个文件已生成:")
print(f"  1. {OUTPUT_2025}")
print(f"  2. {OUTPUT_CROSS}")
print(f"  生成时间: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
