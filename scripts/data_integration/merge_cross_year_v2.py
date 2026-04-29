"""
跨年整合 v2：以2025数据为基底，将2023/2024历史数据匹配合并。

匹配逻辑：
  L1: 标准化科目 + 统一批次代码 + 院校代码(去前导零) + 专业代码
  L2: 标准化科目 + 统一批次代码 + 院校代码(去前导零) + 专业名称
  L3: 标准化科目 + 院校代码(去前导零) + 专业名称 (忽略批次差异)

输出：征集志愿_跨年整合_2025基底.xlsx
  Sheet "跨年整合"：10462行 = 2025全量
  Sheet "未匹配_2023_2024"：历史中未匹配到2025的记录
"""

import sys
import io
import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# Fix Windows console encoding
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

BASE = Path(r"C:\Users\Administrator\Documents\VolunteerHelper\data\13_征集志愿\普通高考")

# ── 1. Load source data ──────────────────────────────────────────────────────

print("Loading source files...")
df25 = pd.read_excel(BASE / "征集志愿_2025_合并.xlsx")
df24 = pd.read_excel(BASE / "征集志愿_2024_合并.xlsx")
df23 = pd.read_excel(BASE / "征集志愿_2023_合并.xlsx")
mapping_batch = pd.read_excel(BASE / "批次类别跨年映射表.xlsx", sheet_name="批次映射")
mapping_subject = pd.read_excel(BASE / "批次类别跨年映射表.xlsx", sheet_name="科目映射")

print(f"  2025: {len(df25)} rows, 2024: {len(df24)} rows, 2023: {len(df23)} rows")
print(f"  Batch mapping: {len(mapping_batch)} rows")

# ── 2. Build lookup dicts from mapping table ─────────────────────────────────

# Subject normalization: 文科→历史, 理科→物理
SUBJECT_MAP = {}
for _, row in mapping_subject.iterrows():
    SUBJECT_MAP[row["2023/2024科目"]] = row["2025科目"]


def normalize_subject(s):
    """Normalize subject to 2025 standard (历史/物理)."""
    if pd.isna(s):
        return s
    return SUBJECT_MAP.get(str(s).strip(), str(s).strip())


# Build batch code lookup: (year_prefix + 录取批次 + 招生类型) → 统一批次代码
# For each year, create a dict: (录取批次, 招生类型) → 统一批次代码
def build_batch_lookup(mapping_df, year):
    """Build (录取批次, 招生类型) → 统一批次代码 for a given year."""
    col_batch = f"{year}录取批次"
    col_type = f"{year}招生类型"
    lookup = {}
    for _, row in mapping_df.iterrows():
        batch = row[col_batch]
        btype = row[col_type]
        code = row["统一批次代码"]
        if pd.notna(batch) and pd.notna(btype):
            key = (str(batch).strip(), str(btype).strip())
            # Don't overwrite if already set (first match wins)
            if key not in lookup:
                lookup[key] = code
    return lookup


batch_lookup_2025 = build_batch_lookup(mapping_batch, 2025)
batch_lookup_2024 = build_batch_lookup(mapping_batch, 2024)
batch_lookup_2023 = build_batch_lookup(mapping_batch, 2023)

# Supplemental mappings for combos not covered in the mapping table
# These are minor variants / edge cases discovered from the actual data
_supp_2025 = [
    ("本科批A段(国家专项)", "国家专项计划征集志愿", "BK_A_GJ"),
    ("本科提前批B段", "其他", "BK_TQ_A"),
    ("本科批B段", "本科层次职业教育人才培养改革试点", "BK_B"),
    ("本科批B段", "非西藏生源定向西藏就业", "BK_B"),
]
for b, t, code in _supp_2025:
    batch_lookup_2025[(b, t)] = code

_supp_2024 = [
    ("本科一批", "一类模式预科", "BK_1LM"),
    ("本科提前批", "司法", "BK_TQ_A"),
    ("本科提前批", "国家专项", "BK_A_GJ"),
]
for b, t, code in _supp_2024:
    batch_lookup_2024[(b, t)] = code

_supp_2023 = [
    ("高水平运动队", "高水平运动队", "BK_B"),
]
for b, t, code in _supp_2023:
    batch_lookup_2023[(b, t)] = code

# Handle the quote-variant entries by scanning actual data for unresolved combos
# and matching them to known codes based on keyword patterns
_keyword_to_code = {
    "少数民族语言授课为主": {
        "高职": "ZK_1LM",
        "专科": "ZK_1LM",
        "预科": "BK_YK",
        "本科批(预科)": "BK_YK",
    },
    "加授少数民族语文": {
        "高职": "ZK_2LM",
        "专科": "ZK_2LM",
    },
}


def _patch_quote_variants(lookup, df, year_label):
    """Add mappings for entries with mixed-quote 少数民族 types."""
    combos = df.groupby(["录取批次", "招生类型"]).size().reset_index(name="cnt")
    patched = 0
    for _, r in combos.iterrows():
        batch = str(r["录取批次"]).strip() if pd.notna(r["录取批次"]) else ""
        btype = str(r["招生类型"]).strip() if pd.notna(r["招生类型"]) else ""
        key = (batch, btype)
        if key in lookup:
            continue
        for keyword, batch_map in _keyword_to_code.items():
            if keyword in btype:
                for batch_kw, code in batch_map.items():
                    if batch_kw in batch:
                        lookup[key] = code
                        patched += 1
                        break
                break
    if patched:
        print(f"  Patched {patched} quote-variant entries for {year_label}")

_patch_quote_variants(batch_lookup_2025, df25, "2025")
_patch_quote_variants(batch_lookup_2024, df24, "2024")
_patch_quote_variants(batch_lookup_2023, df23, "2023")

print(f"  Batch lookup sizes (with supplements): 2025={len(batch_lookup_2025)}, 2024={len(batch_lookup_2024)}, 2023={len(batch_lookup_2023)}")


# ── 3. Standardization helpers ───────────────────────────────────────────────

def normalize_school_code(code):
    """Strip leading zeros for comparison."""
    if pd.isna(code):
        return ""
    s = str(code).strip()
    # Handle float-like codes (e.g., 16.0)
    if '.' in s:
        try:
            s = str(int(float(s)))
        except ValueError:
            pass
    # Strip leading zeros
    stripped = s.lstrip('0')
    return stripped if stripped else '0'


def normalize_major_code(code):
    """Normalize major code to string, stripping whitespace."""
    if pd.isna(code):
        return ""
    s = str(code).strip()
    # Handle float-like codes
    if '.' in s:
        try:
            s = str(int(float(s)))
        except ValueError:
            pass
    return s


def normalize_major_name(name):
    """Normalize major name for fuzzy matching."""
    if pd.isna(name):
        return ""
    return str(name).strip()


def get_batch_code(batch, btype, lookup):
    """Get unified batch code from lookup."""
    if pd.isna(batch) or pd.isna(btype):
        return None
    key = (str(batch).strip(), str(btype).strip())
    return lookup.get(key)


# ── 4. Add internal keys to each dataframe ───────────────────────────────────

def prepare_df(df, batch_lookup, year_label):
    """Add normalized columns for matching (internal only)."""
    df = df.copy()
    df["_subject"] = df["科目"].apply(normalize_subject)
    df["_school"] = df["院校代码"].apply(normalize_school_code)
    df["_major_code"] = df["专业代码"].apply(normalize_major_code)
    df["_major_name"] = df["专业名称"].apply(normalize_major_name)
    df["_batch_code"] = df.apply(
        lambda r: get_batch_code(r["录取批次"], r["招生类型"], batch_lookup), axis=1
    )
    df["_matched"] = False
    df["_idx"] = df.index  # preserve original index
    return df


print("\nPreparing data...")
df25p = prepare_df(df25, batch_lookup_2025, "2025")
df24p = prepare_df(df24, batch_lookup_2024, "2024")
df23p = prepare_df(df23, batch_lookup_2023, "2023")

# Check batch code coverage
for label, dfp in [("2025", df25p), ("2024", df24p), ("2023", df23p)]:
    no_code = dfp["_batch_code"].isna().sum()
    print(f"  {label}: {no_code}/{len(dfp)} rows without batch code")
    if no_code > 0:
        missing = dfp[dfp["_batch_code"].isna()][["录取批次", "招生类型"]].drop_duplicates()
        for _, r in missing.iterrows():
            print(f"    -> ({r['录取批次']}, {r['招生类型']})")


# ── 5. Build indexes for historical data ─────────────────────────────────────

def build_indexes(dfp):
    """Build L1/L2/L3 indexes as dicts of key → list of row indices."""
    idx_l1 = {}  # (subject, batch_code, school, major_code) → [indices]
    idx_l2 = {}  # (subject, batch_code, school, major_name) → [indices]
    idx_l3 = {}  # (subject, school, major_name) → [indices]

    for i, row in dfp.iterrows():
        subj = row["_subject"]
        bc = row["_batch_code"]
        sch = row["_school"]
        mc = row["_major_code"]
        mn = row["_major_name"]

        if subj and sch and mc and bc:
            key1 = (subj, bc, sch, mc)
            idx_l1.setdefault(key1, []).append(i)

        if subj and sch and mn and bc:
            key2 = (subj, bc, sch, mn)
            idx_l2.setdefault(key2, []).append(i)

        if subj and sch and mn:
            key3 = (subj, sch, mn)
            idx_l3.setdefault(key3, []).append(i)

    return idx_l1, idx_l2, idx_l3


print("\nBuilding indexes...")
idx24_l1, idx24_l2, idx24_l3 = build_indexes(df24p)
idx23_l1, idx23_l2, idx23_l3 = build_indexes(df23p)


# ── 6. Match 2025 rows against historical data ──────────────────────────────

def find_match(row25, dfp_hist, idx_l1, idx_l2, idx_l3):
    """Try L1 → L2 → L3 matching. Return (matched_row_idx, level) or (None, None)."""
    subj = row25["_subject"]
    bc = row25["_batch_code"]
    sch = row25["_school"]
    mc = row25["_major_code"]
    mn = row25["_major_name"]

    # L1: subject + batch_code + school + major_code
    if subj and bc and sch and mc:
        key1 = (subj, bc, sch, mc)
        candidates = idx_l1.get(key1, [])
        for idx in candidates:
            if not dfp_hist.at[idx, "_matched"]:
                dfp_hist.at[idx, "_matched"] = True
                return idx, "L1"

    # L2: subject + batch_code + school + major_name
    if subj and bc and sch and mn:
        key2 = (subj, bc, sch, mn)
        candidates = idx_l2.get(key2, [])
        for idx in candidates:
            if not dfp_hist.at[idx, "_matched"]:
                dfp_hist.at[idx, "_matched"] = True
                return idx, "L2"

    # L3: subject + school + major_name (ignore batch)
    if subj and sch and mn:
        key3 = (subj, sch, mn)
        candidates = idx_l3.get(key3, [])
        for idx in candidates:
            if not dfp_hist.at[idx, "_matched"]:
                dfp_hist.at[idx, "_matched"] = True
                return idx, "L3"

    return None, None


print("\nMatching 2025 → 2024...")
match_2024 = []  # list of (hist_idx, level) per 2025 row
for i, row25 in df25p.iterrows():
    hist_idx, level = find_match(row25, df24p, idx24_l1, idx24_l2, idx24_l3)
    match_2024.append((hist_idx, level))

print("Matching 2025 → 2023...")
match_2023 = []
for i, row25 in df25p.iterrows():
    hist_idx, level = find_match(row25, df23p, idx23_l1, idx23_l2, idx23_l3)
    match_2023.append((hist_idx, level))


# ── 7. Build output dataframe ────────────────────────────────────────────────

print("\nBuilding output...")

# 2025 original fields (keep original column names, rename for output)
out = pd.DataFrame()

# === 2025 original fields ===
out["科目"] = df25["科目"]
out["录取批次"] = df25["录取批次"]
out["招生类型"] = df25["招生类型"]
out["降分政策"] = df25["降分政策"]
out["院校代码"] = df25["院校代码"]
out["院校名称"] = df25["院校名称"]
out["办学性质"] = df25["办学性质"]
out["院校地址"] = df25["院校地址"]
out["院校备注"] = df25["院校备注"]
out["专业组代码"] = df25["专业组代码"]
out["再选科目要求"] = df25["再选科目要求"]
out["专业代码"] = df25["专业代码"]
out["专业名称"] = df25["专业名称"]
out["专业备注"] = df25["专业备注"]
out["收费标准"] = df25["收费标准"]

# === 2025 征集数据 ===
out["2025_第1次组计划数"] = df25["第1次专业组计划数"]
out["2025_第1次计划数"] = df25["第1次专业计划数"]
out["2025_第1次来源网页"] = df25["第1次来源网页"]
out["2025_第2次组计划数"] = df25["第2次专业组计划数"]
out["2025_第2次计划数"] = df25["第2次专业计划数"]
out["2025_第2次来源网页"] = df25["第2次来源网页"]
out["2025_第3次组计划数"] = df25["第3次专业组计划数"]
out["2025_第3次计划数"] = df25["第3次专业计划数"]
out["2025_第3次来源网页"] = df25["第3次来源网页"]

# === 2024 历史 ===
hist_2024_major_code = []
hist_2024_batch = []
hist_2024_type = []
hist_2024_plan1 = []
hist_2024_plan2 = []
hist_2024_plan3 = []
hist_2024_plan4 = []

for hist_idx, level in match_2024:
    if hist_idx is not None:
        r = df24.iloc[hist_idx]
        hist_2024_major_code.append(r["专业代码"])
        hist_2024_batch.append(r["录取批次"])
        hist_2024_type.append(r["招生类型"])
        hist_2024_plan1.append(r["第1次专业计划数"])
        hist_2024_plan2.append(r["第2次专业计划数"])
        hist_2024_plan3.append(r["第3次专业计划数"])
        hist_2024_plan4.append(r["第4次专业计划数"])
    else:
        hist_2024_major_code.append(np.nan)
        hist_2024_batch.append(np.nan)
        hist_2024_type.append(np.nan)
        hist_2024_plan1.append(np.nan)
        hist_2024_plan2.append(np.nan)
        hist_2024_plan3.append(np.nan)
        hist_2024_plan4.append(np.nan)

out["2024_专业代码"] = hist_2024_major_code
out["2024_录取批次"] = hist_2024_batch
out["2024_招生类型"] = hist_2024_type
out["2024_第1次计划数"] = hist_2024_plan1
out["2024_第2次计划数"] = hist_2024_plan2
out["2024_第3次计划数"] = hist_2024_plan3
out["2024_第4次计划数"] = hist_2024_plan4

# === 2023 历史 ===
hist_2023_major_code = []
hist_2023_batch = []
hist_2023_type = []
hist_2023_plan1 = []
hist_2023_plan2 = []
hist_2023_plan3 = []
hist_2023_plan4 = []

for hist_idx, level in match_2023:
    if hist_idx is not None:
        r = df23.iloc[hist_idx]
        hist_2023_major_code.append(r["专业代码"])
        hist_2023_batch.append(r["录取批次"])
        hist_2023_type.append(r["招生类型"])
        hist_2023_plan1.append(r["第1次专业计划数"])
        hist_2023_plan2.append(r["第2次专业计划数"])
        hist_2023_plan3.append(r["第3次专业计划数"])
        hist_2023_plan4.append(r["第4次专业计划数"])
    else:
        hist_2023_major_code.append(np.nan)
        hist_2023_batch.append(np.nan)
        hist_2023_type.append(np.nan)
        hist_2023_plan1.append(np.nan)
        hist_2023_plan2.append(np.nan)
        hist_2023_plan3.append(np.nan)
        hist_2023_plan4.append(np.nan)

out["2023_专业代码"] = hist_2023_major_code
out["2023_录取批次"] = hist_2023_batch
out["2023_招生类型"] = hist_2023_type
out["2023_第1次计划数"] = hist_2023_plan1
out["2023_第2次计划数"] = hist_2023_plan2
out["2023_第3次计划数"] = hist_2023_plan3
out["2023_第4次计划数"] = hist_2023_plan4

# === 元信息 ===
out["匹配层级_2024"] = [level for _, level in match_2024]
out["匹配层级_2023"] = [level for _, level in match_2023]
out["2025_页码"] = df25["页码"]
out["2025_征集次数"] = df25["征集次数"]
out["2025_批次ID"] = df25["批次ID"]
out["2025_校正备注"] = df25["校正备注"]


# ── 8. Build unmatched sheet ─────────────────────────────────────────────────

unmatched_2024 = df24[~df24p["_matched"]].copy()
unmatched_2024.insert(0, "来源年份", 2024)
unmatched_2023 = df23[~df23p["_matched"]].copy()
unmatched_2023.insert(0, "来源年份", 2023)

# Align columns for concatenation
common_cols = ["来源年份", "科目", "录取批次", "招生类型", "降分政策",
               "院校代码", "院校名称", "办学性质", "院校地址", "院校备注",
               "专业组代码", "再选科目要求", "专业代码", "专业名称", "专业备注", "收费标准"]

# Add plan columns per year
plan_cols_24 = ["第1次专业计划数", "第2次专业计划数", "第3次专业计划数", "第4次专业计划数"]
plan_cols_23 = ["第1次专业计划数", "第2次专业计划数", "第3次专业计划数", "第4次专业计划数"]

unmatched_records = []
for _, r in unmatched_2024.iterrows():
    rec = {"来源年份": 2024}
    for c in common_cols[1:]:
        rec[c] = r.get(c, np.nan)
    for c in plan_cols_24:
        rec[c] = r.get(c, np.nan)
    rec["页码"] = r.get("页码", np.nan)
    rec["征集次数"] = r.get("征集次数", np.nan)
    rec["批次ID"] = r.get("批次ID", np.nan)
    unmatched_records.append(rec)

for _, r in unmatched_2023.iterrows():
    rec = {"来源年份": 2023}
    for c in common_cols[1:]:
        rec[c] = r.get(c, np.nan)
    for c in plan_cols_23:
        rec[c] = r.get(c, np.nan)
    rec["页码"] = r.get("页码", np.nan)
    rec["征集次数"] = r.get("征集次数", np.nan)
    rec["批次ID"] = r.get("批次ID", np.nan)
    unmatched_records.append(rec)

df_unmatched = pd.DataFrame(unmatched_records) if unmatched_records else pd.DataFrame()


# ── 9. Statistics ─────────────────────────────────────────────────────────────

matched_24_count = sum(1 for idx, _ in match_2024 if idx is not None)
matched_23_count = sum(1 for idx, _ in match_2023 if idx is not None)
unmatched_24_count = len(unmatched_2024)
unmatched_23_count = len(unmatched_2023)

# Per-level stats
level_stats_24 = {}
for _, level in match_2024:
    if level:
        level_stats_24[level] = level_stats_24.get(level, 0) + 1

level_stats_23 = {}
for _, level in match_2023:
    if level:
        level_stats_23[level] = level_stats_23.get(level, 0) + 1

print("\n" + "=" * 60)
print("跨年整合统计")
print("=" * 60)
print(f"2025 基底行数: {len(df25)}")
print(f"\n2024 匹配:")
print(f"  总记录: {len(df24)}")
print(f"  已匹配: {matched_24_count} ({matched_24_count/len(df24)*100:.1f}%)")
for lv in ["L1", "L2", "L3"]:
    cnt = level_stats_24.get(lv, 0)
    print(f"    {lv}: {cnt}")
print(f"  未匹配: {unmatched_24_count} ({unmatched_24_count/len(df24)*100:.1f}%)")

print(f"\n2023 匹配:")
print(f"  总记录: {len(df23)}")
print(f"  已匹配: {matched_23_count} ({matched_23_count/len(df23)*100:.1f}%)")
for lv in ["L1", "L2", "L3"]:
    cnt = level_stats_23.get(lv, 0)
    print(f"    {lv}: {cnt}")
print(f"  未匹配: {unmatched_23_count} ({unmatched_23_count/len(df23)*100:.1f}%)")

print(f"\n输出 '跨年整合' sheet: {len(out)} rows x {len(out.columns)} cols")
print(f"输出 '未匹配_2023_2024' sheet: {len(df_unmatched)} rows")


# ── 10. Write Excel with formatting ──────────────────────────────────────────

OUTPUT = BASE / "征集志愿_跨年整合_2025基底.xlsx"
print(f"\nWriting to {OUTPUT}...")

with pd.ExcelWriter(OUTPUT, engine="openpyxl") as writer:
    out.to_excel(writer, sheet_name="跨年整合", index=False)
    if len(df_unmatched) > 0:
        df_unmatched.to_excel(writer, sheet_name="未匹配_2023_2024", index=False)
    else:
        pd.DataFrame({"说明": ["所有历史记录均已匹配"]}).to_excel(
            writer, sheet_name="未匹配_2023_2024", index=False
        )

# Apply formatting
wb = load_workbook(OUTPUT)
font_normal = Font(name="微软雅黑", size=10)
font_header = Font(name="微软雅黑", size=10, bold=True)
header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

for ws_name in wb.sheetnames:
    ws = wb[ws_name]
    # Freeze first row
    ws.freeze_panes = "A2"

    for row in ws.iter_rows():
        for cell in row:
            cell.font = font_normal
            cell.alignment = Alignment(vertical="center", wrap_text=False)

    # Bold + fill header row
    for cell in ws[1]:
        cell.font = font_header
        cell.fill = header_fill

    # Auto-width (capped)
    for col_idx, col_cells in enumerate(ws.columns, 1):
        max_len = 0
        for cell in col_cells:
            if cell.value:
                val_len = len(str(cell.value))
                if val_len > max_len:
                    max_len = val_len
        # Cap width at 40, min at 8
        width = min(max(max_len + 2, 8), 40)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

wb.save(OUTPUT)
print("Done!")
