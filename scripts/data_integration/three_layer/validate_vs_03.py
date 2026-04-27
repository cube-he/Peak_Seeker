# -*- coding: utf-8 -*-
"""Cross-validate school_registry.json against 03_专家版主表 source tables.

Extracts school-level fields from each 03 source (mixed enrollment+school),
deduplicates by school code, compares against registry values.

Column mappings verified 2026-04-27 by printing full headers + sample data.

Sources:
  A: 13-2026-四川-专家版 (header=row3, data=row4+, code=Col05)
  C: 万能版0624 Sheet1 (header=row1, data=row2+, code=Col04)
  F: 曦鸿仕总库 (header=row1, data=row2+, code=Col02)
  H: 清洗后修改 (header=row1, data=row2+, code=Col01)
"""
from __future__ import annotations

import json
from pathlib import Path
import openpyxl

from three_layer.conflict_reporter import ConflictReporter, ConflictRecord

DATA_03 = Path(__file__).resolve().parents[3] / "data" / "03_专家版主表"
OUTPUT_DIR = Path(__file__).resolve().parent.parent / "three_layer_output"


def _safe(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _classify_diff(a: str, b: str) -> str:
    if not a or not b:
        return "one_side_missing"
    try:
        diff = abs(float(a.rstrip("%")) - float(b.rstrip("%")))
        return "numeric_small" if diff <= 3 else "numeric_large"
    except (ValueError, TypeError):
        return "text_minor" if a[:3] == b[:3] else "text_major"


def _extract_schools_A() -> dict[str, dict]:
    """Source A: 13-2026-四川-专家版.
    Header: row 3. Data: row 4+. School code: Col05.

    院校级字段 (verified):
      Col46 所在省, Col47 城市, Col48 城市水平标签,
      Col49 院校标签, Col50 院校水平(=background),
      Col51 更名合并转设, Col52 隶属单位, Col53 类型,
      Col54 公私性质, Col55 本科/专科,
      Col56 保研率, Col57 院校排名,
      Col59 全校硕士专业数, Col61 全校博士专业数
    """
    path = DATA_03 / "13-2026-四川-专家版.xlsx"
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    schools = {}
    for i, row in enumerate(ws.iter_rows(min_row=4, values_only=True)):
        code = _safe(row[5]).zfill(4) if row[5] else ""
        if not code or code == "0000":
            continue
        if code in schools:
            continue
        schools[code] = {
            "province": _safe(row[46]),         # 所在省
            "city": _safe(row[47]),             # 城市
            "cityTier": _safe(row[48]),         # 城市水平标签
            "tier": _safe(row[49]),             # 院校标签
            "background": _safe(row[50]),       # 院校水平
            "evolution": _safe(row[51]),        # 更名合并转设
            "authority": _safe(row[52]),        # 隶属单位
            "type": _safe(row[53]),             # 类型
            "nature": _safe(row[54]),           # 公私性质
            "level": _safe(row[55]),            # 本科/专科
            "postgraduateRate": _safe(row[56]), # 保研率
            "composite": _safe(row[57]),        # 院校排名
            "masterPrograms": _safe(row[59]),   # 全校硕士专业数
            "doctoralPrograms": _safe(row[61]), # 全校博士专业数
        }
    wb.close()
    return schools


def _extract_schools_C() -> dict[str, dict]:
    """Source C: 万能版0624 Sheet1.
    Header: row 1. Data: row 2+. School code: Col04.

    院校级字段 (verified):
      Col04 院校代码, Col05 院校名称,
      Col06 所在省, Col07 城市, Col08 院校档次(=tier),
      Col09 院校背景(=background), Col10 变迁史(=evolution),
      Col11 隶属单位, Col14 类型, Col15 大学排名(=composite),
      Col16 硕士点数, Col17 博士点数, Col18 保研率
    """
    path = DATA_03 / "四川2025新高考专业组三维大数据正式版(机构万能版0624.xlsx"
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb["Sheet1"]
    schools = {}
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        code = _safe(row[4]).zfill(4) if row[4] else ""  # Col04 院校代码
        if not code or code == "0000" or not code.isdigit():
            continue
        if code in schools:
            continue
        schools[code] = {
            "province": _safe(row[6]),          # Col06 所在省
            "city": _safe(row[7]),              # Col07 城市
            "tier": _safe(row[8]),              # Col08 院校档次
            "background": _safe(row[9]),        # Col09 院校背景
            "evolution": _safe(row[10]),        # Col10 变迁史
            "authority": _safe(row[11]),        # Col11 隶属单位
            "type": _safe(row[14]),             # Col14 类型
            "composite": _safe(row[15]),        # Col15 大学排名
            "masterPrograms": _safe(row[16]),   # Col16 硕士点数
            "doctoralPrograms": _safe(row[17]), # Col17 博士点数
            "postgraduateRate": _safe(row[18]), # Col18 保研率
        }
    wb.close()
    return schools


def _extract_schools_F() -> dict[str, dict]:
    """Source F: 曦鸿仕总库.
    Header: row 1. Data: row 2+. School code: Col02.

    院校级字段 (verified):
      Col02 院校代码, Col03 院校名称,
      Col29 所在省, Col30 城市,
      Col31 院校标签(=tier), Col32 院校水平(=background),
      Col33 城市水平(=cityTier), Col34 本科/专科(=level),
      Col35 隶属单位, Col36 类型, Col37 公私性质(=nature),
      Col38 院校排名(=composite), Col39 硕博点数(格式:"7/40"),
      Col40 保研率
    """
    path = DATA_03 / "曦鸿仕文化2025新高考专业组模拟3月版.xlsx"
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb["总库"]
    schools = {}
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        code = _safe(row[2]).zfill(4) if row[2] else ""  # Col02
        if not code or code == "0000" or not code.isdigit():
            continue
        if code in schools:
            continue

        # 硕博点数 格式 "7/40" → 拆分
        shuobo = _safe(row[39]) if len(row) > 39 else ""
        master = ""
        doctoral = ""
        if "/" in shuobo:
            parts = shuobo.split("/")
            doctoral = parts[0].strip()  # 博士点数在前
            master = parts[1].strip()    # 硕士点数在后

        schools[code] = {
            "province": _safe(row[29]) if len(row) > 29 else "",   # Col29 所在省
            "city": _safe(row[30]) if len(row) > 30 else "",       # Col30 城市
            "tier": _safe(row[31]) if len(row) > 31 else "",       # Col31 院校标签
            "background": _safe(row[32]) if len(row) > 32 else "", # Col32 院校水平
            "cityTier": _safe(row[33]) if len(row) > 33 else "",   # Col33 城市水平
            "level": _safe(row[34]) if len(row) > 34 else "",      # Col34 本科/专科
            "authority": _safe(row[35]) if len(row) > 35 else "",  # Col35 隶属单位
            "type": _safe(row[36]) if len(row) > 36 else "",       # Col36 类型
            "nature": _safe(row[37]) if len(row) > 37 else "",     # Col37 公私性质
            "composite": _safe(row[38]) if len(row) > 38 else "",  # Col38 院校排名
            "masterPrograms": master,                               # 从Col39拆分
            "doctoralPrograms": doctoral,                           # 从Col39拆分
            "postgraduateRate": _safe(row[40]) if len(row) > 40 else "",  # Col40 保研率
        }
    wb.close()
    return schools


def _extract_schools_H() -> dict[str, dict]:
    """Source H: 清洗后修改.
    Header: row 1. Data: row 2+. School code: Col01.

    院校级字段 (verified):
      Col57 院校省份, Col58 院校城市, Col59 城市等级,
      Col60 院校类型, Col61 办学性质, Col62 隶属部门,
      Col63 院校标签(=tier), Col64 院校层级(=level),
      Col65 院校排名, Col69 保研率,
      Col70 是否双一流, Col81 硕士点数量, Col83 博士点数量
    """
    path = DATA_03 / "2026四川高考志愿_清洗后_修改.xlsx"
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    schools = {}
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        code = _safe(row[1]).zfill(4) if row[1] else ""  # Col01 院校代码
        if not code or code == "0000" or not code.isdigit():
            continue
        if code in schools:
            continue
        schools[code] = {
            "province": _safe(row[57]) if len(row) > 57 else "",    # Col57 院校省份
            "city": _safe(row[58]) if len(row) > 58 else "",        # Col58 院校城市
            "cityTier": _safe(row[59]) if len(row) > 59 else "",    # Col59 城市等级
            "type": _safe(row[60]) if len(row) > 60 else "",        # Col60 院校类型
            "nature": _safe(row[61]) if len(row) > 61 else "",      # Col61 办学性质
            "authority": _safe(row[62]) if len(row) > 62 else "",   # Col62 隶属部门
            "tier": _safe(row[63]) if len(row) > 63 else "",        # Col63 院校标签
            "level": _safe(row[64]) if len(row) > 64 else "",       # Col64 院校层级
            "composite": _safe(row[65]) if len(row) > 65 else "",   # Col65 院校排名
            "postgraduateRate": _safe(row[69]) if len(row) > 69 else "",  # Col69 保研率
            "masterPrograms": _safe(row[81]) if len(row) > 81 else "",    # Col81 硕士点数量
            "doctoralPrograms": _safe(row[83]) if len(row) > 83 else "",  # Col83 博士点数量
        }
    wb.close()
    return schools


# Registry field → nested path
_COMPARE_MAP = {
    "province": "location.province",
    "city": "location.city",
    "cityTier": "location.cityTier",
    "type": "basic.type",
    "nature": "basic.nature",
    "authority": "basic.authority",
    "level": "basic.level",
    "composite": "rankings.composite",
    "masterPrograms": "academics.masterPrograms",
    "doctoralPrograms": "academics.doctoralPrograms",
    "postgraduateRate": "academics.postgraduateRate",
    "tier": "tags.tier",
    "background": "tags.background",
    "evolution": "history.evolution",
}


def _get_nested(obj: dict, path: str):
    parts = path.split(".")
    for p in parts:
        if isinstance(obj, dict):
            obj = obj.get(p)
        else:
            return None
    return obj


def validate(source_label: str, source_schools: dict, registry: dict,
             reporter: ConflictReporter) -> dict:
    matched = 0
    diffs = 0
    for code, src in source_schools.items():
        if code not in registry["schools"]:
            continue
        matched += 1
        school = registry["schools"][code]

        for src_field, reg_path in _COMPARE_MAP.items():
            src_val = src.get(src_field, "")
            if not src_val or src_val == "/":
                continue
            reg_val = _safe(_get_nested(school, reg_path))
            if not reg_val:
                continue

            # Normalize for comparison
            sv = src_val.rstrip("%").strip()
            rv = reg_val.rstrip("%").strip()
            if sv == rv:
                continue

            diff_type = _classify_diff(sv, rv)
            reporter.add_conflict(ConflictRecord(
                school_code=code,
                school_name=school.get("name", ""),
                subject="—", batch="—", major_code="—", major_name="—",
                field_name=reg_path,
                current_value=rv,
                new_value=sv,
                source=source_label,
                match_type="exact",
                confidence=1.0,
                diff_type=diff_type,
            ))
            diffs += 1

    return {"matched": matched, "total": len(source_schools), "diffs": diffs}


def main():
    reg_path = OUTPUT_DIR / "school_registry.json"
    with open(reg_path, "r", encoding="utf-8") as f:
        registry = json.load(f)

    reporter = ConflictReporter(output_dir=OUTPUT_DIR)

    sources = [
        ("03_A_李老师", _extract_schools_A),
        ("03_C_万能版", _extract_schools_C),
        ("03_F_曦鸿仕", _extract_schools_F),
        ("03_H_清洗后", _extract_schools_H),
    ]

    for label, extractor in sources:
        print(f"Extracting {label}...")
        schools = extractor()
        print(f"  Extracted {len(schools)} unique schools")
        # Sanity check: verify a known school
        if "0001" in schools:
            pku = schools["0001"]
            print(f"  Sanity check 0001: province={pku.get('province')}, "
                  f"type={pku.get('type')}, authority={pku.get('authority')}, "
                  f"composite={pku.get('composite')}")
        stats = validate(label, schools, registry, reporter)
        print(f"  Validated: {stats['matched']}/{stats['total']} matched, {stats['diffs']} diffs")
        print()

    if reporter.count > 0:
        csv_path = reporter.write("school_registry_03_validation.csv")
        print(f"Total conflicts: {reporter.count} → {csv_path}")
        summary = reporter.summary()
        print(f"  By source: {summary['by_source']}")
        print(f"  By type: {summary['by_diff_type']}")

        # Per-field breakdown
        from collections import Counter
        import csv
        with open(csv_path, "r", encoding="utf-8-sig") as f:
            csv_rows = list(csv.DictReader(f))
        by_sf = Counter((r["source"], r["field_name"]) for r in csv_rows)
        print("\n  Top conflicts by source×field:")
        for (src, field), count in by_sf.most_common(25):
            print(f"    {src} | {field}: {count}")
    else:
        print("No conflicts.")


if __name__ == "__main__":
    main()
