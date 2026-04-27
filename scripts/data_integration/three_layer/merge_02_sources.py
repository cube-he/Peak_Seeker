# -*- coding: utf-8 -*-
"""Merge all 02_全国基础库 sources into school_registry.json.

Sources merged:
1. 大学排名_全国.xlsx → rankings (detailed rank values)
2. 招生章程结构化_全国_2025.xlsx → admissionRules
3. 院校满意度_全国_阳光高考.xlsx → satisfaction
4. 学科评估_全国.xlsx → academics.assessmentDetails
5. 院校简介_全国_阳光高考_partial.xlsx → history, links

Join keys:
- 大学排名/学科评估: 院校代码(国标) → registry ids.nationalCode
- 招生章程/满意度/简介: 阳光高考ID → registry ids.yangguangId
"""
from __future__ import annotations

import json
from pathlib import Path
import openpyxl

from three_layer.conflict_reporter import ConflictReporter, ConflictRecord

DATA_02 = Path(__file__).resolve().parents[3] / "data" / "02_全国基础库"
OUTPUT_DIR = Path(__file__).resolve().parent.parent / "three_layer_output"


def _safe_str(v) -> str | None:
    if v is None: return None
    s = str(v).strip()
    return s if s else None


def _safe_int(v) -> int | None:
    if v is None: return None
    try: return int(v)
    except (ValueError, TypeError): return None


def _safe_float(v) -> float | None:
    if v is None: return None
    try: return round(float(str(v).strip().rstrip("%")), 2)
    except (ValueError, TypeError): return None


def _read_xlsx(filename: str, key_col: str) -> tuple[list[str], dict[str, dict]]:
    """Read xlsx, return (headers, {key → row_dict})."""
    path = DATA_02 / filename
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    headers = []
    data = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(c).strip() if c else "" for c in row]
            continue
        rec = {headers[j]: row[j] for j in range(min(len(headers), len(row)))}
        key = _safe_str(rec.get(key_col))
        if key:
            data[key] = rec
    wb.close()
    return headers, data


def _build_bridge(registry: dict) -> tuple[dict, dict]:
    """Build nationalCode→code and yangguangId→code lookup maps.

    02 sources use mixed national codes:
    - 5-digit (10001 = 北京大学) — exact match
    - 6-digit (100011 = 北京大学校本部) — prefix match on first 5 digits
    So nc_map stores the 5-digit code as key.
    """
    nc_map = {}
    yg_map = {}
    for code, s in registry["schools"].items():
        ids = s.get("ids", {})
        nc = ids.get("nationalCode")
        yg = ids.get("yangguangId")
        if nc:
            nc_map[str(nc)] = code
        if yg:
            yg_map[str(yg)] = code
    return nc_map, yg_map


def _resolve_nc(nc_code: str, nc_map: dict) -> str | None:
    """Resolve a national code (5 or 6 digit) to registry code.

    Tries exact match first, then prefix match (first 5 digits).
    """
    nc = str(nc_code).strip()
    if nc in nc_map:
        return nc_map[nc]
    if len(nc) == 6 and nc[:5] in nc_map:
        return nc_map[nc[:5]]
    return None


def _set_if_empty(school: dict, path: str, value, reporter: ConflictReporter,
                  code: str, source: str):
    """Set a nested value only if currently None/empty. Report conflict if different."""
    parts = path.split(".")
    obj = school
    for p in parts[:-1]:
        obj = obj.setdefault(p, {})
    field = parts[-1]
    current = obj.get(field)

    if value is None:
        return

    if current is None or current == "" or current == 0:
        obj[field] = value
        return

    # Both have values — check if different
    if str(current).strip() != str(value).strip():
        reporter.add_conflict(ConflictRecord(
            school_code=code,
            school_name=school.get("name", ""),
            subject="—", batch="—", major_code="—", major_name="—",
            field_name=path,
            current_value=str(current),
            new_value=str(value),
            source=source,
            match_type="exact",
            confidence=1.0,
            diff_type=_classify_diff(current, value),
        ))


def _classify_diff(a, b) -> str:
    try:
        diff = abs(float(a) - float(b))
        return "numeric_small" if diff <= 3 else "numeric_large"
    except (ValueError, TypeError):
        sa, sb = str(a).strip(), str(b).strip()
        if sa[:3] == sb[:3]:
            return "text_minor"
        return "text_major"


def merge_rankings(registry: dict, nc_map: dict, reporter: ConflictReporter):
    """Merge 02/大学排名_全国.xlsx → rankings detail fields."""
    _, data = _read_xlsx("大学排名_全国.xlsx", "院校代码")
    matched = 0

    for nc_code, rec in data.items():
        reg_code = _resolve_nc(nc_code, nc_map)
        if not reg_code:
            continue
        matched += 1
        school = registry["schools"][reg_code]

        field_map = {
            "rankings.qs": ("QS排名", _safe_int),
            "rankings.usNews": ("USNews排名", _safe_int),
            "rankings.alumni": ("校友会排名", _safe_int),
            "rankings.arwu": ("软科排名", _safe_int),
            "rankings.wushulian": ("武书连排名", _safe_int),
            "rankings.moe": ("教育部排名", _safe_int),
            "rankings.overallRank": ("综合排名", _safe_int),
            "rankings.overallScore": ("综合评分", _safe_float),
            "rankings.popularity": ("热度", _safe_int),
            "links.logo": ("Logo地址", _safe_str),
            "links.banner": ("Banner地址", _safe_str),
        }

        for target, (col, conv) in field_map.items():
            val = conv(rec.get(col))
            if val == 0 and "排名" in col:
                val = None  # 0 means no ranking
            _set_if_empty(school, target, val, reporter, reg_code, "02_大学排名")

    print(f"  大学排名: matched {matched}/{len(data)}")


def merge_admission_rules(registry: dict, yg_map: dict, reporter: ConflictReporter):
    """Merge 02/招生章程结构化_全国_2025.xlsx → admissionRules."""
    _, data = _read_xlsx("招生章程结构化_全国_2025.xlsx", "阳光高考ID")
    matched = 0

    for yg_id, rec in data.items():
        reg_code = yg_map.get(yg_id)
        if not reg_code:
            continue
        matched += 1
        school = registry["schools"][reg_code]

        field_map = {
            "admissionRules.filingRatio": "调档比例",
            "admissionRules.majorAllocation": "专业分配规则",
            "admissionRules.tiebreakRule": "同分规则",
            "admissionRules.foreignLanguageReq": "外语要求",
            "admissionRules.subjectScoreReq": "单科要求",
            "admissionRules.healthRestrictions": "体检限制",
            "admissionRules.bonusPolicy": "加分政策",
            "admissionRules.tuition": "学费",
            "admissionRules.majorTransferRestrictions": "转专业限制",
            "admissionRules.adjustmentPolicy": "服从调剂",
        }

        for target, col in field_map.items():
            val = _safe_str(rec.get(col))
            _set_if_empty(school, target, val, reporter, reg_code, "02_招生章程")

    print(f"  招生章程: matched {matched}/{len(data)}")


def merge_satisfaction(registry: dict, yg_map: dict, reporter: ConflictReporter):
    """Merge 02/院校满意度_全国_阳光高考.xlsx → satisfaction."""
    _, data = _read_xlsx("院校满意度_全国_阳光高考.xlsx", "阳光高考ID")
    matched = 0

    for yg_id, rec in data.items():
        reg_code = yg_map.get(yg_id)
        if not reg_code:
            continue
        matched += 1
        school = registry["schools"][reg_code]

        # Overall
        _set_if_empty(school, "satisfaction.overall.score",
                      _safe_float(rec.get("综合满意度")), reporter, reg_code, "02_满意度")
        _set_if_empty(school, "satisfaction.overall.count",
                      _safe_int(rec.get("综合评价人数")), reporter, reg_code, "02_满意度")
        stars = [_safe_int(rec.get(f"综合{i}星")) or 0 for i in range(1, 6)]
        if any(s > 0 for s in stars):
            cur = school.get("satisfaction", {}).get("overall", {}).get("stars", [0]*5)
            if cur == [0]*5:
                school.setdefault("satisfaction", {}).setdefault("overall", {})["stars"] = stars

        # Living
        _set_if_empty(school, "satisfaction.living.score",
                      _safe_float(rec.get("生活满意度")), reporter, reg_code, "02_满意度")
        live_stars = [_safe_int(rec.get(f"生活{i}星")) or 0 for i in range(1, 6)]
        if any(s > 0 for s in live_stars):
            cur = school.get("satisfaction", {}).get("living", {}).get("stars", [0]*5)
            if cur == [0]*5:
                school.setdefault("satisfaction", {}).setdefault("living", {})["stars"] = live_stars

        # Environment
        _set_if_empty(school, "satisfaction.environment.score",
                      _safe_float(rec.get("环境满意度")), reporter, reg_code, "02_满意度")
        env_stars = [_safe_int(rec.get(f"环境{i}星")) or 0 for i in range(1, 6)]
        if any(s > 0 for s in env_stars):
            cur = school.get("satisfaction", {}).get("environment", {}).get("stars", [0]*5)
            if cur == [0]*5:
                school.setdefault("satisfaction", {}).setdefault("environment", {})["stars"] = env_stars

    print(f"  院校满意度: matched {matched}/{len(data)}")


def merge_assessment(registry: dict, nc_map: dict, reporter: ConflictReporter):
    """Merge 02/学科评估_全国.xlsx → academics.assessmentDetails.

    学科评估是多行数据(每校多个学科), 聚合成一个摘要字符串。
    """
    _, data_raw = _read_xlsx("学科评估_全国.xlsx", "院校代码")

    # 重新读取为列表（因为同一院校多行）
    path = DATA_02 / "学科评估_全国.xlsx"
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    headers = []
    by_school: dict[str, list] = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(c).strip() if c else "" for c in row]
            continue
        rec = {headers[j]: row[j] for j in range(min(len(headers), len(row)))}
        nc = _safe_str(rec.get("院校代码"))
        if nc:
            by_school.setdefault(nc, []).append(rec)
    wb.close()

    matched = 0
    for nc_code, records in by_school.items():
        reg_code = _resolve_nc(nc_code, nc_map)
        if not reg_code:
            continue
        matched += 1
        school = registry["schools"][reg_code]

        # Build assessment summary: "A+:数学,物理; A:化学; B+:生物"
        by_grade: dict[str, list] = {}
        for rec in records:
            grade = _safe_str(rec.get("评估类型名称"))
            subject = _safe_str(rec.get("名称")) or _safe_str(rec.get("专业名称"))
            if grade and subject:
                by_grade.setdefault(grade, []).append(subject)

        if by_grade:
            # Sort by grade quality
            grade_order = ["A+", "A", "A-", "B+", "B", "B-", "C+", "C", "C-"]
            parts = []
            for g in grade_order:
                if g in by_grade:
                    parts.append(f"{g}:{','.join(by_grade[g])}")
            summary = "; ".join(parts)

            academics = school.setdefault("academics", {})
            if not academics.get("assessmentDetails"):
                academics["assessmentDetails"] = summary

    print(f"  学科评估: matched {matched}/{len(by_school)} schools")


def merge_intro(registry: dict, yg_map: dict, reporter: ConflictReporter):
    """Merge 02/院校简介_全国_阳光高考_partial.xlsx → misc fields."""
    _, data = _read_xlsx("院校简介_全国_阳光高考_partial.xlsx", "阳光高考ID")
    matched = 0

    for yg_id, rec in data.items():
        reg_code = yg_map.get(yg_id)
        if not reg_code:
            continue
        matched += 1
        school = registry["schools"][reg_code]

        # Address supplement
        addr = _safe_str(rec.get("通讯地址"))
        _set_if_empty(school, "location.address", addr, reporter, reg_code, "02_简介")

    print(f"  院校简介: matched {matched}/{len(data)}")


def main():
    reg_path = OUTPUT_DIR / "school_registry.json"
    with open(reg_path, "r", encoding="utf-8") as f:
        registry = json.load(f)

    nc_map, yg_map = _build_bridge(registry)
    reporter = ConflictReporter(output_dir=OUTPUT_DIR)

    print("Merging 02 sources into school_registry.json...")
    merge_rankings(registry, nc_map, reporter)
    merge_admission_rules(registry, yg_map, reporter)
    merge_satisfaction(registry, yg_map, reporter)
    merge_assessment(registry, nc_map, reporter)
    merge_intro(registry, yg_map, reporter)

    # Write updated registry
    with open(reg_path, "w", encoding="utf-8") as f:
        json.dump(registry, f, ensure_ascii=False, indent=2)
    print(f"\nUpdated registry: {reg_path}")

    # Write conflict report
    if reporter.count > 0:
        csv_path = reporter.write("school_registry_02_merge_conflicts.csv")
        print(f"Conflicts: {reporter.count} → {csv_path}")
        summary = reporter.summary()
        print(f"  By source: {summary['by_source']}")
        print(f"  By type: {summary['by_diff_type']}")
    else:
        print("No conflicts during 02 merge.")


if __name__ == "__main__":
    main()
