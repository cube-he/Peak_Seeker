# -*- coding: utf-8 -*-
"""
合并2025年全部已校验xlsx，按征集次数去重合并。

合并key（2025有专业组）:
  科目 + 录取批次 + 招生类型 + 院校代码 + 专业组代码 + 专业代码

同一key多次征集时:
  - 征集次数: 分号合并
  - 页码: 分号合并
  - 来源网页: 按征集次数分列（第N次 -> 第N次来源网页列）
  - 批次ID: 分号合并
  - 校正备注: 分号合并
  - 核心数据取首次值

输出: data/13_征集志愿/普通高考/征集志愿_2025_合并.xlsx
"""
from __future__ import annotations

import glob
import os
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

# ==================== 配置 ====================
BASE_DIR = Path(__file__).resolve().parent.parent.parent
DATA_DIR = BASE_DIR / "data" / "13_征集志愿" / "普通高考"
OUTPUT_PATH = DATA_DIR / "征集志愿_2025_合并.xlsx"

# 统一27列输出结构
OUTPUT_COLS = [
    "年份", "科目", "录取批次", "招生类型", "降分政策",
    "院校代码", "院校名称", "办学性质", "院校地址", "院校备注",
    "调档线", "专业组代码", "再选科目要求", "专业组计划数",
    "专业代码", "专业名称", "专业备注", "专业计划数", "收费标准",
    "页码", "征集次数",
    "第1次来源网页", "第2次来源网页", "第3次来源网页", "第4次来源网页",
    "批次ID", "校正备注",
]

# 合并key列
MERGE_KEY_COLS = ["科目", "录取批次", "招生类型", "院校代码", "专业组代码", "专业代码"]

# 分号合并列
SEMICOLON_MERGE_COLS = ["征集次数", "页码", "批次ID", "校正备注"]

# 列名映射：源文件列名 -> 统一列名
NAME_MAP = {
    "科类": "科目",
    "招生类别": "招生类型",
}


def find_source_files() -> list[Path]:
    """找到所有2025年已校验xlsx，排除旧版3335/3336和原始备份。"""
    pattern = str(DATA_DIR / "**" / "*2025*已校验.xlsx")
    files = sorted(glob.glob(pattern, recursive=True))
    files = [
        Path(f) for f in files
        if "3335" not in f and "3336" not in f and "原始备份" not in f
    ]
    return files


def extract_batch_id(filepath: Path) -> str:
    """从目录名提取批次ID（首段数字）。"""
    dirname = filepath.parent.name
    return dirname.split("_")[0]


def normalize_columns(df: pd.DataFrame, batch_id: str) -> pd.DataFrame:
    """
    按列名匹配，将各文件的不同列结构映射到统一27列。
    不硬编码位置，完全靠列名。
    """
    # 重命名已知别名
    df = df.rename(columns=NAME_MAP)

    # 如果没有"年份"列，补充
    if "年份" not in df.columns:
        df["年份"] = 2025

    # 如果没有"批次ID"列，从目录名提取
    if "批次ID" not in df.columns:
        df["批次ID"] = batch_id

    # 来源网页 -> 按征集次数分列
    # 先把"来源网页"映射到对应的"第N次来源网页"
    if "来源网页" in df.columns and "征集次数" in df.columns:
        for _, row in df.iterrows():
            round_num = row.get("征集次数")
            url = row.get("来源网页")
            if pd.notna(round_num) and pd.notna(url):
                try:
                    rn = int(float(round_num))
                    col_name = f"第{rn}次来源网页"
                    if col_name in OUTPUT_COLS:
                        df.at[_, col_name] = url
                except (ValueError, TypeError):
                    pass
        df = df.drop(columns=["来源网页"], errors="ignore")

    # 确保所有输出列都存在
    for col in OUTPUT_COLS:
        if col not in df.columns:
            df[col] = None

    # 只保留输出列，按顺序排列
    return df[OUTPUT_COLS].copy()


def merge_rows(group: pd.DataFrame) -> pd.Series:
    """
    同一key的多行合并策略：
    - 核心数据取首次值（征集次数最小的那次）
    - 分号合并列用分号拼接
    - 来源网页按征集次数分到不同列
    """
    # 按征集次数排序，首次值优先
    group = group.copy()

    # 解析征集次数用于排序
    def parse_first_round(val):
        if pd.isna(val):
            return 999
        try:
            return int(float(str(val).split(";")[0]))
        except (ValueError, TypeError):
            return 999

    group["_sort_key"] = group["征集次数"].apply(parse_first_round)
    group = group.sort_values("_sort_key")

    # 核心数据取首次值
    result = group.iloc[0].copy()

    # 分号合并列
    for col in SEMICOLON_MERGE_COLS:
        vals = []
        for v in group[col]:
            if pd.notna(v):
                s = str(v).strip()
                if s and s not in vals:
                    vals.append(s)
        result[col] = ";".join(vals) if vals else None

    # 来源网页列合并（取各行非空值）
    for i in range(1, 5):
        col = f"第{i}次来源网页"
        vals = group[col].dropna().unique()
        if len(vals) > 0:
            result[col] = str(vals[0])

    # 清理临时列
    result = result.drop("_sort_key", errors="ignore")

    return result


def main():
    print("=" * 70)
    print("合并2025年征集志愿已校验文件")
    print("=" * 70)

    # 1. 扫描源文件
    files = find_source_files()
    print(f"\n找到 {len(files)} 个源文件:")
    for f in files:
        bid = extract_batch_id(f)
        print(f"  [{bid}] {f.name}")

    # 2. 读取并归一化所有文件
    all_dfs = []
    total_input_rows = 0
    for f in files:
        bid = extract_batch_id(f)
        df = pd.read_excel(f, dtype=str)
        n = len(df)
        total_input_rows += n
        print(f"  [{bid}] 读取 {n} 行, 列数={len(df.columns)}")
        df_norm = normalize_columns(df, bid)
        all_dfs.append(df_norm)

    # 3. 合并所有数据
    combined = pd.concat(all_dfs, ignore_index=True)
    print(f"\n合并前总行数: {total_input_rows}")
    print(f"拼接后总行数: {len(combined)}")

    # 4. 按合并key去重合并
    # 构建合并key（处理空值）
    def make_key(row):
        parts = []
        for col in MERGE_KEY_COLS:
            v = row[col]
            parts.append(str(v).strip() if pd.notna(v) else "")
        return "|".join(parts)

    combined["_merge_key"] = combined.apply(make_key, axis=1)

    # 统计重复
    dup_counts = combined["_merge_key"].value_counts()
    n_dup_keys = (dup_counts > 1).sum()
    n_dup_rows = dup_counts[dup_counts > 1].sum()
    print(f"重复key数: {n_dup_keys} (涉及 {n_dup_rows} 行)")

    # 分组合并
    merged_rows = []
    for key, group in combined.groupby("_merge_key", sort=False):
        if len(group) == 1:
            merged_rows.append(group.iloc[0].drop("_merge_key"))
        else:
            merged_rows.append(merge_rows(group).drop("_merge_key"))

    result = pd.DataFrame(merged_rows, columns=OUTPUT_COLS)
    print(f"合并后总行数: {len(result)}")
    print(f"去重减少: {total_input_rows - len(result)} 行")

    # 5. 输出到xlsx
    print(f"\n输出到: {OUTPUT_PATH}")
    result.to_excel(OUTPUT_PATH, index=False, sheet_name="征集志愿", engine="openpyxl")

    # 6. 格式化：微软雅黑10pt，首行冻结+加粗
    wb = load_workbook(OUTPUT_PATH)
    ws = wb.active

    # 微软雅黑 10pt 全表
    font_normal = Font(name="微软雅黑", size=10)
    font_bold = Font(name="微软雅黑", size=10, bold=True)

    # 首行加粗
    for cell in ws[1]:
        cell.font = font_bold
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 数据行普通字体
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.font = font_normal

    # 冻结首行
    ws.freeze_panes = "A2"

    wb.save(OUTPUT_PATH)
    print("格式化完成: 微软雅黑10pt, 首行冻结+加粗")

    # 7. 统计汇总
    print("\n" + "=" * 70)
    print("统计汇总")
    print("=" * 70)
    print(f"源文件数: {len(files)}")
    print(f"合并前总行数: {total_input_rows}")
    print(f"合并后总行数: {len(result)}")
    print(f"去重减少: {total_input_rows - len(result)} 行")
    print(f"重复key数: {n_dup_keys}")

    # 按批次统计
    if "批次ID" in result.columns:
        bid_stats = result["批次ID"].value_counts()
        print(f"\n按批次ID分布:")
        for bid, cnt in bid_stats.items():
            print(f"  {bid}: {cnt} 行")

    # 检查征集次数分布
    if "征集次数" in result.columns:
        round_stats = result["征集次数"].value_counts()
        print(f"\n征集次数分布:")
        for rnd, cnt in sorted(round_stats.items(), key=lambda x: str(x[0])):
            print(f"  {rnd}: {cnt} 行")


if __name__ == "__main__":
    main()
