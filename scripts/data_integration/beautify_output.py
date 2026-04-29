"""
美化 院校信息表.xlsx 和 专业招生主表.xlsx 的格式。

变更内容：
- 表头：微软雅黑 10pt 加粗，白字深蓝底，居中
- 数据行：微软雅黑 10pt，交替灰白底色
- 全部单元格加细边框
- 冻结窗格：首行 + 前2列（代码+名称）
- 自动筛选
- 列宽自适应（采样前100行，上限40字符）
- 百分比列格式化
- JSON 列（双一流专业/特色专业/地址）解析为可读文本
- 代码列（院校代码、专业代码等）强制文本格式，避免前导零丢失
"""

import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from pathlib import Path
from copy import copy

BASE_DIR = Path(__file__).resolve().parent.parent.parent / "data" / "03_专家版主表" / "output"

# ── 样式定义 ──────────────────────────────────────────────
FONT_CN = "微软雅黑"
HEADER_FONT = Font(name=FONT_CN, size=10, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

DATA_FONT = Font(name=FONT_CN, size=10)
DATA_ALIGN = Alignment(vertical="center", wrap_text=False)
DATA_ALIGN_CENTER = Alignment(horizontal="center", vertical="center")

EVEN_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
ODD_FILL = PatternFill(fill_type=None)

THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
HEADER_BORDER = Border(
    left=Side(style="thin", color="1F4E79"),
    right=Side(style="thin", color="1F4E79"),
    top=Side(style="thin", color="1F4E79"),
    bottom=Side(style="medium", color="1F4E79"),
)


def calc_col_widths(ws, max_width=40, sample_rows=100):
    """基于表头和前 sample_rows 行数据计算列宽。"""
    widths = {}
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for row_idx in range(1, min(sample_rows + 2, ws.max_row + 1)):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                # 中文字符占 2 个宽度单位
                text = str(val)[:80]
                length = sum(2 if ord(c) > 127 else 1 for c in text)
                max_len = max(max_len, length)
        # 加 2 padding，限制上限
        widths[col_idx] = min(max_len + 2, max_width)
    return widths


def parse_json_cell(val):
    """将 JSON 字符串解析为可读文本。"""
    if val is None or not isinstance(val, str):
        return val
    try:
        data = json.loads(val)
    except (json.JSONDecodeError, TypeError):
        return val

    # 双一流专业：提取学科名称列表
    if "sylSubjectsGroup" in data:
        names = []
        for group in data.get("sylSubjectsGroup", []):
            round_name = group.get("evaluations", "")
            subjects = group.get("subjects", [])
            for s in subjects:
                name = s.get("name", "")
                if name and name not in names:
                    names.append(name)
        return "；".join(names) if names else val

    # 特色专业：提取国家级 + 省级名称
    if "countries" in data or "provinces" in data:
        parts = []
        for item in data.get("countries", []):
            parts.append(item.get("name", ""))
        for item in data.get("provinces", []):
            parts.append(f"[省]{item.get('name', '')}")
        return "；".join(p for p in parts if p) if parts else val

    return val


def parse_address_cell(val):
    """将地址列的 Python repr / JSON 转为可读文本。"""
    if val is None or not isinstance(val, str):
        return val
    try:
        # 尝试 JSON
        data = json.loads(val)
    except (json.JSONDecodeError, TypeError):
        try:
            # 尝试 Python eval（来自 repr 的 list of dict）
            import ast
            data = ast.literal_eval(val)
        except Exception:
            return val

    if isinstance(data, list):
        parts = []
        for campus in data:
            name = campus.get("name", "")
            addr = campus.get("address", "")
            parts.append(f"{name}: {addr}" if name else addr)
        return " | ".join(p for p in parts if p)
    return val


def identify_center_cols(headers):
    """识别应该居中对齐的列（数值型、代码型、是否型）。"""
    center_keywords = [
        "代码", "等级", "排名", "数量", "人数", "年份", "学制",
        "比例", "是否", "热度", "评分", "满意度", "星",
        "最低分", "最高分", "平均分", "位次", "位",
        "计划", "录取", "投档", "顺序",
    ]
    center_idxs = set()
    for i, h in enumerate(headers):
        if h and any(kw in h for kw in center_keywords):
            center_idxs.add(i)
    return center_idxs


# 必须存为文本格式的列，以及零填充规则
# key=列名, value=zfill 位数（0 表示不填充，仅转文本）
TEXT_COL_RULES = {
    "院校代码": 4,       # 四川招生代码，4位零填充
    "专业代码": 0,       # 2位，含字母(B1/A2)，保持原样
    "专业组代码": 0,     # 3位，保持原样
    "国标代码": 5,       # 教育部国标代码，5位
    "学校标识码": 10,    # 学校标识码，10位
    "省份代码": 2,       # 省份代码，2位
    "阳光高考ID": 0,     # 不定长，不填充
}


def force_text_format(ws, headers, total_rows):
    """将代码列转为文本格式，并按规则补齐前导零。"""
    for col_name, zfill_width in TEXT_COL_RULES.items():
        if col_name not in headers:
            continue
        col_idx = headers.index(col_name) + 1
        converted = 0
        for row_idx in range(2, total_rows + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is None:
                continue
            val = cell.value
            # 数值转字符串
            if not isinstance(val, str):
                if isinstance(val, float) and val == int(val):
                    val = int(val)
                val = str(val)
            # 零填充
            if zfill_width > 0 and val.isdigit():
                val = val.zfill(zfill_width)
            if val != str(cell.value) or cell.number_format != numbers.FORMAT_TEXT:
                cell.value = val
                cell.number_format = numbers.FORMAT_TEXT
                converted += 1
        if converted > 0:
            print(f"  文本化 [{col_name}]: {converted} 个单元格 (zfill={zfill_width})")


def beautify(filepath):
    """美化单个 Excel 文件。"""
    print(f"正在处理: {filepath.name}")
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    total_rows = ws.max_row
    total_cols = ws.max_column

    # ── 0. 代码列强制文本格式 ──
    force_text_format(ws, headers, total_rows)

    # ── 1. 数据清洗：JSON/地址列 ──
    json_target_cols = {"双一流专业", "特色专业"}
    address_cols = {"地址"}
    for col_name in json_target_cols | address_cols:
        if col_name not in headers:
            continue
        col_idx = headers.index(col_name) + 1
        parser = parse_address_cell if col_name in address_cols else parse_json_cell
        changed = 0
        for row_idx in range(2, total_rows + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            old = cell.value
            new = parser(old)
            if new != old:
                cell.value = new
                changed += 1
        print(f"  清洗 [{col_name}]: {changed} 个单元格")

    # ── 2. 列宽 ──
    widths = calc_col_widths(ws)
    for col_idx, w in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = max(w, 6)

    # ── 3. 识别居中列 ──
    center_cols = identify_center_cols(headers)

    # ── 4. 表头样式 ──
    ws.row_dimensions[1].height = 28
    for col_idx in range(1, total_cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = HEADER_BORDER

    # ── 5. 数据行样式 ──
    for row_idx in range(2, total_rows + 1):
        fill = EVEN_FILL if row_idx % 2 == 0 else ODD_FILL
        for col_idx in range(1, total_cols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = DATA_FONT
            cell.fill = fill
            cell.border = THIN_BORDER
            if (col_idx - 1) in center_cols:
                cell.alignment = DATA_ALIGN_CENTER
            else:
                cell.alignment = DATA_ALIGN

        # 进度
        if row_idx % 10000 == 0:
            print(f"  进度: {row_idx}/{total_rows}")

    # ── 6. 冻结窗格（首行 + 前2列） ──
    ws.freeze_panes = "C2"

    # ── 7. 自动筛选 ──
    last_col_letter = get_column_letter(total_cols)
    ws.auto_filter.ref = f"A1:{last_col_letter}{total_rows}"

    # ── 保存 ──
    wb.save(filepath)
    print(f"  完成! {total_rows}行 x {total_cols}列\n")


if __name__ == "__main__":
    for name in ["院校信息表.xlsx", "专业招生主表.xlsx"]:
        beautify(BASE_DIR / name)
    print("全部完成。")
