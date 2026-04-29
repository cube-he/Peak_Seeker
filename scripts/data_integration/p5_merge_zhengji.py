"""
征集志愿 2023-2025 跨年整合脚本

基于批次类别跨年映射表，将三年征集志愿数据按
  院校+专业 匹配合并到同一行，体现趋势变化。

匹配三层：
  L1 精确：统一科目 + 统一批次代码 + 院校代码 + 专业代码
  L2 名称：统一科目 + 统一批次代码 + 院校代码 + 专业名称
  L3 模糊批次：统一科目 + 院校代码 + 专业名称
"""

import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl.styles import Font, Alignment
from collections import defaultdict
import sys, datetime

sys.stdout.reconfigure(encoding="utf-8")

BASE = Path(r"C:\Users\Administrator\Documents\VolunteerHelper\data\13_征集志愿\普通高考")

def normalize_quotes(s):
    """统一引号：\u201c\u201d → \x22，\uff08 → (, \uff09 → )"""
    return s.replace("\u201c", '"').replace("\u201d", '"').replace("\uff08", "(").replace("\uff09", ")")


# ── 1. 读取映射表 ──────────────────────────────────────────────

mapping_file = BASE / "批次类别跨年映射表.xlsx"
df_batch_map = pd.read_excel(mapping_file, sheet_name="批次映射")
df_subject_map = pd.read_excel(mapping_file, sheet_name="科目映射")

# 科目映射：文科→历史, 理科→物理, 历史→历史, 物理→物理
subject_map = {}
for _, row in df_subject_map.iterrows():
    subject_map[str(row["2023/2024科目"]).strip()] = str(row["2025科目"]).strip()

# 批次映射：(年份, 录取批次, 招生类型) → (统一批次代码, 统一批次名称)
batch_lookup = {}
for _, row in df_batch_map.iterrows():
    code = row["统一批次代码"]
    name = row["统一批次名称"]
    if pd.isna(code):
        continue
    for year in ["2023", "2024", "2025"]:
        batch_col = f"{year}录取批次"
        type_col = f"{year}招生类型"
        batch_val = row[batch_col]
        type_val = row[type_col]
        if pd.notna(batch_val) and pd.notna(type_val):
            key = (year, normalize_quotes(str(batch_val).strip()), normalize_quotes(str(type_val).strip()))
            batch_lookup[key] = (str(code).strip(), str(name).strip())

# 补充映射表中缺失的边缘组合
extra_mappings = [
    # 2025 征集志愿特殊措辞
    ("2025", '本科批A段(国家专项)', '国家专项计划征集志愿', 'BK_A_GJ', '本科批A段(国家专项)'),
    ("2025", '本科批B段', '本科层次职业教育人才培养改革试点', 'BK_B', '本科批B段'),
    ("2025", '本科批B段', '非西藏生源定向西藏就业', 'BK_B', '本科批B段'),
    ("2025", '本科批(预科)', '原"少数民族语言授课为主"', 'BK_1LM', '一类模式本科'),
    ("2025", '本科提前批B段', '其他', 'BK_TQ_B', '本科提前批B段'),
    # 2024 边缘组合
    ("2024", '本科提前批', '国家专项', 'BK_A_GJ', '本科批A段(国家专项)'),
    ("2024", '本科提前批', '司法', 'BK_TQ_A', '本科提前批A段'),
    ("2024", '本科一批', '一类模式预科', 'BK_1LM', '一类模式本科'),
    # 2023 边缘
    ("2023", '高水平运动队', '高水平运动队', 'BK_GSY', '高水平运动队'),
]
for yr, batch, typ, code, name in extra_mappings:
    key = (yr, normalize_quotes(batch), normalize_quotes(typ))
    if key not in batch_lookup:
        batch_lookup[key] = (code, name)

print(f"批次映射条目数(含补充): {len(batch_lookup)}")
print(f"科目映射: {subject_map}")


# ── 2. 读取三年数据并标准化 ────────────────────────────────────

def standardize(df, year_str):
    """标准化科目和批次代码"""
    df = df.copy()
    df["年份"] = int(year_str)

    # 科目标准化
    df["统一科目"] = df["科目"].astype(str).str.strip().map(subject_map)
    unmapped_subj = df[df["统一科目"].isna()]["科目"].unique()
    if len(unmapped_subj) > 0:
        print(f"  [WARN] {year_str} 未映射科目: {unmapped_subj}, 原值兜底")
        df["统一科目"] = df["统一科目"].fillna(df["科目"].astype(str).str.strip())

    # 批次标准化
    def lookup_batch(row):
        batch_str = normalize_quotes(str(row["录取批次"]).strip())
        type_str = normalize_quotes(str(row["招生类型"]).strip())
        key = (year_str, batch_str, type_str)
        result = batch_lookup.get(key)
        if result:
            return pd.Series(result, index=["统一批次代码", "统一批次名称"])
        # 兜底：原值
        return pd.Series(
            [batch_str, batch_str],
            index=["统一批次代码", "统一批次名称"],
        )

    batch_info = df.apply(lookup_batch, axis=1)
    df["统一批次代码"] = batch_info["统一批次代码"]
    df["统一批次名称"] = batch_info["统一批次名称"]

    unmapped = df[
        df.apply(
            lambda r: (
                year_str,
                normalize_quotes(str(r["录取批次"]).strip()),
                normalize_quotes(str(r["招生类型"]).strip()),
            )
            not in batch_lookup,
            axis=1,
        )
    ]
    if len(unmapped) > 0:
        pairs = (
            unmapped[["录取批次", "招生类型"]]
            .drop_duplicates()
            .values.tolist()
        )
        print(f"  [WARN] {year_str} 未映射批次组合 ({len(unmapped)} 行): {pairs}")

    # 院校代码标准化：不同年份dtype不同(str/int/float)，统一转为去零前缀的纯数字字符串
    def normalize_code(val):
        if pd.isna(val):
            return ""
        s = str(val).strip()
        # 去掉 float 后缀 ".0"
        if s.endswith(".0"):
            s = s[:-2]
        # 去掉前导零(院校代码是数字编号，前导零无意义)
        s = s.lstrip("0") or "0"
        return s

    df["院校代码"] = df["院校代码"].apply(normalize_code)
    df["专业代码"] = df["专业代码"].astype(str).str.strip()
    # 专业代码也清理 NaN → ""
    df["专业代码"] = df["专业代码"].replace("nan", "")
    df["专业名称"] = df["专业名称"].astype(str).str.strip()
    df["专业名称"] = df["专业名称"].replace("nan", "")

    return df


data = {}
for year in [2023, 2024, 2025]:
    fn = BASE / f"征集志愿_{year}_合并.xlsx"
    print(f"\n读取 {fn.name} ...")
    df = pd.read_excel(fn)
    print(f"  原始行数: {len(df)}")
    df = standardize(df, str(year))
    data[year] = df
    print(f"  标准化后: {len(df)} 行")


# ── 3. 分层匹配合并 ──────────────────────────────────────────

def make_keys(row):
    """为一条记录生成三层匹配key"""
    subj = row["统一科目"]
    batch = row["统一批次代码"]
    school = row["院校代码"]
    major_code = row["专业代码"]
    major_name = row["专业名称"]
    k1 = (subj, batch, school, major_code)   # L1 精确
    k2 = (subj, batch, school, major_name)   # L2 名称
    k3 = (subj, school, major_name)          # L3 模糊批次
    return k1, k2, k3


def safe_val(val):
    """安全取值，NaN→None"""
    if pd.isna(val):
        return None
    return val


# 输出行结构
class MergedRow:
    __slots__ = [
        "统一批次代码", "统一批次名称", "统一科目",
        "院校代码", "院校名称", "办学性质", "院校地址",
        "专业组代码_2025", "再选科目要求_2025",
        "专业代码_2023", "专业代码_2024", "专业代码_2025",
        "专业名称", "专业备注", "收费标准",
        # 2023: 4次
        "y23_p1", "y23_p2", "y23_p3", "y23_p4",
        # 2024: 4次
        "y24_p1", "y24_p2", "y24_p3", "y24_p4",
        # 2025: 3次(组+专业)
        "y25_gp1", "y25_p1", "y25_gp2", "y25_p2", "y25_gp3", "y25_p3",
        "匹配层级", "批次ID_2023", "批次ID_2024", "批次ID_2025", "备注",
    ]

    def __init__(self):
        for s in self.__slots__:
            setattr(self, s, None)

    def to_dict(self):
        return {s: getattr(self, s) for s in self.__slots__}


def fill_from_2023(mr, row):
    mr.专业代码_2023 = safe_val(row["专业代码"])
    mr.y23_p1 = safe_val(row.get("第1次专业计划数"))
    mr.y23_p2 = safe_val(row.get("第2次专业计划数"))
    mr.y23_p3 = safe_val(row.get("第3次专业计划数"))
    mr.y23_p4 = safe_val(row.get("第4次专业计划数"))
    mr.批次ID_2023 = safe_val(row.get("批次ID"))
    # 基本信息(如果还没填)
    if mr.专业名称 is None:
        mr.专业名称 = safe_val(row["专业名称"])
    if mr.专业备注 is None:
        mr.专业备注 = safe_val(row.get("专业备注"))
    if mr.收费标准 is None:
        mr.收费标准 = safe_val(row.get("收费标准"))
    if mr.院校名称 is None:
        mr.院校名称 = safe_val(row.get("院校名称"))
    if mr.办学性质 is None:
        mr.办学性质 = safe_val(row.get("办学性质"))
    if mr.院校地址 is None:
        mr.院校地址 = safe_val(row.get("院校地址"))


def fill_from_2024(mr, row):
    mr.专业代码_2024 = safe_val(row["专业代码"])
    mr.y24_p1 = safe_val(row.get("第1次专业计划数"))
    mr.y24_p2 = safe_val(row.get("第2次专业计划数"))
    mr.y24_p3 = safe_val(row.get("第3次专业计划数"))
    mr.y24_p4 = safe_val(row.get("第4次专业计划数"))
    mr.批次ID_2024 = safe_val(row.get("批次ID"))
    if mr.专业名称 is None:
        mr.专业名称 = safe_val(row["专业名称"])
    if mr.专业备注 is None:
        mr.专业备注 = safe_val(row.get("专业备注"))
    # 收费标准取更新年份
    fee = safe_val(row.get("收费标准"))
    if fee is not None:
        mr.收费标准 = fee
    if mr.院校名称 is None:
        mr.院校名称 = safe_val(row.get("院校名称"))
    if mr.办学性质 is None:
        mr.办学性质 = safe_val(row.get("办学性质"))
    if mr.院校地址 is None:
        mr.院校地址 = safe_val(row.get("院校地址"))


def fill_from_2025(mr, row):
    mr.专业代码_2025 = safe_val(row["专业代码"])
    mr.专业组代码_2025 = safe_val(row.get("专业组代码"))
    mr.再选科目要求_2025 = safe_val(row.get("再选科目要求"))
    mr.y25_gp1 = safe_val(row.get("第1次专业组计划数"))
    mr.y25_p1 = safe_val(row.get("第1次专业计划数"))
    mr.y25_gp2 = safe_val(row.get("第2次专业组计划数"))
    mr.y25_p2 = safe_val(row.get("第2次专业计划数"))
    mr.y25_gp3 = safe_val(row.get("第3次专业组计划数"))
    mr.y25_p3 = safe_val(row.get("第3次专业计划数"))
    mr.批次ID_2025 = safe_val(row.get("批次ID"))
    if mr.专业名称 is None:
        mr.专业名称 = safe_val(row["专业名称"])
    # 收费标准取最新
    fee = safe_val(row.get("收费标准"))
    if fee is not None:
        mr.收费标准 = fee
    # 专业备注也取最新(如果有)
    remark = safe_val(row.get("专业备注"))
    if remark is not None:
        mr.专业备注 = remark
    if mr.院校名称 is None:
        mr.院校名称 = safe_val(row.get("院校名称"))
    # 2025有办学性质
    nature = safe_val(row.get("办学性质"))
    if nature is not None:
        mr.办学性质 = nature
    addr = safe_val(row.get("院校地址"))
    if addr is not None:
        mr.院校地址 = addr


# === Step 1: 2023 建行 ===
print("\n=== Step 1: 用2023数据建初始行 ===")

# merged_rows: list of MergedRow
merged_rows = []
# 索引: key → list of (index_in_merged_rows,)
idx_l1 = {}  # L1 key → merged row index
idx_l2 = {}  # L2 key → merged row index
idx_l3 = defaultdict(list)  # L3 key → list of merged row indices

for _, row in data[2023].iterrows():
    mr = MergedRow()
    mr.统一批次代码 = row["统一批次代码"]
    mr.统一批次名称 = row["统一批次名称"]
    mr.统一科目 = row["统一科目"]
    mr.院校代码 = row["院校代码"]
    fill_from_2023(mr, row)
    mr.匹配层级 = None  # 会在后续更新

    idx = len(merged_rows)
    merged_rows.append(mr)

    k1, k2, k3 = make_keys(row)
    # L1: 专业代码可能重复(同校同批同代码)，保留第一个
    if k1 not in idx_l1:
        idx_l1[k1] = idx
    if k2 not in idx_l2:
        idx_l2[k2] = idx
    idx_l3[k3].append(idx)

print(f"  2023 建行: {len(merged_rows)}")
print(f"  L1索引: {len(idx_l1)}, L2索引: {len(idx_l2)}, L3索引: {len(idx_l3)}")


# === Step 2: 2024 匹配到已有行或新建 ===
print("\n=== Step 2: 2024 数据匹配 ===")

match_stats_24 = {"L1": 0, "L2": 0, "L3": 0, "new": 0}
used_24 = set()  # 记录已被使用的 merged_row index，避免重复匹配

for _, row in data[2024].iterrows():
    k1, k2, k3 = make_keys(row)
    matched_idx = None
    level = None

    # L1
    if k1 in idx_l1:
        candidate = idx_l1[k1]
        if merged_rows[candidate].专业代码_2024 is None:
            matched_idx = candidate
            level = "L1"

    # L2
    if matched_idx is None and k2 in idx_l2:
        candidate = idx_l2[k2]
        if merged_rows[candidate].专业代码_2024 is None:
            matched_idx = candidate
            level = "L2"

    # L3
    if matched_idx is None and k3 in idx_l3:
        for candidate in idx_l3[k3]:
            if merged_rows[candidate].专业代码_2024 is None:
                matched_idx = candidate
                level = "L3"
                break

    if matched_idx is not None:
        mr = merged_rows[matched_idx]
        fill_from_2024(mr, row)
        if mr.匹配层级 is None:
            mr.匹配层级 = level
        else:
            # 取最松的层级
            levels_order = {"L1": 1, "L2": 2, "L3": 3}
            if levels_order.get(level, 0) > levels_order.get(mr.匹配层级, 0):
                mr.匹配层级 = level
        # L3匹配时记录批次差异
        if level == "L3":
            batch_note = f"2024批次[{row['录取批次']}/{row['招生类型']}]≠2023"
            mr.备注 = (mr.备注 + "; " + batch_note) if mr.备注 else batch_note
        match_stats_24[level] += 1
    else:
        # 新建行
        mr = MergedRow()
        mr.统一批次代码 = row["统一批次代码"]
        mr.统一批次名称 = row["统一批次名称"]
        mr.统一科目 = row["统一科目"]
        mr.院校代码 = row["院校代码"]
        fill_from_2024(mr, row)

        idx = len(merged_rows)
        merged_rows.append(mr)

        # 也加入索引，供2025匹配
        if k1 not in idx_l1:
            idx_l1[k1] = idx
        if k2 not in idx_l2:
            idx_l2[k2] = idx
        idx_l3[k3].append(idx)

        match_stats_24["new"] += 1

print(f"  2024 匹配: L1={match_stats_24['L1']}, L2={match_stats_24['L2']}, L3={match_stats_24['L3']}, 新建={match_stats_24['new']}")
print(f"  当前总行数: {len(merged_rows)}")


# === Step 3: 2025 匹配到已有行或新建 ===
print("\n=== Step 3: 2025 数据匹配 ===")

# 为2025重建索引（含2023+2024已有行）
# 因为2025同一院校同一专业可能有多个专业组，每个专业组独立成行
# 所以需要更灵活的匹配逻辑

match_stats_25 = {"L1": 0, "L2": 0, "L3": 0, "new": 0}

for _, row in data[2025].iterrows():
    k1, k2, k3 = make_keys(row)
    matched_idx = None
    level = None

    # L1
    if k1 in idx_l1:
        candidate = idx_l1[k1]
        if merged_rows[candidate].专业代码_2025 is None:
            matched_idx = candidate
            level = "L1"

    # L2
    if matched_idx is None and k2 in idx_l2:
        candidate = idx_l2[k2]
        if merged_rows[candidate].专业代码_2025 is None:
            matched_idx = candidate
            level = "L2"

    # L3
    if matched_idx is None and k3 in idx_l3:
        for candidate in idx_l3[k3]:
            if merged_rows[candidate].专业代码_2025 is None:
                matched_idx = candidate
                level = "L3"
                break

    if matched_idx is not None:
        mr = merged_rows[matched_idx]
        fill_from_2025(mr, row)
        if mr.匹配层级 is None:
            mr.匹配层级 = level
        else:
            levels_order = {"L1": 1, "L2": 2, "L3": 3}
            if levels_order.get(level, 0) > levels_order.get(mr.匹配层级, 0):
                mr.匹配层级 = level
        # L3批次差异备注
        if level == "L3":
            batch_note = f"2025批次[{row['录取批次']}/{row['招生类型']}]≠历史"
            mr.备注 = (mr.备注 + "; " + batch_note) if mr.备注 else batch_note
        # 更新统一批次(优先2025的批次信息)
        mr.统一批次代码 = row["统一批次代码"]
        mr.统一批次名称 = row["统一批次名称"]
        match_stats_25[level] += 1
    else:
        mr = MergedRow()
        mr.统一批次代码 = row["统一批次代码"]
        mr.统一批次名称 = row["统一批次名称"]
        mr.统一科目 = row["统一科目"]
        mr.院校代码 = row["院校代码"]
        fill_from_2025(mr, row)

        idx = len(merged_rows)
        merged_rows.append(mr)

        if k1 not in idx_l1:
            idx_l1[k1] = idx
        if k2 not in idx_l2:
            idx_l2[k2] = idx
        idx_l3[k3].append(idx)

        match_stats_25["new"] += 1

print(f"  2025 匹配: L1={match_stats_25['L1']}, L2={match_stats_25['L2']}, L3={match_stats_25['L3']}, 新建={match_stats_25['new']}")
print(f"  最终总行数: {len(merged_rows)}")


# === 标记匹配层级 ===
for mr in merged_rows:
    has_23 = mr.专业代码_2023 is not None
    has_24 = mr.专业代码_2024 is not None
    has_25 = mr.专业代码_2025 is not None
    if has_23 + has_24 + has_25 == 1:
        mr.匹配层级 = "单年"
    elif mr.匹配层级 is None:
        mr.匹配层级 = "L1"  # 默认


# ── 4. 输出 ──────────────────────────────────────────────────

print("\n=== 生成输出 ===")

# 构造 DataFrame
col_map = {
    "统一批次代码": "统一批次代码",
    "统一批次名称": "统一批次名称",
    "统一科目": "统一科目",
    "院校代码": "院校代码",
    "院校名称": "院校名称",
    "办学性质": "办学性质",
    "院校地址": "院校地址",
    "专业组代码_2025": "专业组代码(2025)",
    "再选科目要求_2025": "再选科目要求(2025)",
    "专业代码_2023": "专业代码(2023)",
    "专业代码_2024": "专业代码(2024)",
    "专业代码_2025": "专业代码(2025)",
    "专业名称": "专业名称",
    "专业备注": "专业备注",
    "收费标准": "收费标准",
    "y23_p1": "2023_第1次计划数",
    "y23_p2": "2023_第2次计划数",
    "y23_p3": "2023_第3次计划数",
    "y23_p4": "2023_第4次计划数",
    "y24_p1": "2024_第1次计划数",
    "y24_p2": "2024_第2次计划数",
    "y24_p3": "2024_第3次计划数",
    "y24_p4": "2024_第4次计划数",
    "y25_gp1": "2025_第1次组计划数",
    "y25_p1": "2025_第1次计划数",
    "y25_gp2": "2025_第2次组计划数",
    "y25_p2": "2025_第2次计划数",
    "y25_gp3": "2025_第3次组计划数",
    "y25_p3": "2025_第3次计划数",
    "匹配层级": "匹配层级",
    "批次ID_2023": "2023批次ID",
    "批次ID_2024": "2024批次ID",
    "批次ID_2025": "2025批次ID",
    "备注": "备注",
}

rows_data = [mr.to_dict() for mr in merged_rows]
df_out = pd.DataFrame(rows_data)
df_out = df_out.rename(columns=col_map)

# 排序：统一批次代码 + 院校代码
df_out = df_out.sort_values(["统一批次代码", "院校代码"]).reset_index(drop=True)

# 输出 Excel
out_path = BASE / "征集志愿_2023_2025_跨年整合.xlsx"
with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
    df_out.to_excel(writer, sheet_name="跨年整合", index=False)
    ws = writer.sheets["跨年整合"]

    # 样式：微软雅黑10pt，首行冻结+加粗
    header_font = Font(name="微软雅黑", size=10, bold=True)
    body_font = Font(name="微软雅黑", size=10)

    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_cells in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row_cells:
            cell.font = body_font

    ws.freeze_panes = "A2"

    # 自动列宽(简单估算)
    for col_cells in ws.columns:
        max_len = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            try:
                val = str(cell.value) if cell.value else ""
                # 中文字符按2宽度计算
                length = sum(2 if ord(c) > 127 else 1 for c in val)
                max_len = max(max_len, length)
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

print(f"输出: {out_path}")
print(f"总行数: {len(df_out)}")


# ── 5. 统计 ──────────────────────────────────────────────────

stats_lines = []
stats_lines.append(f"征集志愿 2023-2025 跨年整合统计")
stats_lines.append(f"生成时间: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
stats_lines.append(f"{'='*60}")
stats_lines.append(f"")
stats_lines.append(f"源数据: 2023={len(data[2023])}行, 2024={len(data[2024])}行, 2025={len(data[2025])}行")
stats_lines.append(f"总行数: {len(df_out)}")
stats_lines.append(f"")

# 年份覆盖统计
has_23 = df_out["专业代码(2023)"].notna()
has_24 = df_out["专业代码(2024)"].notna()
has_25 = df_out["专业代码(2025)"].notna()

three_year = has_23 & has_24 & has_25
only_23_24 = has_23 & has_24 & ~has_25
only_24_25 = ~has_23 & has_24 & has_25
only_23_25 = has_23 & ~has_24 & has_25
only_23 = has_23 & ~has_24 & ~has_25
only_24 = ~has_23 & has_24 & ~has_25
only_25 = ~has_23 & ~has_24 & has_25

stats_lines.append(f"三年都有: {three_year.sum()}")

# 三年匹配层级分布
if three_year.sum() > 0:
    level_dist = df_out[three_year]["匹配层级"].value_counts()
    for lv, cnt in level_dist.items():
        stats_lines.append(f"  匹配层级 {lv}: {cnt}")

stats_lines.append(f"仅23+24: {only_23_24.sum()}")
stats_lines.append(f"仅24+25: {only_24_25.sum()}")
stats_lines.append(f"仅23+25: {only_23_25.sum()}")
stats_lines.append(f"仅2023: {only_23.sum()}")
stats_lines.append(f"仅2024: {only_24.sum()}")
stats_lines.append(f"仅2025: {only_25.sum()}")

stats_lines.append(f"")
stats_lines.append(f"各批次匹配情况:")

batch_groups = df_out.groupby("统一批次代码")
for batch_code, grp in sorted(batch_groups, key=lambda x: x[0]):
    batch_name = grp["统一批次名称"].iloc[0]
    g_23 = grp["专业代码(2023)"].notna()
    g_24 = grp["专业代码(2024)"].notna()
    g_25 = grp["专业代码(2025)"].notna()
    n_three = (g_23 & g_24 & g_25).sum()
    n_two = ((g_23.astype(int) + g_24.astype(int) + g_25.astype(int)) == 2).sum()
    n_one = ((g_23.astype(int) + g_24.astype(int) + g_25.astype(int)) == 1).sum()
    stats_lines.append(f"  {batch_code}({batch_name}): 三年={n_three}, 两年={n_two}, 单年={n_one}")

stats_text = "\n".join(stats_lines)
print(f"\n{stats_text}")

stats_path = BASE / "跨年整合统计.txt"
with open(stats_path, "w", encoding="utf-8") as f:
    f.write(stats_text)

print(f"\n统计文件: {stats_path}")
print("完成!")
