# -*- coding: utf-8 -*-
"""Tests for P3.2 pilot scan."""
from openpyxl import Workbook

from scripts.data_integration.p3_pilot_2023_2024 import scan_xlsx


def test_scan_xlsx_ok_row_count(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.append(["院校代码", "专业代码", "计划数"])
    for i in range(10):
        ws.append([f"1000{i}", "01", i])
    p = tmp_path / "x.xlsx"
    wb.save(p)

    r = scan_xlsx(p)
    assert r["n_rows"] == 10
    assert r["status"] == "ok"
    assert r["cols"][:3] == ["院校代码", "专业代码", "计划数"]


def test_scan_xlsx_many_empty_rows_warns(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.append(["a", "b"])
    ws.append(["v1", "v2"])
    for _ in range(5):
        ws.append([None, None])
    p = tmp_path / "x.xlsx"
    wb.save(p)

    r = scan_xlsx(p)
    assert r["status"] == "warn"
    assert r["empty_ratio"] > 0.3


def test_scan_xlsx_empty_workbook(tmp_path):
    wb = Workbook()
    p = tmp_path / "e.xlsx"
    wb.save(p)

    r = scan_xlsx(p)
    # Empty workbook: 0 rows → fail
    assert r["status"] == "fail"
    assert r["n_rows"] == 0
