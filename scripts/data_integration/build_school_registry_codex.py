# -*- coding: utf-8 -*-
"""Build Layer 2 school registry from 03 + 02 sources only.

This script intentionally does not read:
  - data/03_专家版主表/output
  - data/_pipeline

Outputs:
  data/codex_layer2_school_registry/school_registry.csv
  data/codex_layer2_school_registry/school_registry.json
  data/codex_layer2_school_registry/quality_report.md
  data/codex_layer2_school_registry/lineage.json
"""
from __future__ import annotations

import json
import math
import re
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[2]
DATA = ROOT / "data"
OUT = DATA / "codex_layer2_school_registry"

COLLEGE = "\u9662\u6821"
COLLEGE_CODE = "\u9662\u6821\u4ee3\u7801"
SICHUAN = "\u56db\u5ddd"


def data_dir(prefix: str) -> Path:
    return next(p for p in DATA.iterdir() if p.is_dir() and p.name.startswith(prefix))


DIR02 = data_dir("02_")
DIR03 = data_dir("03_")


def clean(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    text = str(value).strip()
    if text == "" or text.lower() in {"nan", "none", "null", "<na>"}:
        return None
    return text


def norm_name(value: Any) -> str | None:
    text = clean(value)
    if text is None:
        return None
    text = re.sub(r"\s+", "", text)
    return text.replace("（", "(").replace("）", ")")


def to_int(value: Any) -> int | None:
    text = clean(value)
    if text is None:
        return None
    text = text.replace(",", "").replace("%", "")
    if text.endswith(".0"):
        text = text[:-2]
    return int(text) if re.fullmatch(r"-?\d+", text) else None


def to_float(value: Any) -> float | None:
    text = clean(value)
    if text is None:
        return None
    text = text.replace(",", "").replace("%", "")
    try:
        return float(text)
    except ValueError:
        return None


def zfill_school_code(value: Any) -> str | None:
    text = clean(value)
    if text is None:
        return None
    if text.endswith(".0"):
        text = text[:-2]
    return text.zfill(4) if re.fullmatch(r"\d+", text) else text


def first_non_empty(values: pd.Series) -> Any:
    for value in values:
        value = clean(value)
        if value is not None:
            return value
    return None


def find_file(required_tokens: list[str]) -> Path:
    matches = []
    for path in DIR02.glob("*.xlsx"):
        if all(token in path.name for token in required_tokens):
            matches.append(path)
    if not matches:
        raise FileNotFoundError(required_tokens)
    return sorted(matches, key=lambda p: (len(p.name), p.name))[0]


def read_header_detect(path: Path, required_cols: list[str]) -> pd.DataFrame:
    for header in range(0, 6):
        df = pd.read_excel(path, dtype=str, header=header)
        cols = {str(c).strip() for c in df.columns}
        if all(col in cols for col in required_cols):
            return df
    return pd.read_excel(path, dtype=str)


def load_expert_main() -> tuple[Path, pd.DataFrame]:
    path = DIR03 / "2026\u56db\u5ddd\u9ad8\u8003\u5fd7\u613f--\u4e13\u5bb6\u7248\u5408\u5e76\u6574\u7406.xlsx"
    if not path.exists():
        raise FileNotFoundError(path)
    df = pd.read_excel(path, sheet_name=SICHUAN, dtype=str)
    if COLLEGE not in df.columns or COLLEGE_CODE not in df.columns:
        raise RuntimeError("Expert source missing required school columns")
    return path, df


def load_base_sources() -> dict[str, tuple[Path, pd.DataFrame]]:
    return {
        "school_db": (find_file(["\u9662\u6821\u5e93", "\u5168\u56fd"]), pd.read_excel(find_file(["\u9662\u6821\u5e93", "\u5168\u56fd"]), dtype=str)),
        "ranking": (find_file(["\u5927\u5b66\u6392\u540d", "\u5168\u56fd"]), pd.read_excel(find_file(["\u5927\u5b66\u6392\u540d", "\u5168\u56fd"]), dtype=str)),
        "registry": (find_file(["\u5168\u56fd\u9ad8\u6821\u5b8c\u6574\u540d\u5f55"]), read_header_detect(find_file(["\u5168\u56fd\u9ad8\u6821\u5b8c\u6574\u540d\u5f55"]), ["\u5b66\u6821\u540d\u79f0"])),
        "charter": (find_file(["\u62db\u751f\u7ae0\u7a0b\u7ed3\u6784\u5316"]), pd.read_excel(find_file(["\u62db\u751f\u7ae0\u7a0b\u7ed3\u6784\u5316"]), dtype=str)),
        "satisfaction": (find_file(["\u9662\u6821\u6ee1\u610f\u5ea6"]), pd.read_excel(find_file(["\u9662\u6821\u6ee1\u610f\u5ea6"]), dtype=str)),
        "discipline_eval": (find_file(["\u5b66\u79d1\u8bc4\u4f30"]), pd.read_excel(find_file(["\u5b66\u79d1\u8bc4\u4f30"]), dtype=str)),
    }


EXPERT_FIELD_ALIASES = {
    "school_code": ["\u9662\u6821\u4ee3\u7801"],
    "school_name": ["\u9662\u6821"],
    "province": ["\u9662\u6821\u7701\u4efd", "\u6240\u5728\u7701", "\u7701\u4efd"],
    "city": ["\u9662\u6821\u57ce\u5e02", "\u57ce\u5e02"],
    "city_tier": ["\u57ce\u5e02\u7b49\u7ea7", "\u57ce\u5e02\u6c34\u5e73"],
    "school_type_03": ["\u9662\u6821\u7c7b\u578b", "\u7c7b\u578b"],
    "nature_03": ["\u529e\u5b66\u6027\u8d28", "\u516c\u79c1\u6027\u8d28"],
    "authority_03": ["\u96b6\u5c5e\u90e8\u95e8", "\u96b6\u5c5e\u5355\u4f4d"],
    "labels_03": ["\u9662\u6821\u6807\u7b7e"],
    "tier_03": ["\u9662\u6821\u6863\u6b21", "\u9662\u6821\u6c34\u5e73"],
    "background_03": ["\u9662\u6821\u80cc\u666f"],
    "level_03": ["\u9662\u6821\u5c42\u7ea7"],
    "rank_03": ["\u9662\u6821\u6392\u540d", "\u5927\u5b66\u6392\u540d"],
    "master_count_03": ["\u7855\u58eb\u70b9\u6570\u91cf", "\u7855\u58eb\u70b9\u6570"],
    "doctor_count_03": ["\u535a\u58eb\u70b9\u6570\u91cf", "\u535a\u58eb\u70b9\u6570"],
    "master_subjects_03": ["\u7855\u58eb\u70b9\u4e13\u4e1a"],
    "doctor_subjects_03": ["\u535a\u58eb\u70b9\u4e13\u4e1a"],
    "postgraduate_rate_03": ["\u4fdd\u7814\u7387"],
    "is_double_first_class_03": ["\u662f\u5426\u53cc\u4e00\u6d41"],
    "evolution_03": ["\u53d8\u8fc1\u53f2"],
    "merge_history_03": ["\u66f4\u540d\u5408\u5e76\u60c5\u51b5"],
    "major_transfer_03": ["\u8f6c\u4e13\u4e1a\u60c5\u51b5", "\u8f6c\u4e13\u4e1a"],
    "admission_guide_03": ["\u62db\u751f\u7b80\u7ae0"],
}


def extract_expert_registry(expert: pd.DataFrame) -> pd.DataFrame:
    df = expert.copy()
    df["_school_code"] = df[COLLEGE_CODE].map(zfill_school_code)
    df = df.dropna(subset=["_school_code"])
    records = []
    for code, group in df.groupby("_school_code", sort=True):
        record = {"school_code": code}
        for final_name, aliases in EXPERT_FIELD_ALIASES.items():
            for alias in aliases:
                if alias in group.columns:
                    value = first_non_empty(group[alias])
                    if value is not None:
                        record[final_name] = value
                        break
        record["source03_rows"] = int(len(group))
        record["source03_subjects"] = "|".join(sorted({x for x in group.get("\u79d1\u76ee", pd.Series(dtype=str)).map(clean).dropna()}))
        record["source03_batches"] = "|".join(sorted({x for x in group.get("\u6279\u6b21", pd.Series(dtype=str)).map(clean).dropna()}))
        records.append(record)
    registry = pd.DataFrame(records)
    registry["school_name_norm"] = registry["school_name"].map(norm_name)
    return registry


def merge_source(
    registry: pd.DataFrame,
    source_df: pd.DataFrame,
    source_name_col: str,
    fields: list[str],
    prefix: str,
) -> tuple[pd.DataFrame, dict[str, Any]]:
    src = source_df.copy()
    if source_name_col not in src.columns:
        return registry, {"error": f"missing source name column: {source_name_col}"}
    src["school_name_norm"] = src[source_name_col].map(norm_name)
    src = src.dropna(subset=["school_name_norm"]).drop_duplicates("school_name_norm", keep="first")
    available = [field for field in fields if field in src.columns]
    renamed = {field: f"{prefix}_{field}" for field in available}
    keep = ["school_name_norm"] + available
    merged = registry.merge(src[keep].rename(columns=renamed), on="school_name_norm", how="left")
    marker = f"{prefix}_{available[0]}" if available else None
    matched = int(merged[marker].notna().sum()) if marker else 0
    return merged, {
        "matched": matched,
        "total": int(len(registry)),
        "match_rate": round(matched / len(registry), 4) if len(registry) else 0,
        "fields_added": list(renamed.values()),
    }


def add_discipline_summary(registry: pd.DataFrame, eval_df: pd.DataFrame) -> tuple[pd.DataFrame, dict[str, Any]]:
    if "\u9662\u6821\u540d\u79f0" not in eval_df.columns or "\u5c42\u7ea7" not in eval_df.columns:
        return registry, {"error": "discipline eval missing columns"}
    df = eval_df.copy()
    df["school_name_norm"] = df["\u9662\u6821\u540d\u79f0"].map(norm_name)
    df = df.dropna(subset=["school_name_norm"])
    grade_order = ["A+", "A", "A-", "B+", "B", "B-", "C+", "C", "C-"]

    def summarize(values: pd.Series) -> str:
        counts = Counter(clean(v) for v in values if clean(v))
        parts = [f"{grade}:{counts[grade]}" for grade in grade_order if counts.get(grade)]
        return ", ".join(parts)

    summary = df.groupby("school_name_norm")["\u5c42\u7ea7"].apply(summarize).reset_index(name="discipline_eval_summary_02")
    merged = registry.merge(summary, on="school_name_norm", how="left")
    matched = int(merged["discipline_eval_summary_02"].notna().sum())
    return merged, {
        "matched": matched,
        "total": int(len(registry)),
        "match_rate": round(matched / len(registry), 4) if len(registry) else 0,
    }


def finalize_types(registry: pd.DataFrame) -> pd.DataFrame:
    numeric_float = ["postgraduate_rate_03"]
    numeric_int_patterns = ["count", "rank", "\u6570", "\u6392\u540d", "\u70ed\u5ea6", "\u5e74\u4efd"]
    for col in registry.columns:
        if col in numeric_float:
            registry[col] = registry[col].map(to_float)
        elif any(pattern in col for pattern in numeric_int_patterns):
            registry[col] = registry[col].map(lambda x: to_int(x) if clean(x) is not None else None)
    registry = registry.drop(columns=["school_name_norm"])
    ordered = [
        "school_code", "school_name", "province", "city", "city_tier", "school_type_03",
        "nature_03", "authority_03", "level_03", "labels_03", "tier_03", "background_03",
        "rank_03", "master_count_03", "doctor_count_03", "postgraduate_rate_03",
        "is_double_first_class_03", "admission_guide_03", "source03_rows", "source03_subjects",
        "source03_batches",
    ]
    remaining = [c for c in registry.columns if c not in ordered]
    return registry[[c for c in ordered if c in registry.columns] + remaining]


CHINESE_FIELD_NAMES = {
    "school_code": "院校代码",
    "school_name": "院校名称",
    "province": "院校省份",
    "city": "院校城市",
    "city_tier": "城市等级",
    "school_type_03": "院校类型",
    "nature_03": "办学性质",
    "authority_03": "隶属部门",
    "level_03": "院校层级",
    "labels_03": "院校标签",
    "tier_03": "院校档次",
    "background_03": "院校背景",
    "rank_03": "院校排名",
    "master_count_03": "硕士点数量",
    "doctor_count_03": "博士点数量",
    "postgraduate_rate_03": "保研率",
    "is_double_first_class_03": "是否双一流",
    "admission_guide_03": "招生简章",
    "source03_rows": "03源记录数",
    "source03_subjects": "03源科目",
    "source03_batches": "03源批次",
    "master_subjects_03": "硕士点专业",
    "doctor_subjects_03": "博士点专业",
    "merge_history_03": "更名合并情况",
    "major_transfer_03": "转专业情况",
    "discipline_eval_summary_02": "学科评估摘要",
}


SOURCE_SUFFIXES = {
    "schooldb02_": "_院校库",
    "registry02_": "_全国名录",
    "charter02_": "_招生章程",
    "sat02_": "_院校满意度",
}


def to_chinese_columns(registry: pd.DataFrame) -> pd.DataFrame:
    """Rename output columns to Chinese source-style names.

    When the same source field exists in multiple tables, append a source suffix
    so fields remain unique and traceable.
    """
    rename: dict[str, str] = {}
    for column in registry.columns:
        if column in CHINESE_FIELD_NAMES:
            rename[column] = CHINESE_FIELD_NAMES[column]
            continue
        for prefix, suffix in SOURCE_SUFFIXES.items():
            if column.startswith(prefix):
                rename[column] = f"{column.removeprefix(prefix)}{suffix}"
                break
    output = registry.rename(columns=rename)
    # Guard against accidental duplicate names after source-style renaming.
    seen: dict[str, int] = {}
    final_columns = []
    for column in output.columns:
        if column not in seen:
            seen[column] = 1
            final_columns.append(column)
        else:
            seen[column] += 1
            final_columns.append(f"{column}_{seen[column]}")
    output.columns = final_columns
    if "院校代码" in output.columns:
        output["院校代码"] = output["院校代码"].map(lambda x: str(x).zfill(4) if clean(x) is not None else x)
    return output


def write_xlsx_text_code(df: pd.DataFrame, path: Path) -> None:
    """Write XLSX and keep 院校代码 as a 4-character text field."""
    wb = Workbook()
    ws = wb.active
    ws.title = "院校注册表"
    ws.append(list(df.columns))
    code_col_idx = list(df.columns).index("院校代码") + 1 if "院校代码" in df.columns else None
    for row in df.itertuples(index=False, name=None):
        ws.append(list(row))
    if code_col_idx:
        for cell in ws.iter_cols(min_col=code_col_idx, max_col=code_col_idx, min_row=2, max_row=ws.max_row):
            for item in cell:
                item.value = str(item.value).zfill(4) if item.value is not None else None
                item.number_format = "@"
    ws.freeze_panes = "A2"
    for idx, column in enumerate(df.columns, start=1):
        width = min(max(len(str(column)) + 2, 10), 32)
        ws.column_dimensions[get_column_letter(idx)].width = width
    wb.save(path)


def build_registry() -> tuple[pd.DataFrame, dict[str, Any]]:
    expert_path, expert = load_expert_main()
    sources = load_base_sources()
    registry = extract_expert_registry(expert)
    stats: dict[str, Any] = {
        "expert_source": str(expert_path.relative_to(ROOT)),
        "expert_rows": int(len(expert)),
        "expert_school_count": int(len(registry)),
    }

    school_db_path, school_db = sources["school_db"]
    registry, stats["school_db"] = merge_source(registry, school_db, "\u4e2d\u6587\u540d\u79f0", [
        "\u4ee3\u7801", "\u56fd\u6807\u4ee3\u7801", "\u529e\u5b66\u6027\u8d28", "\u5b66\u5386\u5c42\u6b21",
        "\u96b6\u5c5e\u90e8\u95e8", "\u9662\u6821\u7c7b\u578b", "\u9662\u6821\u7279\u8272",
        "\u7701\u4efd\u4ee3\u7801", "\u7701\u4efd\u540d\u79f0", "\u57ce\u5e02", "\u70ed\u5ea6",
        "\u7efc\u5408\u6392\u540d", "\u8f6f\u79d1\u6392\u540d", "\u6821\u53cb\u4f1a\u6392\u540d",
        "QS\u6392\u540d", "USNews\u6392\u540d", "\u6559\u80b2\u90e8\u6392\u540d",
        "\u7537\u751f\u6bd4\u4f8b", "\u5973\u751f\u6bd4\u4f8b", "\u5efa\u6821\u5e74\u4efd",
        "\u4fdd\u7814\u7387", "\u5347\u5b66\u7387", "\u53cc\u4e00\u6d41\u5b66\u79d1\u6570",
        "A\u7c7b\u5b66\u79d1\u6570", "\u56fd\u5bb6\u7ea7\u6570\u91cf", "\u7701\u7ea7\u6570\u91cf",
        "\u53cc\u4e00\u6d41\u4e13\u4e1a", "\u7279\u8272\u4e13\u4e1a", "Logo\u5730\u5740",
        "Banner\u5730\u5740", "\u5730\u5740",
    ], "schooldb02")
    stats["school_db"]["source"] = str(school_db_path.relative_to(ROOT))

    registry_path, registry_src = sources["registry"]
    registry, stats["national_registry"] = merge_source(registry, registry_src, "\u5b66\u6821\u540d\u79f0", [
        "\u9633\u5149\u9ad8\u8003ID", "\u5b66\u6821\u5b98\u7f51", "\u62db\u751f\u7f51\u5740",
        "\u62db\u529e\u7535\u8bdd", "\u53cc\u4e00\u6d41",
    ], "registry02")
    stats["national_registry"]["source"] = str(registry_path.relative_to(ROOT))

    charter_path, charter = sources["charter"]
    registry, stats["charter"] = merge_source(registry, charter, "\u5b66\u6821\u540d\u79f0", [
        "\u9633\u5149\u9ad8\u8003ID", "\u6709\u7ae0\u7a0b", "\u8c03\u6863\u6bd4\u4f8b",
        "\u4e13\u4e1a\u5206\u914d\u89c4\u5219", "\u540c\u5206\u89c4\u5219", "\u5916\u8bed\u8981\u6c42",
        "\u5355\u79d1\u8981\u6c42", "\u4f53\u68c0\u9650\u5236", "\u52a0\u5206\u653f\u7b56",
        "\u5b66\u8d39", "\u8f6c\u4e13\u4e1a\u9650\u5236", "\u670d\u4ece\u8c03\u5242",
        "\u6765\u6e90\u7f51\u5740", "\u91c7\u96c6\u65f6\u95f4",
    ], "charter02")
    stats["charter"]["source"] = str(charter_path.relative_to(ROOT))

    sat_path, sat = sources["satisfaction"]
    registry, stats["satisfaction"] = merge_source(registry, sat, "\u9662\u6821\u540d\u79f0", [
        "\u9633\u5149\u9ad8\u8003ID", "\u7efc\u5408\u6ee1\u610f\u5ea6", "\u7efc\u5408\u8bc4\u4ef7\u4eba\u6570",
        "\u7efc\u54081\u661f", "\u7efc\u54082\u661f", "\u7efc\u54083\u661f", "\u7efc\u54084\u661f", "\u7efc\u54085\u661f",
        "\u751f\u6d3b\u6ee1\u610f\u5ea6", "\u751f\u6d3b1\u661f", "\u751f\u6d3b2\u661f", "\u751f\u6d3b3\u661f",
        "\u751f\u6d3b4\u661f", "\u751f\u6d3b5\u661f", "\u73af\u5883\u6ee1\u610f\u5ea6", "\u73af\u58831\u661f",
        "\u73af\u58832\u661f", "\u73af\u58833\u661f", "\u73af\u58834\u661f", "\u73af\u58835\u661f", "\u6765\u6e90\u7f51\u5740",
    ], "sat02")
    stats["satisfaction"]["source"] = str(sat_path.relative_to(ROOT))

    eval_path, eval_df = sources["discipline_eval"]
    registry, stats["discipline_eval"] = add_discipline_summary(registry, eval_df)
    stats["discipline_eval"]["source"] = str(eval_path.relative_to(ROOT))

    registry = finalize_types(registry)
    stats["final_columns"] = int(len(registry.columns))
    stats["duplicate_school_code_rows"] = int(registry.duplicated(["school_code"], keep=False).sum())
    return registry, stats


def write_outputs(registry: pd.DataFrame, stats: dict[str, Any]) -> None:
    OUT.mkdir(parents=True, exist_ok=True)
    registry.to_csv(OUT / "school_registry_internal_english.csv", index=False, encoding="utf-8-sig")
    registry.to_json(OUT / "school_registry_internal_english.json", orient="records", force_ascii=False, indent=2)
    output_registry = to_chinese_columns(registry)
    output_registry.to_csv(OUT / "school_registry.csv", index=False, encoding="utf-8-sig")
    output_registry.to_json(OUT / "school_registry.json", orient="records", force_ascii=False, indent=2)
    write_xlsx_text_code(output_registry, OUT / "school_registry.xlsx")
    lineage = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "layer": "Layer 2 - school_registry",
        "inputs": {
            "03_expert_main": stats["expert_source"],
            "02_sources": {
                key: value.get("source")
                for key, value in stats.items()
                if isinstance(value, dict) and value.get("source")
            },
            "excluded": ["data/03_专家版主表/output", "data/_pipeline"],
        },
        "merge_policy": {
            "base": "03专家版主表按院校代码去重，每所院校取首个非空院校级字段",
            "join_key": "院校名称规范化后精确匹配",
            "02_usage": "仅追加补充字段，不覆盖03已有字段",
            "output_columns": "主输出使用中文源字段名；跨来源同名字段追加来源后缀，英文内部字段另存为 school_registry_internal_english.*",
        },
        "stats": stats,
    }
    (OUT / "lineage.json").write_text(json.dumps(lineage, ensure_ascii=False, indent=2), encoding="utf-8")

    lines = [
        "# Layer 2 院校注册表质量报告",
        "",
        f"- 生成时间: {lineage['generated_at']}",
        f"- 主源: `{stats['expert_source']}`",
        "- 未读取: `data/03_专家版主表/output`, `data/_pipeline`",
        "",
        "## 输出",
        "",
        "- `school_registry.csv`（中文字段主表）",
        "- `school_registry.json`（中文字段主表）",
        "- `school_registry.xlsx`（中文字段主表，院校代码按文本保留4位）",
        "- `school_registry_internal_english.csv`（内部英文备份）",
        "- `school_registry_internal_english.json`（内部英文备份）",
        "- `lineage.json`",
        "",
        "## 统计",
        "",
        f"- 专家版主表行数: {stats['expert_rows']:,}",
        f"- 院校数: {stats['expert_school_count']:,}",
        f"- 字段数: {stats['final_columns']:,}",
        f"- 院校代码重复行: {stats['duplicate_school_code_rows']:,}",
        "",
        "## 02基础库匹配率",
        "",
    ]
    for key in ["school_db", "national_registry", "charter", "satisfaction", "discipline_eval"]:
        value = stats[key]
        lines.append(f"- {key}: {value['matched']:,}/{value['total']:,} ({value['match_rate']:.2%})")
    lines.extend([
        "",
        "## 口径说明",
        "",
        "- `03` 字段保留源文件中文字段名作为主字段。",
        "- `02` 补充字段保留源文件中文字段名，并追加来源后缀，如 `_院校库`、`_全国名录`、`_招生章程`、`_院校满意度`。",
        "- 学科评估按院校聚合为 `学科评估摘要`，格式如 `A+:3, A:5`。",
        "- 本层不使用 `01_核心录取数据`，因为 01 属于录取事实/分数线层，不是院校注册信息层。",
    ])
    (OUT / "quality_report.md").write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    registry, stats = build_registry()
    write_outputs(registry, stats)
    print(json.dumps({
        "output": str(OUT),
        "rows": len(registry),
        "columns": len(registry.columns),
        "duplicate_school_code_rows": stats["duplicate_school_code_rows"],
        "school_db_match_rate": stats["school_db"]["match_rate"],
        "charter_match_rate": stats["charter"]["match_rate"],
        "satisfaction_match_rate": stats["satisfaction"]["match_rate"],
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
