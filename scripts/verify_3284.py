# -*- coding: utf-8 -*-
"""Batch 3284 verification script - generates verified xlsx"""
import sys, io, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from collections import defaultdict

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
src_path = os.path.join(BASE, 'data', '13_征集志愿', '普通高考', '本科批次',
    '3284_2023_文理综合_本科一批_征集志愿_第一次',
    '3284_2023_文理综合_本科一批_征集志愿_第一次_mimo-v2-omni.xlsx')
out_path = os.path.join(BASE, 'data', '13_征集志愿', '普通高考', '本科批次',
    '3284_2023_文理综合_本科一批_征集志愿_第一次',
    '3284_2023_文理综合_本科一批_征集志愿_第一次_已校验.xlsx')

# Load source
src_wb = openpyxl.load_workbook(src_path)
src_ws = src_wb.active

# Create output
out_wb = openpyxl.Workbook()
out_ws = out_wb.active
out_ws.title = '征集志愿'

# Styles
header_font = Font(name='微软雅黑', size=10, bold=True)
data_font = Font(name='微软雅黑', size=10)
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
green_fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
red_fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')

# Headers
headers = ['科类', '院校代码', '院校名称', '院校地址', '专业代码', '专业名称',
           '专业备注', '专业计划数', '收费标准', '院校备注', '页码', '校正备注']
for col, h in enumerate(headers, 1):
    cell = out_ws.cell(1, col, h)
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

widths = [6, 10, 20, 30, 8, 25, 40, 10, 10, 40, 6, 40]
for col, w in enumerate(widths, 1):
    out_ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w


def split_major_name(name):
    """Split bracket content from major name to notes field."""
    if not name:
        return name, ''
    idx = -1
    for i, ch in enumerate(str(name)):
        if ch in '(\uff08':
            idx = i
            break
    if idx == -1:
        return str(name), ''
    pure_name = str(name)[:idx]
    bracket_content = str(name)[idx:]
    parts = []
    temp = bracket_content
    while temp:
        if temp[0] in '(\uff08':
            depth = 0
            end = -1
            for j, c in enumerate(temp):
                if c in '(\uff08':
                    depth += 1
                elif c in ')\uff09':
                    depth -= 1
                    if depth == 0:
                        end = j
                        break
            if end > 0:
                parts.append(temp[1:end])
                temp = temp[end+1:]
            else:
                parts.append(temp[1:])
                break
        else:
            temp = temp[1:]
    note = '\uff1b'.join(parts) if parts else ''
    return pure_name, note


def clean_vmark(text):
    """Remove [V] marks from text."""
    if not text:
        return str(text) if text is not None else '', False
    s = str(text)
    if '[V]' in s:
        return s.replace('[V]', '').strip(), True
    return s, False


def write_row(ws, row_num, values, font=None, fill=None):
    for col, val in enumerate(values, 1):
        cell = ws.cell(row_num, col, val)
        if font:
            cell.font = font
        if fill:
            cell.fill = fill


# Read all source rows
all_rows = []
for row in range(2, src_ws.max_row + 1):
    vals = [src_ws.cell(row, c).value for c in range(1, 12)]
    all_rows.append(vals)

# Build page -> rows mapping
page_rows = defaultdict(list)
for i, vals in enumerate(all_rows):
    page = vals[10]
    page_rows[page].append({
        'idx': i,
        'kl': vals[0] or '', 'code': vals[1] or '', 'name': vals[2] or '',
        'addr': vals[3] or '', 'mjcode': str(vals[4]) if vals[4] is not None else '',
        'mjname': str(vals[5]) if vals[5] else '', 'mjnote': str(vals[6]) if vals[6] else '',
        'plan': vals[7], 'fee': vals[8], 'yxnote': str(vals[9]) if vals[9] else '',
        'page': page
    })

# ===== MISSING DATA DEFINITIONS =====
# 文科 2106 辽宁工程技术大学 (pages 6-7 cross-page loss)
wk_2106 = [
    ['文科', '2106', '辽宁工程技术大学', '辽宁省阜新市中华路47号', '62',
     '英语', '阜新校区中华路校园；招收英语语种考生', 2, '4800',
     '认同四川省少数民族地区加分项目，但分值最高20分。', 7],
    ['文科', '2106', '辽宁工程技术大学', '辽宁省阜新市中华路47号', '64',
     '财务管理', '阜新校区玉龙校园', 1, '5200',
     '认同四川省少数民族地区加分项目，但分值最高20分。', 7],
]

# 文科 2308 哈尔滨医科大学 (pages 7-8 cross-page loss)
wk_2308 = [
    ['文科', '2308', '哈尔滨医科大学', '哈尔滨市南岗区保健路157号', '0N',
     '公共事业管理', None, 1, '5000',
     '认同四川省少数民族地区加分项目，但分值最高20分。', 8],
]

# 理科 Page 20 entire page missing
page20_data = [
    ['理科', '0365', '石河子大学', '新疆石河子市北四路221号', '4Z',
     '中药学', None, 1, '4000', '认同四川省少数民族地区加分项目，但分值最高20分。', 20],
    ['理科', '0365', '石河子大学', '新疆石河子市北四路221号', '52',
     '工商管理类', '包含专业:财务管理、工商管理、人力资源管理、市场营销', 2, '3200',
     '认同四川省少数民族地区加分项目，但分值最高20分。', 20],
    ['理科', '0365', '石河子大学', '新疆石河子市北四路221号', '53',
     '经济学类', '包含专业:金融学、经济学、国际经济与贸易', 4, '3200',
     '认同四川省少数民族地区加分项目，但分值最高20分。', 20],
    ['理科', '0365', '石河子大学', '新疆石河子市北四路221号', '5D',
     '社会工作', None, 2, '3100', '认同四川省少数民族地区加分项目，但分值最高20分。', 20],
    ['理科', '0365', '石河子大学', '新疆石河子市北四路221号', '5E',
     '应急管理', None, 1, '3100', '认同四川省少数民族地区加分项目，但分值最高20分。', 20],
    ['理科', '0371', '大连民族大学', '辽宁省大连开发区辽河西路18号', '37',
     '区块链工程', None, 1, '5200',
     '认同四川省少数民族地区加分项目，但分值最高20分。金石滩校区就读。', 20],
    ['理科', '9120', '海军军医大学', '上海市杨浦区翔殷路800号', '09',
     '护理学', '护师', 1, '5000',
     '认同四川省少数民族地区加分项目，但分值最高20分。招收无军籍考生。只招收普通高中应届毕业生，考生政治面貌为共青团员或中共党员。', 20],
    ['理科', '1104', '北京印刷学院', '北京市大兴区兴华大街(二段)1号', '14',
     '包装工程', '不招色盲、色弱', 2, '4600',
     '认同四川省少数民族地区加分项目，但分值最高20分。', 20],
    ['理科', '1104', '北京印刷学院', '北京市大兴区兴华大街(二段)1号', '22',
     '机械类', '包含专业:机械工程、自动化', 1, '4600',
     '认同四川省少数民族地区加分项目，但分值最高20分。', 20],
    ['理科', '1110', '北京第二外国语学院', '北京市朝阳区定福庄南里1号', '1D',
     '金融学', '资本市场与投融资；外语单科成绩不低于90分', 1, '4200',
     '认同四川省少数民族地区加分项目，但分值最高20分。', 20],
    ['理科', '1111', '北京物资学院', '北京市通州区富河大街321号', '15',
     '信息管理与信息系统', '数智化管理实验班', 1, '4200',
     '认同四川省少数民族地区加分项目，但分值最高20分。', 20],
    ['理科', '1114', '中国戏曲学院', '北京市丰台区万泉寺400号', '0N',
     '艺术管理', '国际文化交流；外语单科成绩达到110分', 1, '8000', '本校是实分投档。', 20],
    ['理科', '1121', '首都经济贸易大学', '北京市丰台区花乡张家路口121号', '14',
     '安全工程', '注安师；色盲色弱限报', 1, '5500',
     '认同四川省少数民族地区加分项目，但分值最高20分。', 20],
]

# 理科 4506 广西医科大学 (pages 41-42 cross-page)
lk_4506 = [
    ['理科', '4506', '广西医科大学', '广西壮族自治区南宁市青秀区双拥路22号', '12',
     '临床药学', None, 1, '6400',
     '认同四川省少数民族地区加分项目，但分值最高20分。不招色盲色弱；武鸣校区就读。', 42],
]

# 理科 5177 成都东软学院 (pages 46-47 cross-page)
lk_5177 = [
    ['理科', '5177', '成都东软学院', '四川省成都市都江堰青城山东软大道1号', '03',
     '软件工程', '民办院校', 16, '18000', None, 47],
]

# ===== PROCESS ALL ROWS =====
out_row = 2
stats = {'match': 0, 'corrected': 0, 'missing': 0}

all_pages = sorted(set(list(page_rows.keys()) + [20]))

for page in all_pages:
    rows = page_rows.get(page, [])

    # Insert page 20 missing data
    if page == 20:
        for mrow in page20_data:
            mn = mrow[6] or ''
            mjn, br = split_major_name(mrow[5])
            if br:
                mn = (br + '\uff1b' + mn) if mn else br
            vals = [mrow[0], mrow[1], mrow[2], mrow[3], mrow[4], mjn, mn,
                    mrow[7], str(mrow[8]) if mrow[8] else '', mrow[9] or '', mrow[10],
                    '缺漏行(整页OCR缺失page20)']
            write_row(out_ws, out_row, vals, font=data_font, fill=green_fill)
            out_row += 1
            stats['missing'] += 1
        continue

    for d in rows:
        cnotes = []
        kl = d['kl']
        code = d['code']
        name = d['name']
        addr = d['addr']
        mjcode = d['mjcode']
        mjname = d['mjname']
        mjnote = d['mjnote']
        plan = d['plan']
        fee = d['fee']
        yxnote = d['yxnote']

        # === Cross-page institution fix ===
        if page == 11 and not code and '工商管理' in mjname:
            code = '4106'
            name = '河南科技大学'
            addr = '河南省洛阳市洛龙区开元大道263号'
            cnotes.append('补全跨页丢失院校:4106河南科技大学')

        if page == 16 and not code and mjcode == 'H3':
            code = '0923'
            name = '东北林业大学'
            addr = '黑龙江省哈尔滨市香坊区和兴路26号'
            cnotes.append('补全跨页丢失院校:0923东北林业大学')

        if page == 24 and code == '0T' and name == '环境工程':
            kl = '理科'
            code = '2102'
            name = '沈阳航空航天大学'
            addr = '辽宁省沈阳市沈北新区道义南大街37号'
            mjcode = '0T'
            mjname = '环境工程'
            mjnote = ''
            yxnote = '认同四川省少数民族地区加分项目，但分值最高20分。'
            cnotes.append('修正列偏移→补全2102沈阳航空航天大学')

        if page == 50 and not code and mjcode in ['B1', 'D1', 'E5', 'G3', 'K1']:
            code = '6407'
            name = '宁夏医科大学'
            addr = '银川市兴庆区胜利街1160号'
            cnotes.append('补全跨页丢失院校:6407宁夏医科大学')

        # === [V] cleanup ===
        cleaned, had_v = clean_vmark(mjname)
        if had_v:
            mjname = cleaned
            cnotes.append('清除[V]标记')

        if '[V]' in mjcode:
            mjcode = mjcode.replace('[V]', '').replace(' ', '').strip()
            cnotes.append('清除专业代码中[V]')

        if mjnote and '[V]' in mjnote:
            mjnote = mjnote.replace('[V]', '').strip()
            cnotes.append('清除备注中[V]')

        # === Specific code corrections ===
        if str(code) == '5924' and kl == '文科' and str(mjcode) == '1Z' and '酒店' in mjname:
            mjcode = '12'
            cnotes.append('代码修正:1Z->12')

        if str(code) == '2308' and '生物信息' in mjname:
            raw_code = mjcode.replace('[V]', '').replace(' ', '').strip()
            if raw_code in ('01', '0I'):
                mjcode = '0J'
                cnotes.append('代码修正:01->0J')

        if str(code) == '5119' and str(mjcode) == '6' and '预防' in mjname:
            mjcode = '06'
            cnotes.append('代码修正:6->06')

        if str(code) == '5119' and str(mjcode) == '7' and '检验' in mjname:
            mjcode = '07'
            cnotes.append('代码修正:7->07')

        if str(code) == '5924' and kl == '理科' and str(mjcode) == '2Z' and '园艺' in mjname:
            mjcode = 'ZZ'
            cnotes.append('代码修正:2Z->ZZ')

        # Clean trailing [V] from code like "0E [V]"
        if ' ' in mjcode:
            parts = mjcode.split()
            mjcode = parts[0]
            if len(parts) > 1 and parts[1] == '[V]':
                if '清除专业代码中[V]' not in cnotes:
                    cnotes.append('清除专业代码中[V]')

        # === Fix specific data issues ===
        # R6: 0841 fee and yxnote missing
        if str(code) == '0841' and str(mjcode) == 'M0':
            if not fee:
                fee = '待定'
                cnotes.append('收费补全:待定')
            if not yxnote:
                yxnote = '只承认教育部规定的全国性加分政策。马来西亚分校就读。'
                cnotes.append('院校备注补全')

        # === Bracket split ===
        mjname_clean, bracket = split_major_name(mjname)
        if bracket:
            if mjnote:
                mjnote = bracket + '\uff1b' + mjnote
            else:
                mjnote = bracket
            mjname = mjname_clean
            cnotes.append('括号拆分→备注')

        # === Write row ===
        fill = yellow_fill if cnotes else None
        note = '\uff1b'.join(cnotes) if cnotes else ''
        fee_s = str(fee) if fee is not None else ''
        row_vals = [kl, code, name, addr, mjcode, mjname, mjnote, plan, fee_s,
                    yxnote, page, note]
        write_row(out_ws, out_row, row_vals, font=data_font, fill=fill)
        out_row += 1
        if cnotes:
            stats['corrected'] += 1
        else:
            stats['match'] += 1

    # Insert missing rows after current page's data
    if page == 7:
        for mrow in wk_2106:
            mn = mrow[6] or ''
            vals = [mrow[0], mrow[1], mrow[2], mrow[3], mrow[4], mrow[5], mn,
                    mrow[7], str(mrow[8]), mrow[9] or '', mrow[10],
                    '缺漏行(文科2106辽宁工程技术大学跨页丢失)']
            write_row(out_ws, out_row, vals, font=data_font, fill=green_fill)
            out_row += 1
            stats['missing'] += 1

    if page == 8:
        for mrow in wk_2308:
            mn = mrow[6] or ''
            vals = [mrow[0], mrow[1], mrow[2], mrow[3], mrow[4], mrow[5], mn,
                    mrow[7], str(mrow[8]), mrow[9] or '', mrow[10],
                    '缺漏行(文科2308哈尔滨医科大学跨页丢失)']
            write_row(out_ws, out_row, vals, font=data_font, fill=green_fill)
            out_row += 1
            stats['missing'] += 1

    if page == 42:
        for mrow in lk_4506:
            mn = mrow[6] or ''
            vals = [mrow[0], mrow[1], mrow[2], mrow[3], mrow[4], mrow[5], mn,
                    mrow[7], str(mrow[8]), mrow[9] or '', mrow[10],
                    '缺漏行(理科4506广西医科大学跨页丢失)']
            write_row(out_ws, out_row, vals, font=data_font, fill=green_fill)
            out_row += 1
            stats['missing'] += 1

    if page == 47:
        for mrow in lk_5177:
            mn = mrow[6] or ''
            vals = [mrow[0], mrow[1], mrow[2], mrow[3], mrow[4], mrow[5], mn,
                    mrow[7], str(mrow[8]) if mrow[8] else '', mrow[9] or '', mrow[10],
                    '缺漏行(理科5177成都东软学院跨页丢失)']
            write_row(out_ws, out_row, vals, font=data_font, fill=green_fill)
            out_row += 1
            stats['missing'] += 1

    # Save after each page
    out_wb.save(out_path)

# Final save
out_wb.save(out_path)

print(f"\n=== 批次3284校验完成 ===")
print(f"总行数: {out_row - 2}")
print(f"  匹配行: {stats['match']}")
print(f"  修正行(黄色): {stats['corrected']}")
print(f"  缺漏行(绿色): {stats['missing']}")
