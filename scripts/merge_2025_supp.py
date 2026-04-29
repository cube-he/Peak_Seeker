# -*- coding: utf-8 -*-
"""
Merge all 18 verified 2025 supplementary collection xlsx files into one
unified 24-column file.

Column mapping is header-name-based (not positional) to handle the 5
different source layouts (Type 3/4/8/9/10).
"""

import glob, re, os
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# ── Unified output columns (24) ──────────────────────────────────────────
UNIFIED_HEADERS = [
    "年份", "科目", "录取批次", "招生类型", "降分政策",
    "院校代码", "院校名称", "办学性质", "院校地址", "院校备注",
    "调档线", "专业组代码", "再选科目要求", "专业组计划数",
    "专业代码", "专业名称", "专业备注", "专业计划数", "收费标准",
    "页码", "征集次数", "来源网页", "批次ID", "校正备注",
]

# ── Source header → unified column index mapping ─────────────────────────
# Supports alternate names found in different source files
HEADER_TO_COL = {
    "年份":       0,
    "科目":       1,  "科类": 1,
    "录取批次":   2,
    "招生类型":   3,  "招生类别": 3,
    "降分政策":   4,
    "院校代码":   5,
    "院校名称":   6,
    "办学性质":   7,
    "院校地址":   8,
    "院校备注":   9,
    "调档线":    10,
    "专业组代码": 11,
    "再选科目要求": 12,
    "专业组计划数": 13,
    "专业代码":  14,
    "专业名称":  15,
    "专业备注":  16,
    "专业计划数": 17,
    "收费标准":  18,
    "页码":      19,
    "征集次数":  20,
    "来源网页":  21,
    # col 22 = 批次ID (filled programmatically)
    "校正备注":  23,
}


def extract_batch_id(filepath: str) -> str:
    """Extract the leading numeric batch ID (e.g. '4402') from the filename."""
    basename = Path(filepath).name
    m = re.match(r"(\d+)", basename)
    return m.group(1) if m else ""


def build_col_mapping(header_row: list) -> dict:
    """
    Given a source file's header row, return {source_col_idx: unified_col_idx}.
    """
    mapping = {}
    for src_idx, name in enumerate(header_row):
        if name is None:
            continue
        name = str(name).strip()
        if name in HEADER_TO_COL:
            mapping[src_idx] = HEADER_TO_COL[name]
    return mapping


def main():
    base = Path("C:/Users/Administrator/Documents/VolunteerHelper/data/13_征集志愿/普通高考")
    pattern = str(base / "**" / "*2025*已校验.xlsx")
    src_files = sorted(glob.glob(pattern, recursive=True))

    # Exclude backup copies
    src_files = [f for f in src_files if "原始备份" not in f]

    print(f"Found {len(src_files)} source files")
    assert len(src_files) == 18, f"Expected 18 files, got {len(src_files)}"

    all_rows = []

    for fpath in src_files:
        batch_id = extract_batch_id(fpath)
        wb = load_workbook(fpath, read_only=True, data_only=True)
        ws = wb.active

        rows_iter = ws.iter_rows()
        header_cells = next(rows_iter)
        header_row = [c.value for c in header_cells]
        col_map = build_col_mapping(header_row)

        has_year_col = any(str(h).strip() == "年份" for h in header_row if h)

        file_row_count = 0
        for row in rows_iter:
            src_values = [c.value for c in row]
            # Skip completely empty rows
            if all(v is None or str(v).strip() == "" for v in src_values):
                continue

            unified = [None] * 24
            for src_idx, uni_idx in col_map.items():
                if src_idx < len(src_values):
                    unified[uni_idx] = src_values[src_idx]

            # Fill year if not present in source
            if not has_year_col or unified[0] is None:
                unified[0] = "2025"

            # Fill batch ID
            unified[22] = batch_id

            all_rows.append(unified)
            file_row_count += 1

        wb.close()
        print(f"  {batch_id}: {file_row_count} rows  ({Path(fpath).name[:60]})")

    # Sort by batch ID (col 22), then preserve original order within each batch
    all_rows.sort(key=lambda r: int(r[22]) if r[22] and str(r[22]).isdigit() else 0)

    print(f"\nTotal data rows: {len(all_rows)}")

    # ── Write output workbook ────────────────────────────────────────────
    out_path = base / "征集志愿_2025_合并.xlsx"
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "征集志愿"

    header_font = Font(name="微软雅黑", size=10, bold=True)
    data_font = Font(name="微软雅黑", size=10)

    # Write header
    for col_idx, name in enumerate(UNIFIED_HEADERS, start=1):
        cell = ws_out.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Write data
    for row_idx, row_data in enumerate(all_rows, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws_out.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font

    # Freeze first row
    ws_out.freeze_panes = "A2"

    # Auto-width (approximate)
    for col_idx in range(1, 25):
        col_letter = get_column_letter(col_idx)
        ws_out.column_dimensions[col_letter].width = max(
            len(UNIFIED_HEADERS[col_idx - 1]) * 2 + 2, 10
        )

    wb_out.save(str(out_path))
    print(f"\nSaved to: {out_path}")
    print(f"Total rows (incl header): {len(all_rows) + 1}")


if __name__ == "__main__":
    main()
