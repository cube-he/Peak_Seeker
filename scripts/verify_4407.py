"""
校验批次 4407: 2025年 物历综合 本科批次 A段地方专项第二次+B段第一次 征集志愿
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import sys, re

sys.stdout.reconfigure(encoding='utf-8')

BASE = 'C:/Users/Administrator/Documents/VolunteerHelper'
src_path = f'{BASE}/data/13_征集志愿/普通高考/专项计划/4407_2025_物历综合_本科批次_A段地方专项第二次+B段第一次_征集志愿/4407_2025_物历综合_本科批次_A段地方专项第二次+B段第一次_征集志愿_mimo-v2-omni.xlsx'
out_path = f'{BASE}/data/13_征集志愿/普通高考/专项计划/4407_2025_物历综合_本科批次_A段地方专项第二次+B段第一次_征集志愿/4407_2025_物历综合_本科批次_A段地方专项第二次+B段第一次_征集志愿_已校验.xlsx'

wb_src = openpyxl.load_workbook(src_path)
ws_src = wb_src.active

wb_out = openpyxl.Workbook()
ws_out = wb_out.active
ws_out.title = '征集志愿'

font_normal = Font(name='微软雅黑', size=10)
font_bold = Font(name='微软雅黑', size=10, bold=True)
fill_yellow = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
fill_green = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
fill_red = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')

headers = ['科类', '招生类型', '院校代码', '院校名称', '办学性质', '院校地址', '院校备注',
           '专业组代码', '再选科目要求', '专业组计划数', '专业代码', '专业名称', '专业备注',
           '专业计划数', '收费标准', '页码', '校正备注']
for c, h in enumerate(headers, 1):
    cell = ws_out.cell(1, c, h)
    cell.font = font_bold


def split_major_name(name, existing_note):
    """Split bracket content from major name into note field."""
    if not name:
        return name, existing_note
    clean_name = name
    if clean_name.startswith('[V]'):
        clean_name = clean_name[3:]

    idx_half = clean_name.find('(')
    idx_full = clean_name.find('\uff08')  # full-width (

    if idx_half == -1 and idx_full == -1:
        return clean_name, existing_note

    if idx_half == -1:
        idx = idx_full
    elif idx_full == -1:
        idx = idx_half
    else:
        idx = min(idx_half, idx_full)

    base_name = clean_name[:idx].strip()
    bracket_content = clean_name[idx:]

    # Parse all bracket groups
    parts = []
    remaining = bracket_content
    while remaining:
        m = re.match(r'[\uff08(]([^)\uff09]*)[\uff09)](.*)', remaining)
        if m:
            parts.append(m.group(1))
            remaining = m.group(2).strip()
        else:
            # Handle unmatched bracket
            if remaining.startswith('(') or remaining.startswith('\uff08'):
                inner = remaining.strip('()\uff08\uff09')
                if inner:
                    parts.append(inner)
            break

    sep = '\uff1b'  # full-width semicolon
    note_from_brackets = sep.join(parts) if parts else bracket_content.strip('()\uff08\uff09')

    if existing_note:
        final_note = note_from_brackets + sep + existing_note
    else:
        final_note = note_from_brackets

    return base_name, final_note


out_row = 2
corrections_count = 0
type_fix_count = 0
v_fix_count = 0
split_count = 0
crosspage_fix_count = 0

# Track last known school info for cross-page inheritance
last_school = {
    '院校代码': None, '院校名称': None, '办学性质': None,
    '院校地址': None, '院校备注': None
}

for r in range(2, ws_src.max_row + 1):
    kl = ws_src.cell(r, 1).value
    zs = ws_src.cell(r, 2).value
    code = ws_src.cell(r, 3).value
    name_val = ws_src.cell(r, 4).value
    bx = ws_src.cell(r, 5).value
    addr = ws_src.cell(r, 6).value
    ybz = ws_src.cell(r, 7).value
    zyzdm = ws_src.cell(r, 8).value
    zxkm = ws_src.cell(r, 9).value
    zyjhs = ws_src.cell(r, 10).value
    zydm = ws_src.cell(r, 11).value
    zymc = ws_src.cell(r, 12).value
    zybz = ws_src.cell(r, 13).value
    zyjh = ws_src.cell(r, 14).value
    sfbz = ws_src.cell(r, 15).value
    page = ws_src.cell(r, 16).value

    corrections = []

    # --- Fix 0: Cross-page null 院校代码 inheritance ---
    if code is not None:
        last_school['院校代码'] = code
        last_school['院校名称'] = name_val
        last_school['办学性质'] = bx
        last_school['院校地址'] = addr
        last_school['院校备注'] = ybz
    else:
        # Fill from last known school
        code = last_school['院校代码']
        name_val = last_school['院校名称']
        bx = last_school['办学性质']
        addr = last_school['院校地址']
        ybz = last_school['院校备注']
        corrections.append(f'跨页补全院校:{code} {name_val}')
        crosspage_fix_count += 1

    # --- Fix 1: 招生类型 normalization ---
    original_zs = zs
    if page in [4, 5]:
        if zs != '地方专项计划':
            zs = '地方专项计划'
            corrections.append(f'招生类型:{original_zs}\u2192地方专项计划')
            type_fix_count += 1
    elif page and page >= 6:
        code_str = str(code) if code else ''
        # 6595 on page 124 was mis-classified
        if page == 124 and code_str == '6595' and zs == '本科层次职业教育人才培养改革试点':
            zs = '普通类本科'
            corrections.append(f'招生类型:{original_zs}\u2192普通类本科')
            type_fix_count += 1
        elif zs in ['本科层次职业教育人才培养改革试点', '民族班', '非西藏生源定向西藏就业']:
            pass  # correct special types
        elif zs == '普通类' or zs is None:
            zs = '普通类本科'
            corrections.append(f'招生类型:{original_zs}\u2192普通类本科')
            type_fix_count += 1

    # --- Fix 2: [V] removal ---
    if zymc and '[V]' in str(zymc):
        zymc = str(zymc).replace('[V]', '')
        corrections.append('专业名称去[V]')
        v_fix_count += 1

    # --- Fix 3: Split brackets from 专业名称 ---
    if zymc and ('(' in str(zymc) or '\uff08' in str(zymc)):
        original_zymc = str(zymc)
        zymc, zybz = split_major_name(str(zymc), zybz)
        if zymc != original_zymc:
            corrections.append('专业名称拆分')
            split_count += 1

    # --- Fix 4: String formatting ---
    if code is not None:
        code = str(code).zfill(4)
    if zydm is not None:
        zydm = str(zydm)
        if zydm.isdigit() and len(zydm) < 2:
            zydm = zydm.zfill(2)
    if zyzdm is not None:
        zyzdm = str(zyzdm)
    if sfbz is not None:
        sfbz = str(sfbz)

    # Write row
    values = [kl, zs, code, name_val, bx, addr, ybz, zyzdm, zxkm, zyjhs,
              zydm, zymc, zybz, zyjh, sfbz, page]
    for c, val in enumerate(values, 1):
        cell = ws_out.cell(out_row, c, val)
        cell.font = font_normal

    if corrections:
        ws_out.cell(out_row, 17, '\uff1b'.join(corrections)).font = font_normal
        for corr in corrections:
            if '招生类型' in corr:
                ws_out.cell(out_row, 2).fill = fill_yellow
            if '专业名称' in corr:
                ws_out.cell(out_row, 12).fill = fill_yellow
                if zybz:
                    ws_out.cell(out_row, 13).fill = fill_yellow
            if '跨页补全院校' in corr:
                ws_out.cell(out_row, 3).fill = fill_yellow
                ws_out.cell(out_row, 4).fill = fill_yellow
                ws_out.cell(out_row, 5).fill = fill_yellow
                ws_out.cell(out_row, 6).fill = fill_yellow
        corrections_count += len(corrections)

    out_row += 1

# ===== MISSING DATA: Page 45 (OCR skipped entire page) =====
missing_p45 = [
    ['物理类', '普通类本科', '1161', '民政职业大学', '公办', '北京市',
     '执行部委属和外省属高校少数民族地区加分办法1。本科层次职业教育。',
     '101', '不限', 2, '0S', '康复辅助器具技术', '燕郊校区', 2, '6000', 45],
    ['物理类', '普通类本科', '6503', '塔里木大学', '公办', '新疆维吾尔自治区阿拉尔市',
     '执行部委属和外省属高校少数民族地区加分办法1。部分专业对考生身体健康状况（色弱、色盲、单色识别不全等）、外语语种要求等有限制，具体可查阅学校招生章程。',
     '102', '化学', 7, '38', '物理学', '师范', 6, '3500', 45],
    ['物理类', '普通类本科', '6503', '塔里木大学', '公办', '新疆维吾尔自治区阿拉尔市',
     '执行部委属和外省属高校少数民族地区加分办法1。部分专业对考生身体健康状况（色弱、色盲、单色识别不全等）、外语语种要求等有限制，具体可查阅学校招生章程。',
     '102', '化学', 7, '49', '环境工程', None, 1, '3500', 45],
    ['物理类', '普通类本科', '9211', '陆军军医大学', '公办', '重庆市',
     '执行部委属和外省属高校少数民族地区加分办法1。',
     '101', '化学', 1, '03', '预防医学', '无军籍地方生，仅招收普通高中应届毕业生，政治面貌须为共青团员或中共党员。', 1, '4500', 45],
    ['物理类', '普通类本科', '1181', '首都师范大学科德学院', '独立学院', '北京市',
     '执行部委属和外省属高校少数民族地区加分办法1。',
     '101', '化学', 1, '01', '计算机科学与技术', None, 1, '79800', 45],
    ['物理类', '普通类本科', '1182', '北京金融科技学院', '民办院校', '北京市',
     '执行部委属和外省属高校少数民族地区加分办法1。',
     '101', '不限', 2, '02', '金融学', '具体培养模式、收费标准等请查看院校招生章程并咨询院校。', 2, '79800', 45],
    ['物理类', '普通类本科', '1183', '北京邮电大学世纪学院', '独立学院', '北京市',
     '执行部委属和外省属高校少数民族地区加分办法1。',
     '101', '不限', 2, '10', '财务管理', None, 1, '39000', 45],
    ['物理类', '普通类本科', '1183', '北京邮电大学世纪学院', '独立学院', '北京市',
     '执行部委属和外省属高校少数民族地区加分办法1。',
     '101', '不限', 2, '11', '电子商务', None, 1, '34000', 45],
    ['物理类', '普通类本科', '1183', '北京邮电大学世纪学院', '独立学院', '北京市',
     '执行部委属和外省属高校少数民族地区加分办法1。',
     '102', '化学', 2, '02', '通信工程', None, 1, '41000', 45],
    ['物理类', '普通类本科', '1183', '北京邮电大学世纪学院', '独立学院', '北京市',
     '执行部委属和外省属高校少数民族地区加分办法1。',
     '102', '化学', 2, '04', '计算机科学与技术', None, 1, '37000', 45],
    ['物理类', '普通类本科', '1184', '北京工业大学耿丹学院', '独立学院', '北京市',
     '执行部委属和外省属高校少数民族地区加分办法1。',
     '102', '化学', 4, '05', '机械设计制造及其自动化', '机器人方向', 1, '46600', 45],
    ['物理类', '普通类本科', '1184', '北京工业大学耿丹学院', '独立学院', '北京市',
     '执行部委属和外省属高校少数民族地区加分办法1。',
     '102', '化学', 4, '10', '电子信息类', '包含专业:电子信息工程、通信工程', 1, '46600', 45],
    ['物理类', '普通类本科', '1184', '北京工业大学耿丹学院', '独立学院', '北京市',
     '执行部委属和外省属高校少数民族地区加分办法1。',
     '102', '化学', 4, '11', '数字媒体技术', None, 1, '46600', 45],
    ['物理类', '普通类本科', '1184', '北京工业大学耿丹学院', '独立学院', '北京市',
     '执行部委属和外省属高校少数民族地区加分办法1。',
     '102', '化学', 4, '22', '计算机科学与技术', None, 1, '46600', 45],
    ['物理类', '普通类本科', '1185', '北京第二外国语学院中瑞酒店管理学院', '独立学院', '北京市',
     '执行部委属和外省属高校少数民族地区加分办法1。',
     '101', '不限', 5, '04', '财务管理', None, 5, '49800', 45],
]

for row_data in missing_p45:
    for c, val in enumerate(row_data, 1):
        cell = ws_out.cell(out_row, c, val)
        cell.font = font_normal
        cell.fill = fill_green
    ws_out.cell(out_row, 17, '缺漏行：OCR遗漏整页(页45)').font = font_normal
    ws_out.cell(out_row, 17).fill = fill_green
    out_row += 1

# ===== MISSING DATA: Page 8 (1294 continuation from page 7) =====
missing_p8 = [
    ['历史类', '普通类本科', '1294', '天津传媒学院', '民办院校', '天津市',
     '执行部委属和外省属高校少数民族地区加分办法1。',
     '101', '不限', 42, '17', '新闻学', None, 9, '30000', 8],
    ['历史类', '普通类本科', '1294', '天津传媒学院', '民办院校', '天津市',
     '执行部委属和外省属高校少数民族地区加分办法1。',
     '101', '不限', 42, '18', '文化产业管理', None, 8, '24000', 8],
    ['历史类', '普通类本科', '1294', '天津传媒学院', '民办院校', '天津市',
     '执行部委属和外省属高校少数民族地区加分办法1。',
     '101', '不限', 42, '20', '学前教育', None, 10, '24000', 8],
    ['历史类', '普通类本科', '1294', '天津传媒学院', '民办院校', '天津市',
     '执行部委属和外省属高校少数民族地区加分办法1。',
     '101', '不限', 42, '29', '广告学', None, 6, '33000', 8],
    ['历史类', '普通类本科', '1294', '天津传媒学院', '民办院校', '天津市',
     '执行部委属和外省属高校少数民族地区加分办法1。',
     '102', '不限', 2, '28', '英语', '招收英语语种考生', 2, '24000', 8],
]

for row_data in missing_p8:
    for c, val in enumerate(row_data, 1):
        cell = ws_out.cell(out_row, c, val)
        cell.font = font_normal
        cell.fill = fill_green
    ws_out.cell(out_row, 17, '缺漏行：OCR遗漏跨页续写(1294天津传媒学院)').font = font_normal
    ws_out.cell(out_row, 17).fill = fill_green
    out_row += 1

wb_out.save(out_path)
print(f'Output saved: {out_row - 1} total rows (incl header)')
print(f'  招生类型 corrections: {type_fix_count}')
print(f'  [V] removals: {v_fix_count}')
print(f'  专业名称 splits: {split_count}')
print(f'  跨页院校补全: {crosspage_fix_count}')
print(f'  Total corrections: {corrections_count}')
print(f'  Missing rows added (green): {len(missing_p45) + len(missing_p8)}')
