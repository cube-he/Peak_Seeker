"""
校验批次 4413: 物历综合_本科批次_区域教育均衡+省属高校少数民族预科_征集志愿_第一次
将原始图片人工识别结果与 OCR xlsx 逐行比对，生成校验后的 xlsx
"""
import openpyxl
from openpyxl.styles import PatternFill
from copy import copy

# ── 文件路径 ──
BASE = "C:/Users/Administrator/Documents/VolunteerHelper/data/13_征集志愿/普通高考/少数民族/4413_2025_物历综合_本科批次_区域教育均衡+省属高校少数民族预科_征集志愿_第一次"
OCR_FILE = f"{BASE}/4413_2025_物历综合_本科批次_区域教育均衡+省属高校少数民族预科_征集志愿_第一次_mimo-v2-omni.xlsx"
OUT_FILE = f"{BASE}/4413_2025_物历综合_本科批次_区域教育均衡+省属高校少数民族预科_征集志愿_第一次_已校验.xlsx"

HEADERS = ['科类', '招生类型', '院校代码', '院校名称', '办学性质', '院校地址', '院校备注',
           '专业组代码', '再选科目要求', '专业组计划数', '专业代码', '专业名称', '专业备注',
           '专业计划数', '收费标准', '页码', '校正备注']

YELLOW = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
GREEN = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
RED = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")

# ── 从图片人工识别的真实数据 (ground truth) ──
# 表头: 科类, 招生类型, 院校代码, 院校名称, 办学性质, 院校地址, 院校备注,
#       专业组代码, 再选科目要求, 专业组计划数, 专业代码, 专业名称, 专业备注,
#       专业计划数, 收费标准, 页码

gt_rows = [
    # ── 历史类 / 省属高校少数民族预科 ── (第4页)
    ["历史类", "省属高校少数民族预科", "5137", "成都大学", "公办", "四川省成都市", None,
     "108", "不限", 1, "Y5", "少数民族预科(旅游管理)", "预科培养地点为西昌学院。",
     1, 5500, 4],
    ["历史类", "省属高校少数民族预科", "5137", "成都大学", "公办", "四川省成都市", None,
     "109", "化学", 2, "Y8", "少数民族预科(护理学)", "预科培养地点为西昌学院。",
     2, 5500, 4],

    # ── 物理类 / 1. 区域教育均衡发展专项计划 ── (第4-5页)
    ["物理类", "区域教育均衡发展专项计划", "5108", "四川轻化工大学", "公办", "四川省自贡市", None,
     "812", "化学", 1, "8T", "食品科学与工程(雅安市)", None,
     1, 5980, 4],
    ["物理类", "区域教育均衡发展专项计划", "5119", "成都医学院", "公办", "四川省成都市",
     "色盲、色弱考生不予录取。新都校区就读。",
     "103", "化学", 2, "J8", "护理学(凉山州)", None,
     1, 6670, 4],
    ["物理类", "区域教育均衡发展专项计划", "5119", "成都医学院", "公办", "四川省成都市",
     "色盲、色弱考生不予录取。新都校区就读。",
     "103", "化学", 2, "JD", "智能医学工程(凉山州)", None,
     1, 6670, 4],
    ["物理类", "区域教育均衡发展专项计划", "5123", "绵阳师范学院", "公办", "四川省绵阳市", None,
     "108", "不限", 1, "QH", "物流管理(雅安市)", None,
     1, 5520, 4],
    # 5125 院校头在第4页底部，专业在第5页
    ["物理类", "区域教育均衡发展专项计划", "5125", "宜宾学院", "公办", "四川省宜宾市", None,
     "107", "化学", 1, "5D", "制药工程(雅安市)(江北校区A区)", None,
     1, 5980, 5],
    ["物理类", "区域教育均衡发展专项计划", "5130", "阿坝师范学院", "公办", "四川省阿坝州", None,
     "104", "不限", 1, "2G", "学前教育(雅安市)(师范)", None,
     1, 4800, 5],
    ["物理类", "区域教育均衡发展专项计划", "5137", "成都大学", "公办", "四川省成都市", None,
     "105", "化学", 1, "37", "环境工程(凉山州)", None,
     1, 5980, 5],
    ["物理类", "区域教育均衡发展专项计划", "5142", "成都工业学院", "公办", "四川省成都市", None,
     "146", "化学", 1, "47", "汽车服务工程(乐山市)(宜宾校区)", "入学后不允许跨校区转专业。",
     1, 5200, 5],
    ["物理类", "区域教育均衡发展专项计划", "5142", "成都工业学院", "公办", "四川省成都市", None,
     "147", "化学", 3, "48", "汽车服务工程(雅安市)(宜宾校区)", "入学后不允许跨校区转专业。",
     3, 5200, 5],

    # ── 物理类 / 2. 省属高校少数民族预科 ── (第5页)
    ["物理类", "省属高校少数民族预科", "5119", "成都医学院", "公办", "四川省成都市", None,
     "109", "不限", 2, "Y9", "少数民族预科(应用心理学)",
     "预科培养地点为四川民族学院。色盲、色弱考生不予录取。",
     2, 5980, 5],
    ["物理类", "省属高校少数民族预科", "5119", "成都医学院", "公办", "四川省成都市", None,
     "110", "化学", 2, "Y8", "少数民族预科(护理学)",
     "预科培养地点为四川民族学院。色盲、色弱考生不予录取。",
     2, 6670, 5],
    ["物理类", "省属高校少数民族预科", "5142", "成都工业学院", "公办", "四川省成都市", None,
     "130", "化学", 34, "Y1", "少数民族预科(预科培养地点为阿坝师范学院。)", "宜宾校区就读。",
     34, 5500, 5],
]

# ── 读取 OCR xlsx ──
wb_ocr = openpyxl.load_workbook(OCR_FILE)
ws_ocr = wb_ocr.active
ocr_rows = []
for row in ws_ocr.iter_rows(min_row=2, values_only=True):
    ocr_rows.append(list(row))

print(f"OCR rows: {len(ocr_rows)}")
print(f"GT rows:  {len(gt_rows)}")

# ── 构建 key 索引 ──
# key = (院校代码, 专业组代码, 专业代码) — 但需要处理科类差异（同一院校可能在历史类和物理类都出现）
# 使用 (科类, 院校代码, 专业组代码, 专业代码) 作为更精确的 key

def make_key(row):
    """(科类, 院校代码, 专业组代码, 专业代码)"""
    return (str(row[0] or ''), str(row[2] or ''), str(row[7] or ''), str(row[10] or ''))

gt_dict = {}
for r in gt_rows:
    k = make_key(r)
    gt_dict[k] = r

ocr_dict = {}
for r in ocr_rows:
    k = make_key(r)
    ocr_dict[k] = r

gt_keys = set(gt_dict.keys())
ocr_keys = set(ocr_dict.keys())

matched = gt_keys & ocr_keys
missing_in_ocr = gt_keys - ocr_keys  # 图片有但 OCR 没有 → 缺漏
extra_in_ocr = ocr_keys - gt_keys    # OCR 有但图片没有 → 多余

print(f"\nMatched:        {len(matched)}")
print(f"Missing in OCR: {len(missing_in_ocr)}")
print(f"Extra in OCR:   {len(extra_in_ocr)}")

if missing_in_ocr:
    print("\n缺漏的 key:")
    for k in sorted(missing_in_ocr):
        print(f"  {k}")

if extra_in_ocr:
    print("\n多余的 key:")
    for k in sorted(extra_in_ocr):
        print(f"  {k}")

# ── 生成校验后的 xlsx ──
wb_out = openpyxl.Workbook()
ws_out = wb_out.active
ws_out.title = "校验结果"

# 写表头
for col_idx, h in enumerate(HEADERS, 1):
    ws_out.cell(row=1, column=col_idx, value=h)

field_names = HEADERS[:16]  # 不含校正备注

row_num = 2

# 按 GT 原始顺序输出所有行（匹配行+缺漏行），多余行追加在最后
for gt in gt_rows:
    key = make_key(gt)
    ocr = ocr_dict.get(key)

    if ocr is None:
        # 缺漏行 → 绿色
        for col_idx in range(16):
            cell = ws_out.cell(row=row_num, column=col_idx + 1, value=gt[col_idx])
            cell.fill = GREEN
        note_cell = ws_out.cell(row=row_num, column=17, value="缺漏补录")
        note_cell.fill = GREEN
    else:
        # 匹配行 → 逐字段比对
        corrections = []
        for col_idx in range(16):
            gt_val = gt[col_idx]
            ocr_val = ocr[col_idx]
            gt_str = str(gt_val) if gt_val is not None else ''
            ocr_str = str(ocr_val) if ocr_val is not None else ''

            # 数值比较
            try:
                if gt_str and ocr_str and float(gt_str) == float(ocr_str):
                    ws_out.cell(row=row_num, column=col_idx + 1, value=gt_val)
                    continue
            except (ValueError, TypeError):
                pass

            if gt_str == ocr_str:
                ws_out.cell(row=row_num, column=col_idx + 1, value=gt_val)
            else:
                cell = ws_out.cell(row=row_num, column=col_idx + 1, value=gt_val)
                cell.fill = YELLOW
                corrections.append(f"{field_names[col_idx]}:原值[{ocr_str}]→正确值[{gt_str}]")

        if corrections:
            ws_out.cell(row=row_num, column=17, value="; ".join(corrections))

    row_num += 1

# 多余行（OCR 有但图片无）→ 红色，追加在最后
for key in sorted(extra_in_ocr):
    ocr = ocr_dict[key]
    for col_idx in range(16):
        cell = ws_out.cell(row=row_num, column=col_idx + 1, value=ocr[col_idx])
        cell.fill = RED
    note_cell = ws_out.cell(row=row_num, column=17, value="多余行")
    note_cell.fill = RED
    row_num += 1

# 调整列宽
for col_idx in range(1, 18):
    ws_out.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 18

wb_out.save(OUT_FILE)
print(f"\n校验完成，输出: {OUT_FILE}")
print(f"总行数: {row_num - 2}")

# ── 打印差异汇总 ──
print("\n=== 差异汇总 ===")
for key in sorted(matched):
    gt = gt_dict[key]
    ocr = ocr_dict[key]
    diffs = []
    for col_idx in range(16):
        gt_val = gt[col_idx]
        ocr_val = ocr[col_idx]
        gt_str = str(gt_val) if gt_val is not None else ''
        ocr_str = str(ocr_val) if ocr_val is not None else ''
        try:
            if gt_str and ocr_str and float(gt_str) == float(ocr_str):
                continue
        except (ValueError, TypeError):
            pass
        if gt_str != ocr_str:
            diffs.append(f"  {field_names[col_idx]}: [{ocr_str}] → [{gt_str}]")
    if diffs:
        print(f"\nKey {key}:")
        for d in diffs:
            print(d)
