# -*- coding: utf-8 -*-
"""校验批次 3792: 2024年 文理综合 本科一批 征集志愿 第一次"""
import sys, io, json, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ====== Load OCR data ======
input_path = 'data/13_征集志愿/普通高考/本科批次/3792_2024_文理综合_本科一批_征集志愿_第一次/3792_2024_文理综合_本科一批_征集志愿_第一次_mimo-v2-omni.xlsx'
output_path = 'data/13_征集志愿/普通高考/本科批次/3792_2024_文理综合_本科一批_征集志愿_第一次/3792_2024_文理综合_本科一批_征集志愿_第一次_已校验.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

ocr_rows = []
for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), start=2):
    vals = list(row)
    ocr_rows.append({
        'xlsx_row': i,
        'kl': vals[0],
        'yxdm': str(vals[1]) if vals[1] else None,
        'yxmc': vals[2],
        'yxdz': vals[3],
        'zydm': str(vals[4]) if vals[4] else None,
        'zymc': vals[5],
        'zybz': vals[6],
        'zyjhs': vals[7],
        'sfbz': vals[8],
        'yxbz': vals[9],
        'page': int(vals[10]) - 1 if vals[10] else None,
    })

# ====== Styles ======
font_normal = Font(name='微软雅黑', size=10)
font_bold = Font(name='微软雅黑', size=10, bold=True)
fill_yellow = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
fill_green = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
fill_red = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')

# ====== Create output ======
wb_out = openpyxl.Workbook()
ws_out = wb_out.active
ws_out.title = '征集志愿'

headers = ['科类', '院校代码', '院校名称', '院校地址', '专业代码', '专业名称',
           '专业备注', '专业计划数', '收费标准', '院校备注', '页码', '校正备注']
for col, h in enumerate(headers, 1):
    cell = ws_out.cell(row=1, column=col, value=h)
    cell.font = font_bold

ws_out.column_dimensions['A'].width = 8
ws_out.column_dimensions['B'].width = 10
ws_out.column_dimensions['C'].width = 30
ws_out.column_dimensions['D'].width = 40
ws_out.column_dimensions['E'].width = 10
ws_out.column_dimensions['F'].width = 30
ws_out.column_dimensions['G'].width = 30
ws_out.column_dimensions['H'].width = 10
ws_out.column_dimensions['I'].width = 10
ws_out.column_dimensions['J'].width = 45
ws_out.column_dimensions['K'].width = 8
ws_out.column_dimensions['L'].width = 50


# ====== Helper functions ======
def split_zymc(zymc):
    """拆分专业名称 — 第一个(之前是名称，之后所有括号内容拆到备注"""
    if not zymc:
        return zymc, None
    idx = -1
    for i, c in enumerate(zymc):
        if c in ('(', '\uff08'):
            idx = i
            break
    if idx == -1:
        return zymc, None
    name = zymc[:idx]
    rest = zymc[idx:]
    notes = []
    depth = 0
    current = ''
    for c in rest:
        if c in ('(', '\uff08'):
            depth += 1
            if depth == 1:
                current = ''
                continue
        elif c in (')', '\uff09'):
            depth -= 1
            if depth == 0:
                notes.append(current)
                current = ''
                continue
        if depth > 0:
            current += c
    return name.strip(), '\uff1b'.join(notes) if notes else None


def write_row(ws, row_num, data, fill=None):
    for col, val in enumerate(data, 1):
        cell = ws.cell(row=row_num, column=col, value=val)
        cell.font = font_normal
        if fill:
            cell.fill = fill


# ====== Corrections ======

# Row corrections: fix rows with None school info (cross-page data loss)
row_corrections = {
    98: {'yxdm': '5017', 'yxmc': '西南政法大学',
         'yxdz': '重庆市渝北区宝圣大道301号',
         'note': '院校代码/名称/地址补全（跨页丢失）'},
    175: {'yxdm': '0856', 'yxmc': '西南大学',
          'yxdz': '重庆市北碚区',
          'zydm': '9C',
          'note': '院校代码/名称补全（跨页丢失）'},
    227: {'yxdm': '1209', 'yxmc': '天津外国语大学',
          'yxdz': '天津市河西区马场道117号',
          'note': '院校代码/名称/地址补全（跨页丢失）'},
    228: {'yxdm': '1209', 'yxmc': '天津外国语大学',
          'yxdz': '天津市河西区马场道117号',
          'note': '院校代码/名称/地址补全（跨页丢失）'},
    229: {'yxdm': '1209', 'yxmc': '天津外国语大学',
          'yxdz': '天津市河西区马场道117号',
          'note': '院校代码/名称/地址补全（跨页丢失）'},
}

# School name suffix fixes
school_name_fixes = {
    7: '西南民族大学(含增加计划)',
    8: '西南民族大学(含增加计划)',
    57: '南昌大学(含增加计划)',
    117: '四川师范大学(含增加计划)',
    118: '四川师范大学(含增加计划)',
    121: '西华师范大学(含增加计划)',
    122: '西华师范大学(含增加计划)',
    123: '西华师范大学(含增加计划)',
    124: '西华师范大学(含增加计划)',
    127: '成都锦城学院(民办院校)',
    129: '西安科技大学(含增加计划)',
    130: '西安科技大学(含增加计划)',
    144: '大连理工大学(盘锦校区)',
    163: '中国石油大学（北京）克拉玛依校区(含增加计划)',
    176: '西南大学(荣昌校区)',
    177: '西南大学(荣昌校区)',
    180: '华北电力大学(保定)',
    191: '中国民用航空飞行学院(含增加计划)',
    192: '西南民族大学(含增加计划)',
    193: '西南民族大学(含增加计划)',
    194: '西南民族大学(含增加计划)',
    195: '西南民族大学(含增加计划)',
}

# Code fixes
code_fixes = {
    9: {'zydm': '0S'},    # 1110 北京第二外国语学院: 05 -> 0S
    36: {'zydm': '0G'},   # 3239 南京审计大学: 06 -> 0G
}

# Missing schools (to add as green rows)
# (kl, yxdm, yxmc, yxdz, zydm, zymc, zybz, zyjhs, sfbz, yxbz, page)
missing_schools = [
    ('文科', '1211', '天津财经大学', '天津市河西区珠江道25号',
     '0B', '商务英语', '招收英语语种考生', 1, '5400',
     '认同并执行四川省少数民族加分项目和分值。', 6),
    ('文科', '2120', '大连外国语大学', '大连市旅顺口区旅顺南路西段6号',
     '0M', '朝鲜语', '外语口试；建议非朝鲜族考生报考', 1, '8000',
     '认同并执行四川省少数民族加分项目和分值。', 7),
    ('文科', '2120', '大连外国语大学', '大连市旅顺口区旅顺南路西段6号',
     '1D', '旅游管理', None, 1, '8000',
     '认同并执行四川省少数民族加分项目和分值。', 7),
    ('文科', '3111', '上海海洋大学', '上海市浦东新区临港新城沪城环路999号',
     '29', '日语', None, 1, '6500',
     '认同并执行四川省少数民族加分项目和分值。', 8),
    ('文科', '3341', '浙江外国语学院', '杭州市西湖区留和路299号',
     '07', '朝鲜语', None, 2, '4800',
     '认同并执行四川省少数民族加分项目和分值。', 9),
    ('文科', '3341', '浙江外国语学院', '杭州市西湖区留和路299号',
     '21', '旅游管理', None, 2, '4800',
     '认同并执行四川省少数民族加分项目和分值。', 9),
    ('文科', '4115', '河南师范大学', '河南省新乡市建设东路46号',
     '12', '经济学', None, 2, '4840',
     '只承认教育部规定的全国性加分政策。', 12),
    ('文科', '4408', '广州中医药大学', '广东省广州市番禺区大学城外环东路232号',
     '23', '护理学', None, 1, '7660',
     '认同并执行四川省少数民族加分项目和分值。不招色盲色弱考生。', 13),
    ('文科', '5108', '四川轻化工大学', '四川省自贡市自流井区汇兴路519号',
     '12', '国际经济与贸易', None, 11, '5520', None, 15),
    ('文科', '6202', '兰州交通大学', '甘肃省兰州市安宁区安宁西路88号',
     '52', '国际经济与贸易', None, 1, '4000',
     '认同并执行四川省少数民族加分项目和分值。', 17),
    ('文科', '6202', '兰州交通大学', '甘肃省兰州市安宁区安宁西路88号',
     '60', '阿拉伯语', None, 1, '4400',
     '认同并执行四川省少数民族加分项目和分值。', 17),
    ('理科', '0540', '合肥工业大学', '安徽省合肥市屯溪路193号',
     '26', '国际经济与贸易', '中外合作办学；该专业学生入学后不能转专业', 5, '50000',
     '只承认教育部规定的全国性加分政策。', 19),
    ('理科', '0347', '华侨大学', '福建省厦门市集美区集美大道668号',
     '31', '环境工程', '厦门校区', 3, '5460',
     '认同并执行四川省少数民族加分项目和分值。', 22),
    ('理科', '0347', '华侨大学', '福建省厦门市集美区集美大道668号',
     '34', '城乡规划', '厦门校区；要求具有较好的绘画基础', 2, '5460',
     '认同并执行四川省少数民族加分项目和分值。', 22),
    ('理科', '0347', '华侨大学', '福建省厦门市集美区集美大道668号',
     '43', '英语', '泉州校区；招收英语语种考生；外语口试；实行"1+3"两校区融合式培养',
     1, '5460', '认同并执行四川省少数民族加分项目和分值。', 22),
    ('理科', '0347', '华侨大学', '福建省厦门市集美区集美大道668号',
     '45', '翻译', '泉州校区；招收英语语种考生；外语口试；实行"1+3"两校区融合式培养',
     2, '5460', '认同并执行四川省少数民族加分项目和分值。', 22),
    ('理科', '0347', '华侨大学', '福建省厦门市集美区集美大道668号',
     '61', '经济学类', '中美121双学位班，可接轨CFA金融分析师；泉州校区；全英文教学；只招收有专业志愿的考生',
     2, '28000', '认同并执行四川省少数民族加分项目和分值。', 22),
    ('理科', '0347', '华侨大学', '福建省厦门市集美区集美大道668号',
     '63', '国际商务', '全英文教学；泉州校区；只招收有专业志愿的考生',
     3, '28000', '认同并执行四川省少数民族加分项目和分值。', 22),
    ('理科', '1120', '北京工商大学(含增加计划)', '北京市房山区良乡高教园区',
     '11', '经济学', '实验班', 2, '4200',
     '认同并执行四川省少数民族加分项目和分值。', 24),
    ('理科', '1120', '北京工商大学(含增加计划)', '北京市房山区良乡高教园区',
     '18', '工商管理', '数字化管理方向', 5, '4200',
     '认同并执行四川省少数民族加分项目和分值。', 24),
    ('理科', '1302', '河北大学', '河北省保定市五四东路180号',
     '39', '应用心理学', None, 3, '5390',
     '认同并执行四川省少数民族加分项目和分值。', 26),
    ('理科', '1302', '河北大学', '河北省保定市五四东路180号',
     '74', '材料化学', None, 3, '5390',
     '认同并执行四川省少数民族加分项目和分值。', 26),
    ('理科', '2120', '大连外国语大学', '大连市旅顺口区旅顺南路西段6号',
     '1B', '国际经济与贸易', None, 1, '8000',
     '认同并执行四川省少数民族加分项目和分值。', 29),
    ('理科', '2120', '大连外国语大学', '大连市旅顺口区旅顺南路西段6号',
     '1D', '旅游管理', None, 1, '8000',
     '认同并执行四川省少数民族加分项目和分值。', 29),
    ('理科', '2120', '大连外国语大学', '大连市旅顺口区旅顺南路西段6号',
     '1P', '跨境电子商务', None, 1, '8000',
     '认同并执行四川省少数民族加分项目和分值。', 29),
    ('理科', '2211', '长春中医药大学', '吉林省长春市净月国家高新技术产业开发区博硕路1035号',
     '07', '针灸推拿学', None, 2, '6160',
     '认同并执行四川省少数民族加分项目和分值。不招色盲色弱。', 30),
    ('理科', '2211', '长春中医药大学', '吉林省长春市净月国家高新技术产业开发区博硕路1035号',
     '0F', '中药学', None, 1, '6160',
     '认同并执行四川省少数民族加分项目和分值。不招色盲色弱。', 30),
    ('理科', '2444', '黑龙江中医药大学', '黑龙江哈尔滨和平路24号',
     '17', '药学', None, 2, '5000',
     '认同并执行四川省少数民族加分项目和分值。', 31),
    ('理科', '5602', '成都理工大学', '成都市成华区二仙桥东三路1号',
     '65', '商务英语', '宜宾校区；招收英语语种考生；口吃、听力障碍不录。', 4, '5760',
     None, 48),
    ('理科', '5177', '成都东软学院(民办院校)', '四川省成都市都江堰青城山东软大道1号',
     '03', '软件工程', None, 13, '19080', None, 51),
]


# ====== Process ======
out_row = 2

for r in ocr_rows:
    row_num = r['xlsx_row']
    corrections = []

    kl = r['kl']
    yxdm = r['yxdm']
    yxmc = r['yxmc']
    yxdz = r['yxdz']
    zydm = r['zydm']
    zymc = r['zymc']
    zybz = r['zybz']
    zyjhs = r['zyjhs']
    sfbz = r['sfbz']
    yxbz = r['yxbz']
    page = r['page']

    fill = None

    # 1. Row-level corrections (跨页院校信息补全)
    if row_num in row_corrections:
        fix = row_corrections[row_num]
        if 'yxdm' in fix:
            yxdm = fix['yxdm']
        if 'yxmc' in fix:
            yxmc = fix['yxmc']
        if 'yxdz' in fix:
            yxdz = fix['yxdz']
        if 'zydm' in fix:
            zydm = fix['zydm']
        corrections.append(fix['note'])
        fill = fill_yellow

    # 2. School name suffix fixes
    if row_num in school_name_fixes:
        old_name = yxmc
        yxmc = school_name_fixes[row_num]
        corrections.append(f'院校名称修正：{old_name}\u2192{yxmc}')
        fill = fill_yellow

    # 3. Code fixes
    if row_num in code_fixes:
        fix = code_fixes[row_num]
        if 'zydm' in fix:
            old_code = zydm
            zydm = fix['zydm']
            corrections.append(f'专业代码修正：{old_code}\u2192{zydm}')
            fill = fill_yellow

    # 4. Remove [V] marks from zymc AND zydm
    has_v = False
    if zymc and '[V]' in str(zymc):
        zymc = str(zymc).replace('[V]', '').replace('[V] ', '').strip()
        has_v = True
    if zydm and '[V]' in str(zydm):
        zydm = str(zydm).replace('[V]', '').replace('[V] ', '').replace(' [V]', '').strip()
        has_v = True
    if has_v:
        corrections.append('去除[V]标记')
        fill = fill_yellow

    # 5. Split professional name brackets
    if zymc:
        new_name, extra_notes = split_zymc(str(zymc))
        if extra_notes:
            zymc = new_name
            if zybz:
                zybz = str(zybz) + '\uff1b' + extra_notes
            else:
                zybz = extra_notes

    note = '\uff1b'.join(corrections) if corrections else None

    data = [kl, yxdm, yxmc, yxdz, zydm, zymc, zybz, zyjhs, sfbz, yxbz, page, note]
    write_row(ws_out, out_row, data, fill=fill)
    out_row += 1

# Add missing schools (green)
for ms in missing_schools:
    kl, yxdm, yxmc, yxdz, zydm, zymc_raw, zybz, zyjhs, sfbz, yxbz, page = ms

    zymc = zymc_raw
    if zymc:
        new_name, extra = split_zymc(str(zymc))
        if extra:
            zymc = new_name
            if zybz:
                zybz = str(zybz) + '\uff1b' + extra
            else:
                zybz = extra

    data = [kl, yxdm, yxmc, yxdz, zydm, zymc, zybz, zyjhs, sfbz, yxbz, page,
            '缺漏行（跨页院校信息丢失，从原始图片补全）']
    write_row(ws_out, out_row, data, fill=fill_green)
    out_row += 1

# Save
wb_out.save(output_path)

print(f'Output saved: {output_path}')
print(f'Total rows: {out_row - 2} (OCR: {len(ocr_rows)}, Missing added: {len(missing_schools)})')
print(f'Corrections:')
print(f'  Cross-page missing rows added (green): {len(missing_schools)}')
print(f'  Rows with None school info fixed (yellow): {len(row_corrections)}')
print(f'  School name suffix fixes: {len(school_name_fixes)}')
print(f'  Code fixes: {len(code_fixes)}')
print(f'  Page numbers corrected: all rows (OCR page - 1)')
print(f'  Professional name bracket split: applied to all')
