# -*- coding: utf-8 -*-
"""
合并2024年全部已校验征集志愿xlsx，按征集次数去重合并。

合并key: 科目 + 录取批次 + 招生类型 + 院校代码 + 专业代码
同一key多次征集时: 征集次数/页码/批次ID/校正备注 分号合并，来源网页按次数分列。
输出27列统一结构。
"""
import glob
import os
import re
import sys
from collections import OrderedDict

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# ── Target schema (27 columns) ──
TARGET_COLS = [
    "年份", "科目", "录取批次", "招生类型", "降分政策",
    "院校代码", "院校名称", "办学性质", "院校地址", "院校备注",
    "调档线", "专业组代码", "再选科目要求", "专业组计划数",
    "专业代码", "专业名称", "专业备注", "专业计划数", "收费标准",
    "页码", "征集次数",
    "第1次来源网页", "第2次来源网页", "第3次来源网页", "第4次来源网页",
    "批次ID", "校正备注",
]
assert len(TARGET_COLS) == 27

# ── Column alias → unified name ──
ALIAS = {
    "科类": "科目",
    "招生类别": "招生类型",
}

# ── 核心字段：取首次值，有差异追加校正备注 ──
CORE_FIELDS = [
    "科目", "录取批次", "招生类型", "院校代码", "院校名称",
    "院校地址", "调档线", "专业代码", "专业名称", "专业备注",
    "专业计划数", "收费标准", "院校备注",
]

BASE_DIR = "C:/Users/Administrator/Documents/VolunteerHelper/data/13_征集志愿/普通高考"
OUTPUT = os.path.join(BASE_DIR, "征集志愿_2024_合并.xlsx")


def extract_batch_id(filepath: str) -> str:
    """Extract numeric batch ID from directory name like '3780_2024_...'."""
    dirname = os.path.basename(os.path.dirname(filepath))
    m = re.match(r"^(\d+)_", dirname)
    return m.group(1) if m else ""


def s(v) -> str:
    """Safe stringify, strip whitespace."""
    if v is None:
        return ""
    return str(v).strip()


def merge_semicolon(existing: str, new_val: str) -> str:
    """Append new_val with semicolon separator, dedup and preserve order."""
    if not new_val:
        return existing
    if not existing:
        return new_val
    parts = existing.split(";")
    if new_val not in parts:
        parts.append(new_val)
    return ";".join(parts)


def read_xlsx(filepath: str) -> list[dict]:
    """Read one xlsx, return list of dicts with normalized column names."""
    wb = load_workbook(filepath, read_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    raw_headers = list(next(rows_iter))
    headers = [ALIAS.get(h, h) for h in raw_headers]

    batch_id = extract_batch_id(filepath)
    records = []
    for vals in rows_iter:
        row = dict(zip(headers, vals))
        row["_batch_id"] = batch_id
        records.append(row)
    wb.close()
    return records


def main():
    sys.stdout.reconfigure(encoding="utf-8")

    # ── Discover source files ──
    pattern = os.path.join(BASE_DIR, "**", "*2024*已校验*.xlsx")
    files = sorted(glob.glob(pattern, recursive=True))
    files = [f for f in files if "备份" not in f and "R2" not in f and "~$" not in f]
    print(f"源文件数: {len(files)}")

    # ── Read all data ──
    all_rows = []
    for fpath in files:
        rows = read_xlsx(fpath)
        batch_id = extract_batch_id(fpath)
        print(f"  {batch_id}: {len(rows)} rows")
        all_rows.extend(rows)
    print(f"合并前总行数: {len(all_rows)}")

    # ── Merge by key ──
    merged = OrderedDict()

    for row in all_rows:
        key = (
            s(row.get("科目", "")),
            s(row.get("录取批次", "")),
            s(row.get("招生类型", "")),
            s(row.get("院校代码", "")),
            s(row.get("专业代码", "")),
        )

        征集次数 = s(row.get("征集次数", ""))
        页码 = s(row.get("页码", ""))
        来源网页 = s(row.get("来源网页", ""))
        batch_id = s(row.get("_batch_id", ""))
        校正备注 = s(row.get("校正备注", ""))

        if key not in merged:
            target = {col: "" for col in TARGET_COLS}
            target["年份"] = "2024"
            for field in CORE_FIELDS:
                target[field] = s(row.get(field, ""))
            target["征集次数"] = 征集次数
            target["页码"] = 页码
            target["批次ID"] = batch_id
            target["校正备注"] = 校正备注

            # 来源网页按征集次数分列
            try:
                n = int(征集次数)
                if 1 <= n <= 4:
                    target[f"第{n}次来源网页"] = 来源网页
            except (ValueError, TypeError):
                pass

            merged[key] = target
        else:
            target = merged[key]
            target["征集次数"] = merge_semicolon(target["征集次数"], 征集次数)
            target["页码"] = merge_semicolon(target["页码"], 页码)
            target["批次ID"] = merge_semicolon(target["批次ID"], batch_id)
            target["校正备注"] = merge_semicolon(target["校正备注"], 校正备注)

            # 来源网页按征集次数分列
            try:
                n = int(征集次数)
                if 1 <= n <= 4:
                    col_name = f"第{n}次来源网页"
                    if not target[col_name]:
                        target[col_name] = 来源网页
                    elif 来源网页 and 来源网页 != target[col_name]:
                        target[col_name] = merge_semicolon(target[col_name], 来源网页)
            except (ValueError, TypeError):
                pass

            # 核心数据差异检查
            diffs = []
            for field in CORE_FIELDS:
                old_val = s(target.get(field, ""))
                new_val = s(row.get(field, ""))
                if new_val and old_val and new_val != old_val:
                    diffs.append(f"{field}差异:{old_val}→{new_val}")
            if diffs:
                diff_note = "；".join(diffs)
                target["校正备注"] = merge_semicolon(target["校正备注"], diff_note)

    print(f"合并后总行数: {len(merged)}")

    # ── 统计 ──
    count_dist = {}
    for target in merged.values():
        n = len(target["征集次数"].split(";")) if target["征集次数"] else 0
        count_dist[n] = count_dist.get(n, 0) + 1
    print("征集次数分布:")
    for n in sorted(count_dist):
        print(f"  出现{n}次征集: {count_dist[n]}行")

    # ── Write output xlsx ──
    wb = Workbook()
    ws = wb.active
    ws.title = "征集志愿"

    font_normal = Font(name="微软雅黑", size=10)
    font_bold = Font(name="微软雅黑", size=10, bold=True)

    # Header row
    for ci, col_name in enumerate(TARGET_COLS, 1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.font = font_bold
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for ri, target in enumerate(merged.values(), 2):
        for ci, col_name in enumerate(TARGET_COLS, 1):
            val = target.get(col_name, "")
            cell = ws.cell(row=ri, column=ci, value=val if val else None)
            cell.font = font_normal

    # Freeze first row
    ws.freeze_panes = "A2"

    # Auto-width
    for ci in range(1, len(TARGET_COLS) + 1):
        col_letter = get_column_letter(ci)
        header_len = len(TARGET_COLS[ci - 1])
        ws.column_dimensions[col_letter].width = max(header_len * 2.2, 8)

    wb.save(OUTPUT)
    print(f"\n输出文件: {OUTPUT}")
    print("完成!")


if __name__ == "__main__":
    main()
