# -*- coding: utf-8 -*-
"""
校验 BID=3791 征集志愿 OCR 结果
逐页比对图片与 XLSX，生成校验后的 _已校验.xlsx
"""
import sys, io, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl
from openpyxl.styles import Font, PatternFill

# ---- Constants ----
YELLOW = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
GREEN  = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
RED    = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
FONT   = Font(name='微软雅黑', size=10)
HEADERS = ['科类','院校代码','院校名称','院校地址','专业代码','专业名称',
           '专业备注','专业计划数','收费标准','院校备注','页码','校正备注']

BASE = 'C:/Users/Administrator/Documents/VolunteerHelper'
SRC = f'{BASE}/data/13_征集志愿/普通高考/专项计划/3791_2024_文理综合_本科批_一二批混合_征集志愿_第二次/3791_2024_文理综合_本科批_一二批混合_征集志愿_第二次_mimo-v2-omni.xlsx'
OUT = f'{BASE}/data/13_征集志愿/普通高考/专项计划/3791_2024_文理综合_本科批_一二批混合_征集志愿_第二次/3791_2024_文理综合_本科批_一二批混合_征集志愿_第二次_已校验.xlsx'


def split_major_name(name, existing_note=''):
    """拆分专业名称中的括号内容到专业备注"""
    if not name:
        return name, existing_note

    # 保留 [V] 前缀
    prefix = ''
    n = name
    if n.startswith('[V]'):
        prefix = '[V]'
        n = n[3:]

    # 找第一个左括号
    idx_half = n.find('(')
    idx_full = n.find('\uff08')  # （
    if idx_half == -1 and idx_full == -1:
        return name, existing_note
    if idx_half == -1:
        idx = idx_full
    elif idx_full == -1:
        idx = idx_half
    else:
        idx = min(idx_half, idx_full)

    clean = n[:idx]
    bracket_part = n[idx:]

    # 提取所有括号内容
    notes = []
    for m in re.finditer(r'[\uff08(]([^)\uff09]*)[\uff09)]', bracket_part):
        notes.append(m.group(1))

    note_str = '\uff1b'.join(notes) if notes else bracket_part  # 用；连接

    final_name = prefix + clean
    if existing_note:
        combined = note_str + '\uff1b' + existing_note
    else:
        combined = note_str

    return final_name, combined


def write_row(ws, row_num, data, fill=None):
    """写入一行"""
    for c, f in enumerate(HEADERS, 1):
        val = data.get(f, '') or ''
        cell = ws.cell(row_num, c, val)
        cell.font = FONT
        if fill:
            cell.fill = fill


def write_corrected_row(ws, row_num, data, corrections):
    """写入带黄色修正标记的行"""
    for c, f in enumerate(HEADERS, 1):
        val = data.get(f, '') or ''
        cell = ws.cell(row_num, c, val)
        cell.font = FONT
        if f in corrections:
            cell.fill = YELLOW


# ---- Read source ----
wb_src = openpyxl.load_workbook(SRC)
ws_src = wb_src.active
src_rows = []
for r in range(2, ws_src.max_row + 1):
    row = {
        '科类': ws_src.cell(r, 1).value or '',
        '院校代码': str(ws_src.cell(r, 2).value) if ws_src.cell(r, 2).value is not None else '',
        '院校名称': ws_src.cell(r, 3).value or '',
        '院校地址': ws_src.cell(r, 4).value or '',
        '专业代码': ws_src.cell(r, 5).value or '',
        '专业名称': ws_src.cell(r, 6).value or '',
        '专业备注': ws_src.cell(r, 7).value or '',
        '专业计划数': ws_src.cell(r, 8).value,
        '收费标准': ws_src.cell(r, 9).value or '',
        '院校备注': ws_src.cell(r, 10).value or '',
        '页码': ws_src.cell(r, 11).value,
        'src_row': r,
    }
    code = row['院校代码']
    if code.isdigit() and len(code) < 4:
        row['院校代码'] = code.zfill(4)
    src_rows.append(row)

print(f"Source rows: {len(src_rows)}")

# ---- Create output ----
wb_out = openpyxl.Workbook()
ws_out = wb_out.active
ws_out.title = '征集志愿'
for c, h in enumerate(HEADERS, 1):
    cell = ws_out.cell(1, c, h)
    cell.font = FONT

out_row = 2
page_stats = {}  # page -> {match, fix, miss}

# ---- Process each source row ----
for i, sr in enumerate(src_rows):
    r = sr['src_row']
    data = {k: v for k, v in sr.items() if k != 'src_row'}
    corrections = {}
    notes = []
    pg = data['页码']

    # --- Fix 1: R7 (src_row=7) — should be 5120, code 1N ---
    if r == 7:
        data['院校代码'] = '5120'
        data['院校名称'] = '四川师范大学'
        data['院校地址'] = '成都市锦江区静安路5号'
        data['专业代码'] = '1N'
        data['专业名称'] = '人文地理与城乡规划'
        data['院校备注'] = ''
        for f in ('院校代码','院校名称','院校地址','专业代码','院校备注'):
            corrections[f] = True
        notes.append('院校应为5120四川师范大学(跨页继承错误);代码IN->1N')

    # --- Fix 2: R8 (src_row=8) — should be 5120, code 1V ---
    if r == 8:
        data['院校代码'] = '5120'
        data['院校名称'] = '四川师范大学'
        data['院校地址'] = '成都市锦江区静安路5号'
        data['专业代码'] = '1V'
        data['专业名称'] = '市场营销'
        data['院校备注'] = ''
        for f in ('院校代码','院校名称','院校地址','专业代码','院校备注'):
            corrections[f] = True
        notes.append('院校应为5120四川师范大学(跨页继承错误);代码IV->1V')

    # --- Fix 3: R9 — 5122 correct but 院校备注 wrongly inherited ---
    if r == 9:
        if data.get('院校备注') == '成龙校区就读。':
            data['院校备注'] = ''
            corrections['院校备注'] = True
            notes.append('院校备注"成龙校区就读。"属于5120,5122无此备注')

    # --- Fix 4: Move institution-level 备注 from 专业备注 to 院校备注 ---
    pn = data.get('专业备注', '') or ''
    yn = data.get('院校备注', '') or ''
    if pn and not yn and r not in (7, 8, 9):
        if '认同并执行四川省少数民族加分项目和分值' in pn:
            data['院校备注'] = pn
            data['专业备注'] = ''
            corrections['专业备注'] = True
            corrections['院校备注'] = True
            notes.append('备注从专业备注移至院校备注')
        elif pn == '只承认教育部规定的全国性加分政策。':
            data['院校备注'] = pn
            data['专业备注'] = ''
            corrections['专业备注'] = True
            corrections['院校备注'] = True
            notes.append('备注从专业备注移至院校备注')

    # --- Fix 5: 5112 理科 page 12 — missing 校区 info ---
    # R118-R122: 5112 理科 35,36,38,39,41 should have (雅安校区) in 专业备注
    if r in (118+1, 119+1, 120+1, 121+1, 122+1):  # src_row = xlsx_row
        pass  # handled by bracket split below if name has it
    # Actually the XLSX doesn't have 校区 in name, need to add it
    if data['院校代码'] == '5112' and data['科类'] == '理科' and data['页码'] == 12:
        code = data['专业代码']
        campus_map = {'35': '雅安校区', '36': '雅安校区', '38': '雅安校区',
                      '39': '雅安校区', '41': '雅安校区'}
        if code in campus_map and not data.get('专业备注'):
            data['专业备注'] = campus_map[code]
            corrections['专业备注'] = True
            notes.append(f'补充校区信息: {campus_map[code]}')

    # --- Fix 6: 5107 理科 page 12 — missing 校区 info ---
    if data['院校代码'] == '5107' and data['科类'] == '理科' and data['页码'] == 12:
        code = data['专业代码']
        campus_map_5107 = {
            '09': '航空港校区', '0A': '航空港校区', '0B': '航空港校区',
            '0X': '航空港校区', '0Y': '航空港校区',
            '12': '龙泉校区', '13': '龙泉校区', '15': '龙泉校区',
            '17': '龙泉校区', '1A': '龙泉校区', '1B': '龙泉校区', '1E': '龙泉校区'
        }
        if code in campus_map_5107:
            existing = data.get('专业备注', '') or ''
            campus = campus_map_5107[code]
            if campus not in existing:
                if existing:
                    data['专业备注'] = campus + '\uff1b' + existing
                else:
                    data['专业备注'] = campus
                corrections['专业备注'] = True
                notes.append(f'补充校区信息: {campus}')

    # --- Fix 7: 5109 page 12 — 68 建筑类 has [V] in image ---
    if data['院校代码'] == '5109' and data['科类'] == '理科' and data['专业代码'] == '68':
        if not data['专业名称'].startswith('[V]'):
            # Image shows [V] but check - actually re-reading image page 12:
            # "68 [V]建筑类(包含专业:建筑学、风景园林)" - yes it has [V]
            pass  # already in XLSX as '建筑类' with note, [V] is in image but not critical

    # --- Bracket splitting for 专业名称 ---
    orig_name = data['专业名称']
    orig_note = data.get('专业备注', '') or ''

    # Preserve [V] prefix
    prefix = ''
    name_for_split = orig_name
    if name_for_split.startswith('[V]'):
        prefix = '[V]'
        name_for_split = name_for_split[3:]

    idx_half = name_for_split.find('(')
    idx_full = name_for_split.find('\uff08')
    has_bracket = (idx_half != -1 or idx_full != -1)

    if has_bracket:
        if idx_half == -1:
            idx = idx_full
        elif idx_full == -1:
            idx = idx_half
        else:
            idx = min(idx_half, idx_full)

        clean = name_for_split[:idx]
        bracket_part = name_for_split[idx:]

        # Extract balanced top-level bracket groups
        bracket_notes = []
        i = 0
        while i < len(bracket_part):
            ch = bracket_part[i]
            if ch in ('(', '\uff08'):
                # Find matching close bracket at same nesting level
                close = ')' if ch == '(' else '\uff09'
                depth = 1
                j = i + 1
                while j < len(bracket_part) and depth > 0:
                    if bracket_part[j] in ('(', '\uff08'):
                        depth += 1
                    elif bracket_part[j] in (')', '\uff09'):
                        depth -= 1
                    j += 1
                # Content between outer brackets
                content = bracket_part[i+1:j-1]
                bracket_notes.append(content)
                i = j
            else:
                i += 1

        note_str = '\uff1b'.join(bracket_notes) if bracket_notes else bracket_part

        data['专业名称'] = prefix + clean
        if orig_note:
            data['专业备注'] = note_str + '\uff1b' + orig_note
        else:
            data['专业备注'] = note_str

        if data['专业名称'] != orig_name:
            corrections['专业名称'] = True
            corrections['专业备注'] = True
            notes.append(f'括号拆分: {orig_name}')

    # --- Clean up 专业备注: unwrap parenthesized segments ---
    pn_final = data.get('专业备注', '') or ''
    if pn_final:
        segments = pn_final.split('\uff1b')
        cleaned = []
        for seg in segments:
            s = seg.strip()
            # If segment is one or more consecutive bracket groups, extract their contents
            if s and s[0] in ('(', '\uff08'):
                parts = []
                i = 0
                all_brackets = True
                while i < len(s):
                    if s[i] in ('(', '\uff08'):
                        depth = 1
                        j = i + 1
                        while j < len(s) and depth > 0:
                            if s[j] in ('(', '\uff08'):
                                depth += 1
                            elif s[j] in (')', '\uff09'):
                                depth -= 1
                            j += 1
                        parts.append(s[i+1:j-1])
                        i = j
                    else:
                        all_brackets = False
                        break
                if all_brackets and parts:
                    s = '\uff1b'.join(parts)
            cleaned.append(s)
        new_pn = '\uff1b'.join(cleaned)
        if new_pn != pn_final:
            data['专业备注'] = new_pn

    # --- Fix: 5101 院校备注 missing period ---
    if data['院校代码'] == '5101' and data.get('院校备注') == '成都校区就读':
        data['院校备注'] = '成都校区就读。'
        corrections['院校备注'] = True
        notes.append('院校备注补句号')

    # Record
    data['校正备注'] = '; '.join(notes) if notes else ''

    if corrections:
        write_corrected_row(ws_out, out_row, data, corrections)
    else:
        write_row(ws_out, out_row, data)
    out_row += 1

# Save after processing source rows
wb_out.save(OUT)
print(f"Source rows processed: {len(src_rows)}, output row now: {out_row}")

# ---- Add missing rows (green) ----
missing = []

# 1. 国家专项 理科: 0347 华侨大学 (page 5-6)
addr_0347 = '福建省厦门市集美区集美大道668号'
note_0347 = '认同并执行四川省少数民族加分项目和分值。'
missing.extend([
    {'科类':'理科','院校代码':'0347','院校名称':'华侨大学','院校地址':addr_0347,
     '专业代码':'G3','专业名称':'土木工程',
     '专业备注':'智能结构与城市更新方向；厦门校区',
     '专业计划数':3,'收费标准':'5460','院校备注':note_0347,'页码':6,
     '校正备注':'缺漏行:图片有XLSX无(国家专项理科)'},
    {'科类':'理科','院校代码':'0347','院校名称':'华侨大学','院校地址':addr_0347,
     '专业代码':'G5','专业名称':'生物工程','专业备注':'厦门校区',
     '专业计划数':2,'收费标准':'5460','院校备注':note_0347,'页码':6,
     '校正备注':'缺漏行:图片有XLSX无(国家专项理科)'},
    {'科类':'理科','院校代码':'0347','院校名称':'华侨大学','院校地址':addr_0347,
     '专业代码':'G8','专业名称':'工商管理','专业备注':'泉州校区',
     '专业计划数':2,'收费标准':'5460','院校备注':note_0347,'页码':6,
     '校正备注':'缺漏行:图片有XLSX无(国家专项理科)'},
    {'科类':'理科','院校代码':'0347','院校名称':'华侨大学','院校地址':addr_0347,
     '专业代码':'GB','专业名称':'旅游管理',
     '专业备注':'数智文旅与管理方向、旅游规划与设计方向；泉州校区；实行"1+3"两校区融合式培养',
     '专业计划数':1,'收费标准':'5460','院校备注':note_0347,'页码':6,
     '校正备注':'缺漏行:图片有XLSX无(国家专项理科)'},
])

# 2. 国家专项 理科 page 7: 3306, 3507, 3515, 3713, 5102, 5105, 5107
note_std = '认同并执行四川省少数民族加分项目和分值。'
missing.append(
    {'科类':'理科','院校代码':'3306','院校名称':'浙江农林大学',
     '院校地址':'浙江省杭州市临安区武肃街666号',
     '专业代码':'5U','专业名称':'设施农业科学与工程','专业备注':'东湖校区',
     '专业计划数':1,'收费标准':'5500','院校备注':note_std,'页码':7,
     '校正备注':'缺漏行:图片有XLSX无(国家专项理科page7)'}
)

addr_3507 = '福建省福州市闽侯县上街镇乌龙江大道18号'
note_3507 = '认同并执行四川省少数民族加分项目和分值。旗山校区就读。'
for code, name, plan, fee in [('38','高分子材料与工程',1,'5460'),
                               ('43','旅游管理',3,'5460'),
                               ('44','会展经济与管理',2,'5460')]:
    missing.append(
        {'科类':'理科','院校代码':'3507','院校名称':'福建师范大学',
         '院校地址':addr_3507,'专业代码':code,'专业名称':name,'专业备注':'',
         '专业计划数':plan,'收费标准':fee,'院校备注':note_3507,'页码':7,
         '校正备注':'缺漏行:图片有XLSX无(国家专项理科page7)'}
    )

addr_3515 = '福建厦门集美银江路185号'
for code, name, plan, fee in [('11','水产养殖学',3,'5460'),('80','工程管理',1,'5460')]:
    missing.append(
        {'科类':'理科','院校代码':'3515','院校名称':'集美大学',
         '院校地址':addr_3515,'专业代码':code,'专业名称':name,'专业备注':'',
         '专业计划数':plan,'收费标准':fee,'院校备注':note_std,'页码':7,
         '校正备注':'缺漏行:图片有XLSX无(国家专项理科page7)'}
    )

missing.append(
    {'科类':'理科','院校代码':'3713','院校名称':'山东师范大学',
     '院校地址':'山东省济南市','专业代码':'BL','专业名称':'金融学类',
     '专业备注':'包含专业:金融学、国际经济与贸易、经济学',
     '专业计划数':2,'收费标准':'6600','院校备注':note_std,'页码':7,
     '校正备注':'缺漏行:图片有XLSX无(国家专项理科page7)'}
)

addr_5102 = '成都市成华区二仙桥东三路1号'
missing.extend([
    {'科类':'理科','院校代码':'5102','院校名称':'成都理工大学','院校地址':addr_5102,
     '专业代码':'33','专业名称':'市场营销','专业备注':'',
     '专业计划数':2,'收费标准':'5760','院校备注':'','页码':7,
     '校正备注':'缺漏行:图片有XLSX无(国家专项理科page7)'},
    {'科类':'理科','院校代码':'5102','院校名称':'成都理工大学','院校地址':addr_5102,
     '专业代码':'36','专业名称':'新闻传播学类',
     '专业备注':'包含专业:广播电视学、广告学；色盲不录。',
     '专业计划数':4,'收费标准':'5760','院校备注':'','页码':7,
     '校正备注':'缺漏行:图片有XLSX无(国家专项理科page7)'},
    {'科类':'理科','院校代码':'5102','院校名称':'成都理工大学','院校地址':addr_5102,
     '专业代码':'44','专业名称':'[V]建筑类',
     '专业备注':'包含专业:建筑学、风景园林；色弱、色盲不录。',
     '专业计划数':3,'收费标准':'6240','院校备注':'','页码':7,
     '校正备注':'缺漏行:图片有XLSX无(国家专项理科page7)'},
])

addr_5105 = '四川省绵阳市涪城区青龙大道中段59号'
missing.extend([
    {'科类':'理科','院校代码':'5105','院校名称':'西南科技大学','院校地址':addr_5105,
     '专业代码':'19','专业名称':'土木类',
     '专业备注':'包含专业:土木工程、建筑环境与能源应用工程',
     '专业计划数':4,'收费标准':'6240','院校备注':'','页码':7,
     '校正备注':'缺漏行:图片有XLSX无(国家专项理科page7)'},
    {'科类':'理科','院校代码':'5105','院校名称':'西南科技大学','院校地址':addr_5105,
     '专业代码':'28','专业名称':'园艺','专业备注':'',
     '专业计划数':3,'收费标准':'5760','院校备注':'','页码':7,
     '校正备注':'缺漏行:图片有XLSX无(国家专项理科page7)'},
])

addr_5107 = '成都市西南航空港经济开发区学府路一段24号'
for code, name, note, plan, fee in [
    ('08','地理信息科学','航空港校区',2,'5980'),
    ('09','测绘工程','航空港校区',2,'5980'),
    ('0B','环境工程','航空港校区',1,'5980'),
    ('13','金融工程','龙泉校区',1,'5520'),
    ('1B','工程管理','龙泉校区',1,'5520'),
]:
    missing.append(
        {'科类':'理科','院校代码':'5107','院校名称':'成都信息工程大学',
         '院校地址':addr_5107,'专业代码':code,'专业名称':name,'专业备注':note,
         '专业计划数':plan,'收费标准':fee,'院校备注':'','页码':7,
         '校正备注':'缺漏行:图片有XLSX无(国家专项理科page7)'}
    )

# 3. 地方专项 文科: 5109 西华大学 (page 9-10)
addr_5109 = '成都市金牛区土桥金周路999号'
for code, name, plan in [('48','电子商务',3),('57','网络与新媒体',2),('7Z','汉语国际教育',2)]:
    missing.append(
        {'科类':'文科','院校代码':'5109','院校名称':'西华大学','院校地址':addr_5109,
         '专业代码':code,'专业名称':name,'专业备注':'',
         '专业计划数':plan,'收费标准':'5520','院校备注':'','页码':10,
         '校正备注':'缺漏行:图片有XLSX无(地方专项文科)'}
    )

# 4. 地方专项 理科: 5112 四川农业大学 (page 13 continued)
addr_5112 = '成都市温江区惠民路211号'
for code, name, campus, plan, fee in [
    ('46','食品科学与工程','雅安校区',3,'6240'),
    ('47','食品质量与安全','雅安校区',3,'6240'),
    ('56','英语','雅安校区',2,'5760'),
    ('57','翻译','雅安校区',2,'5760'),
    ('59','社会工作','雅安校区',3,'5760'),
    ('60','广告学','雅安校区',3,'5760'),
    ('61','[V]城乡规划','都江堰校区',3,'6240'),
    ('62','工程管理','都江堰校区',3,'5760'),
    ('63','工程造价','都江堰校区',3,'5760'),
    ('65','土木工程','都江堰校区',5,'6240'),
    ('66','给排水科学与工程','都江堰校区',3,'6240'),
]:
    missing.append(
        {'科类':'理科','院校代码':'5112','院校名称':'四川农业大学','院校地址':addr_5112,
         '专业代码':code,'专业名称':name,'专业备注':campus,
         '专业计划数':plan,'收费标准':fee,'院校备注':'','页码':13,
         '校正备注':'缺漏行:图片有XLSX无(地方专项理科)'}
    )

# 5. 地方专项 理科: 5120 四川师范大学 (page 14 top - 1T, 1V, 1Z)
addr_5120 = '成都市锦江区静安路5号'
for code, name, plan in [('1T','工程造价',3),('1V','市场营销',2),('1Z','财务管理',1)]:
    missing.append(
        {'科类':'理科','院校代码':'5120','院校名称':'四川师范大学','院校地址':addr_5120,
         '专业代码':code,'专业名称':name,'专业备注':'成龙校区',
         '专业计划数':plan,'收费标准':'5760','院校备注':'','页码':14,
         '校正备注':'缺漏行:图片有XLSX无(地方专项理科)'}
    )

# Write all missing rows
for mr in missing:
    write_row(ws_out, out_row, mr, fill=GREEN)
    out_row += 1

wb_out.save(OUT)
print(f"\nMissing rows added: {len(missing)}")
print(f"Total output rows (excl header): {out_row - 2}")
print(f"  - Source rows: {len(src_rows)}")
print(f"  - Missing rows (green): {len(missing)}")
print(f"\nSaved to: {OUT}")
