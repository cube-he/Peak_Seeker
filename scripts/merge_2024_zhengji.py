# -*- coding: utf-8 -*-
"""
合并30个2024年已校验征集志愿xlsx，按征集次数展开计划数和来源网页。
输出：征集志愿_2024_合并.xlsx（29列统一结构）

列结构：
  年份, 科目, 录取批次, 招生类型, 降分政策, 院校代码, 院校名称, 办学性质, 院校地址, 院校备注, 调档线,
  专业组代码, 再选科目要求, 专业代码, 专业名称, 专业备注, 收费标准,
  第1次专业计划数, 第1次来源网页, 第2次专业计划数, 第2次来源网页,
  第3次专业计划数, 第3次来源网页, 第4次专业计划数, 第4次来源网页,
  页码, 征集次数, 批次ID, 校正备注

合并key: 科目 + 录取批次 + 招生类型 + 院校代码 + 专业代码
"""
import os
import re
from collections import OrderedDict
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

# ── 统一列结构（29列）──
HEADERS = [
    '年份',             # 0
    '科目',             # 1
    '录取批次',         # 2
    '招生类型',         # 3
    '降分政策',         # 4
    '院校代码',         # 5
    '院校名称',         # 6
    '办学性质',         # 7
    '院校地址',         # 8
    '院校备注',         # 9
    '调档线',           # 10
    '专业组代码',       # 11
    '再选科目要求',     # 12
    '专业代码',         # 13
    '专业名称',         # 14
    '专业备注',         # 15
    '收费标准',         # 16
    '第1次专业计划数',  # 17
    '第1次来源网页',    # 18
    '第2次专业计划数',  # 19
    '第2次来源网页',    # 20
    '第3次专业计划数',  # 21
    '第3次来源网页',    # 22
    '第4次专业计划数',  # 23
    '第4次来源网页',    # 24
    '页码',             # 25
    '征集次数',         # 26
    '批次ID',           # 27
    '校正备注',         # 28
]
H = {h: i for i, h in enumerate(HEADERS)}
NUM_COLS = len(HEADERS)

# 展开列索引
def plan_col(n): return H[f'第{n}次专业计划数']
def url_col(n):  return H[f'第{n}次来源网页']

# 分号拼接列
SEMICOLON_INDICES = [H['页码'], H['征集次数'], H['批次ID'], H['校正备注']]

# 源列名 → 统一列名
NAME_MAP = {
    '科目': '科目', '科类': '科目',
    '录取批次': '录取批次',
    '招生类型': '招生类型', '招生类别': '招生类型',
    '院校代码': '院校代码', '院校名称': '院校名称', '院校地址': '院校地址',
    '调档线': '调档线',
    '专业代码': '专业代码', '专业名称': '专业名称', '专业备注': '专业备注',
    '专业计划数': '专业计划数',  # 特殊：按征集次数展开
    '收费标准': '收费标准', '院校备注': '院校备注',
    '页码': '页码', '征集次数': '征集次数',
    '来源网页': '来源网页',      # 特殊：按征集次数展开
    '校正备注': '校正备注',
}

# 直接映射的非展开字段（源统一名 → 目标列名）
DIRECT_FIELDS = [
    '科目', '录取批次', '招生类型', '院校代码', '院校名称',
    '院校地址', '调档线', '专业代码', '专业名称', '专业备注',
    '收费标准', '院校备注',
]

BASE_DIR = 'C:/Users/Administrator/Documents/VolunteerHelper/data/13_征集志愿/普通高考'
OUTPUT_PATH = os.path.join(BASE_DIR, '征集志愿_2024_合并.xlsx')


def to_str(v):
    if v is None:
        return ''
    return str(v).strip()


def merge_semicolon(existing_row, new_val_str, idx):
    """将 new_val_str 追加到 existing_row[idx]，分号分隔，去重保序"""
    ev = to_str(existing_row[idx])
    nv = new_val_str
    if not nv:
        return
    parts = ev.split(';') if ev else []
    parts = [p for p in parts if p]
    if nv not in parts:
        parts.append(nv)
    existing_row[idx] = ';'.join(parts)


def find_source_files():
    """找到所有2024年已校验xlsx，按batch_id排序"""
    files = []
    for root, dirs, fnames in os.walk(BASE_DIR):
        for fn in fnames:
            if '2024' in fn and '已校验' in fn and fn.endswith('.xlsx'):
                batch_id = fn.split('_')[0]
                files.append((batch_id, os.path.join(root, fn)))
    files.sort(key=lambda x: int(x[0]))
    return files


def main():
    src_files = find_source_files()
    print(f'源文件数: {len(src_files)}')

    merged = OrderedDict()
    total_in = 0
    merge_count = 0

    for batch_id, fpath in src_files:
        wb = load_workbook(fpath, read_only=True)
        ws = wb.active

        # 读 header，建映射
        raw_headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        col_map = {}  # src_col_idx → unified_field_name
        for ci, h in enumerate(raw_headers):
            if h and h in NAME_MAP:
                col_map[ci] = NAME_MAP[h]

        row_count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            # 跳过全空行
            if all(v is None or (isinstance(v, str) and v.strip() == '') for v in row):
                continue

            # 读源数据到 dict
            src = {}
            for ci, uname in col_map.items():
                src[uname] = row[ci]

            # 征集次数
            n_raw = src.get('征集次数')
            if n_raw is None:
                print(f'  WARN: 征集次数为空, batch={batch_id}')
                continue
            n = int(n_raw)
            if n < 1 or n > 4:
                print(f'  WARN: 征集次数={n} 超出1-4, batch={batch_id}')
                continue

            # 构建 unified row
            urow = [None] * NUM_COLS
            urow[H['年份']] = '2024'

            # 直接映射字段
            for field in DIRECT_FIELDS:
                if field in src and src[field] is not None:
                    urow[H[field]] = src[field]

            # 展开字段：专业计划数 → 第n次专业计划数
            urow[plan_col(n)] = src.get('专业计划数')
            # 展开字段：来源网页 → 第n次来源网页
            urow[url_col(n)] = src.get('来源网页')

            # 分号拼接字段
            urow[H['页码']] = to_str(src.get('页码'))
            urow[H['征集次数']] = to_str(n)
            urow[H['批次ID']] = batch_id
            urow[H['校正备注']] = to_str(src.get('校正备注'))

            # 合并key: 科目 + 录取批次 + 招生类型 + 院校代码 + 专业代码
            key = '|'.join([
                to_str(src.get('科目')),
                to_str(src.get('录取批次')),
                to_str(src.get('招生类型')),
                to_str(src.get('院校代码')),
                to_str(src.get('专业代码')),
            ])

            if key not in merged:
                merged[key] = urow
            else:
                existing = merged[key]
                # 合并展开列（不覆盖已有值）
                pi, ui = plan_col(n), url_col(n)
                if urow[pi] is not None and existing[pi] is None:
                    existing[pi] = urow[pi]
                if urow[ui] is not None and existing[ui] is None:
                    existing[ui] = urow[ui]
                # 合并分号拼接列
                for idx in SEMICOLON_INDICES:
                    merge_semicolon(existing, to_str(urow[idx]), idx)
                merge_count += 1

            row_count += 1

        wb.close()
        total_in += row_count
        print(f'  {batch_id}: {row_count} rows')

    total_out = len(merged)
    print(f'\n统计:')
    print(f'  源文件: {len(src_files)}')
    print(f'  读入行数: {total_in}')
    print(f'  合并次数: {merge_count}')
    print(f'  输出行数: {total_out}')

    # ── 写 xlsx ──
    wb_out = Workbook()
    ws = wb_out.active
    ws.title = '征集志愿'

    font_h = Font(name='微软雅黑', size=10, bold=True)
    font_d = Font(name='微软雅黑', size=10)

    for ci, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = font_h

    for ri, row_data in enumerate(merged.values(), 2):
        for ci, v in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.font = font_d

    ws.freeze_panes = 'A2'
    wb_out.save(OUTPUT_PATH)
    print(f'\n输出: {OUTPUT_PATH}')
    print(f'行数: {total_out} (不含表头)')


if __name__ == '__main__':
    main()
