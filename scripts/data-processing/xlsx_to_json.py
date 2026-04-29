"""
xlsx → JSON converter for VolunteerHelper data pipeline.

Reads master xlsx files from data/03_专家版主表/output/ and produces
JSON files compatible with import_to_db.ts.

Outputs:
  - universities_enriched.json
  - majors_enriched.json
  - enrollment_plans_enriched.json
  - admission_records_filled.json

Usage:
  python scripts/data-processing/xlsx_to_json.py
  python scripts/data-processing/xlsx_to_json.py --xlsx-dir=path --out-dir=path
"""

import argparse
import json
import os
import sys
from pathlib import Path

import openpyxl


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _str(val) -> str | None:
    """Convert to stripped string or None."""
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def _int(val) -> int | None:
    """Convert to int or None."""
    if val is None:
        return None
    try:
        return int(float(str(val)))
    except (ValueError, TypeError):
        return None


def _float(val) -> float | None:
    """Convert to float or None."""
    if val is None:
        return None
    try:
        return float(str(val))
    except (ValueError, TypeError):
        return None


def _bool_from_str(val) -> bool:
    """'是' → True, everything else → False."""
    return _str(val) == '是'


def _any_not_none(*vals) -> bool:
    """True if at least one value is not None."""
    return any(v is not None for v in vals)


# ---------------------------------------------------------------------------
# University converter
# ---------------------------------------------------------------------------

def convert_university_row(row: tuple) -> dict:
    """Convert a single row from 院校信息表.xlsx to a university dict.

    Row indices follow the column spec (0-89).
    """
    tags_raw = _str(row[8])  # 院校档次 e.g. "985/211/双一流/国重点/保研资格"
    tags = [t.strip() for t in tags_raw.split('/')] if tags_raw else []

    master_count = _int(row[15])
    doctoral_count = _int(row[16])

    return {
        'enrollCode': _int(row[0]) or _str(row[0]),
        'code': _str(row[49]),           # 国标代码
        'name': _str(row[1]),
        'province': _str(row[2]),
        'city': _str(row[3]),
        'type': _str(row[5]),
        'level': _str(row[11]),          # 院校层级
        'runningNature': _str(row[6]),
        'isDoubleFirstClass': _bool_from_str(row[12]),
        'is985': '985' in tags,
        'is211': '211' in tags,
        'tags': tags,
        'grade': _str(row[4]),           # 城市等级
        'department': _str(row[7]),
        'ranking': _int(row[14]) or _str(row[14]),
        'admissionGuide': _str(row[23]),
        'renameHistory': _str(row[24]),
        'transferDifficulty': _str(row[22]),
        'postgradRate': _str(row[21]),
        'disciplineEvaluationLevel': _str(row[25]),
        'softRating': None,              # not in university sheet
        'softRanking': _int(row[60]),
        'masterProgramCount': master_count,
        'doctoralProgramCount': doctoral_count,
        'hasMasterProgram': (master_count or 0) > 0,
        'hasDoctoralProgram': (doctoral_count or 0) > 0,
        'masterPrograms': _str(row[17]),
        'doctoralPrograms': _str(row[18]),
        'maleRatio': _int(row[42]),
        'femaleRatio': _int(row[43]),
        'createdYear': _str(row[41]),
        'logoUrl': _str(row[53]),
        'aClassDisciplineCount': _int(row[64]),
        'rankingSoft': _int(row[60]),
        'rankingAlumni': _int(row[46]),
        'rankingQS': _int(row[44]),
        'rankingUSNews': _int(row[45]),
        'satisfactionOverall': _float(row[69]),
        'satisfactionLife': _float(row[76]),
        'satisfactionEnviron': _float(row[83]),
        'satisfactionCount': _int(row[70]),
        # Charter fields from xlsx columns 26-33
        'charterFilingRatio': _str(row[26]),
        'charterMajorAllocation': _str(row[27]),
        'charterTiebreaker': _str(row[28]),
        'charterLanguageReq': _str(row[31]),
        'charterPhysicalReq': _str(row[29]),
        'charterBonusPolicy': _str(row[33]),
        'charterAdjustment': _str(row[30]),
        # Employment / further study
        # col40 = 升学率 (further study rate), not employment rate
        'employmentRate': None,
        'furtherStudyRate': _str(row[40]) if _str(row[40]) and '%' in str(row[40]) else (
            f"{row[40]}%" if _int(row[40]) is not None else None
        ),
        'avgSalary': None,
        'topEmployers': None,
    }


# ---------------------------------------------------------------------------
# Major converter
# ---------------------------------------------------------------------------

def _infer_level(batch_str: str) -> str:
    """从录取批次推断学历层级：含'专科'→专科，否则→本科。"""
    if batch_str and '专科' in batch_str:
        return '专科'
    return '本科'


def convert_majors_from_main_table(rows: list[tuple]) -> list[dict]:
    """Deduplicate majors by (专业名, 层级) and return unique major dicts."""
    seen = set()
    majors = []

    for row in rows:
        name = _str(row[12])
        if not name:
            continue
        batch = _str(row[5]) or ''
        level = _infer_level(batch)
        key = (name, level)
        if key in seen:
            continue
        seen.add(key)

        majors.append({
            'name': name,
            'code': _str(row[4]),
            'category': _str(row[15]),      # 门类
            'level': level,
            'discipline': _str(row[14]),     # 专业类
            'type': None,
            'notes': _str(row[16]),          # 专业备注
            'majorLevel': _str(row[65]),     # 专业水平
            'softRating': _str(row[62]),     # 软科评级
        })

    return majors


# ---------------------------------------------------------------------------
# Enrollment plan converter
# ---------------------------------------------------------------------------

def convert_enrollment_plans_from_row(row: tuple) -> list[dict]:
    """Convert a single row from 专业招生主表.xlsx to up to 3 enrollment plan dicts.

    Produces records for years 2025, 2024, 2023 based on available plan counts.
    """
    # Shared fields across all years for this row
    shared = {
        'universityEnrollCode': _str(row[1]),
        'majorName': _str(row[12]),
        'majorCode': _str(row[4]),
        'groupCode': _str(row[3]),
        'subjects': _str(row[6]),
        'batch': _str(row[5]),
        'recruitType': _str(row[9]),
        'province': '四川',
        'level': _infer_level(_str(row[5]) or ''),
        'subjectRequirements': _str(row[19]),
        'isNew': _bool_from_str(row[18]),
        'oldBatch': _str(row[7]),
        'disciplineEval': _str(row[64]),
        'isNationalFeature': _bool_from_str(row[66]),
        'majorRanking': _str(row[67]),
        'majorHonor': _str(row[68]),
        'localMasterPoint': _str(row[69]),
        'localDoctoralPoint': _str(row[70]),
        'softRating': _str(row[62]),
        'planNotes': _str(row[17]),
    }

    plans = []

    # 2025: planCount=col21, groupPlanCount=col20, tuition=col23, duration=col22
    plan_25 = _int(row[21])
    if plan_25 is not None:
        plans.append({
            **shared,
            'year': 2025,
            'planCount': plan_25,
            'groupPlanCount': _int(row[20]),
            'tuition': _int(row[23]),
            'duration': _str(row[22]),
        })

    # 2024: planCount=col39
    plan_24 = _int(row[39])
    if plan_24 is not None:
        plans.append({
            **shared,
            'year': 2024,
            'planCount': plan_24,
            'groupPlanCount': None,
            'tuition': None,
            'duration': None,
        })

    # 2023: planCount=col47
    plan_23 = _int(row[47])
    if plan_23 is not None:
        plans.append({
            **shared,
            'year': 2023,
            'planCount': plan_23,
            'groupPlanCount': None,
            'tuition': None,
            'duration': None,
        })

    return plans


# ---------------------------------------------------------------------------
# Admission record converter
# ---------------------------------------------------------------------------

def convert_admission_records_from_row(row: tuple) -> list[dict]:
    """Convert a single row from 专业招生主表.xlsx to up to 4 admission record dicts.

    Produces records for years 2025, 2024, 2023, 2022.
    Skips a year if ALL score/count fields are null.
    """
    shared = {
        'universityEnrollCode': _str(row[1]),
        'majorName': _str(row[12]),
        'majorCode': _str(row[4]),
        'groupCode': _str(row[3]),
        'subjects': _str(row[6]),
        'batch': _str(row[5]),
        'recruitType': _str(row[9]),
        'province': '四川',
        'level': _infer_level(_str(row[5]) or ''),
    }

    records = []

    # 2025: cols 24-35
    vals_25 = {
        'filingMinScore': _int(row[24]),
        'filingMinRank': _int(row[25]),
        'groupAdmissionCount': _int(row[26]),
        'groupMinScore': _int(row[27]),
        'groupMinRank': _int(row[28]),
        'majorAdmissionCount': _int(row[29]),
        'majorMinScore': _int(row[30]),
        'majorMinRank': _int(row[31]),
        'majorAvgScore': _int(row[32]),
        'majorAvgRank': _int(row[33]),
        'majorMaxScore': _int(row[34]),
        'majorMaxRank': _int(row[35]),
    }
    if _any_not_none(*vals_25.values()):
        records.append({**shared, 'year': 2025, **vals_25})

    # 2024: cols 36-46
    vals_24 = {
        'groupMinScore': _int(row[36]),
        'groupMinRank': _int(row[37]),
        'groupAdmissionCount': _int(row[38]),
        'majorAdmissionCount': _int(row[40]),
        'majorMinScore': _int(row[41]),
        'majorMinRank': _int(row[42]),
        'majorAvgScore': _int(row[43]),
        'majorAvgRank': _int(row[44]),
        'majorMaxScore': _int(row[45]),
        'majorMaxRank': _int(row[46]),
    }
    if _any_not_none(*vals_24.values()):
        records.append({**shared, 'year': 2024, **vals_24})

    # 2023: cols 48-54
    vals_23 = {
        'majorAdmissionCount': _int(row[48]),
        'majorMinScore': _int(row[49]),
        'majorMinRank': _int(row[50]),
        'majorAvgScore': _int(row[51]),
        'majorAvgRank': _int(row[52]),
        'majorMaxScore': _int(row[53]),
        'majorMaxRank': _int(row[54]),
    }
    if _any_not_none(*vals_23.values()):
        records.append({**shared, 'year': 2023, **vals_23})

    # 2022: cols 55-61
    vals_22 = {
        'majorAdmissionCount': _int(row[55]),
        'majorMinScore': _int(row[56]),
        'majorMinRank': _int(row[57]),
        'majorAvgScore': _int(row[58]),
        'majorAvgRank': _int(row[59]),
        'majorMaxScore': _int(row[60]),
        'majorMaxRank': _int(row[61]),
    }
    if _any_not_none(*vals_22.values()):
        records.append({**shared, 'year': 2022, **vals_22})

    return records


# ---------------------------------------------------------------------------
# File-level converters
# ---------------------------------------------------------------------------

def convert_universities(xlsx_path: str) -> list[dict]:
    """Read 院校信息表.xlsx and return list of university dicts."""
    print(f'Reading {xlsx_path} ...')
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    universities = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        # Skip completely empty rows
        if not any(cell is not None for cell in row[:5]):
            continue
        universities.append(convert_university_row(row))
    wb.close()
    print(f'  → {len(universities)} universities')
    return universities


def convert_main_table(xlsx_path: str):
    """Read 专业招生主表.xlsx and return (majors, enrollment_plans, admission_records)."""
    print(f'Reading {xlsx_path} ...')
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(cell is not None for cell in row[:5]):
            continue
        rows.append(row)
    wb.close()
    print(f'  → {len(rows)} data rows')

    # Majors (deduplicated)
    print('Converting majors ...')
    majors = convert_majors_from_main_table(rows)
    print(f'  → {len(majors)} unique majors')

    # Enrollment plans
    print('Converting enrollment plans ...')
    plans = []
    for row in rows:
        plans.extend(convert_enrollment_plans_from_row(row))
    print(f'  → {len(plans)} enrollment plans')

    # Admission records
    print('Converting admission records ...')
    records = []
    for row in rows:
        records.extend(convert_admission_records_from_row(row))
    print(f'  → {len(records)} admission records')

    return majors, plans, records


def write_json(data: list, out_path: str):
    """Write list of dicts to JSON file with UTF-8 encoding."""
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    with open(out_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f'  Written {len(data)} records → {out_path}')


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description='Convert xlsx master tables to JSON')
    parser.add_argument(
        '--xlsx-dir',
        default=None,
        help='Directory containing xlsx files (default: data/03_专家版主表/output/)',
    )
    parser.add_argument(
        '--out-dir',
        default=None,
        help='Output directory for JSON files (default: scripts/data-processing/output/)',
    )
    args = parser.parse_args()

    # Resolve project root (go up from this script's location)
    project_root = Path(__file__).resolve().parent.parent.parent

    xlsx_dir = Path(args.xlsx_dir) if args.xlsx_dir else project_root / 'data' / '03_专家版主表' / 'output'
    out_dir = Path(args.out_dir) if args.out_dir else project_root / 'scripts' / 'data-processing' / 'output'

    uni_xlsx = xlsx_dir / '院校信息表.xlsx'
    main_xlsx = xlsx_dir / '专业招生主表.xlsx'

    if not uni_xlsx.exists():
        print(f'ERROR: {uni_xlsx} not found', file=sys.stderr)
        sys.exit(1)
    if not main_xlsx.exists():
        print(f'ERROR: {main_xlsx} not found', file=sys.stderr)
        sys.exit(1)

    print('=' * 60)
    print('xlsx → JSON converter')
    print('=' * 60)

    # 1. Universities
    universities = convert_universities(str(uni_xlsx))
    write_json(universities, str(out_dir / 'universities_enriched.json'))

    # 2. Main table → majors + plans + records
    majors, plans, records = convert_main_table(str(main_xlsx))
    write_json(majors, str(out_dir / 'majors_enriched.json'))
    write_json(plans, str(out_dir / 'enrollment_plans_enriched.json'))
    write_json(records, str(out_dir / 'admission_records_filled.json'))

    print()
    print('=' * 60)
    print('Done!')
    print(f'  universities_enriched.json:      {len(universities)} records')
    print(f'  majors_enriched.json:             {len(majors)} records')
    print(f'  enrollment_plans_enriched.json:   {len(plans)} records')
    print(f'  admission_records_filled.json:    {len(records)} records')
    print('=' * 60)


if __name__ == '__main__':
    main()
