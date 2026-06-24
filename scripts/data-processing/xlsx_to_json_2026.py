"""
86-col master xlsx → JSON converter for the 2026 go-live.

Reads 四川-2026-专家版数据_批次标准化.xlsx (Sheet1, 86 cols) BY COLUMN NAME
(not positional index) and produces JSON compatible with import_to_db.ts:
  - majors_enriched.json
  - enrollment_plans_enriched.json   (year=2026 + historical 2025/2024/2023)
  - admission_records_filled.json    (2025/2024/2023)

Universities still come from 院校信息表.xlsx via xlsx_to_json.convert_universities.

Fails fast if a required column is missing (vs the old positional converter that
silently mis-read columns). See design:
  docs/superpowers/specs/2026-06-24-86col-converter-design.md
"""

import argparse
import json
import os
import sys
from pathlib import Path

import openpyxl

# Columns that MUST exist; absence means the input is not the expected 86-col master.
REQUIRED_COLS = [
    "科类", "批次", "招生类型", "院校代码", "专业组代码", "专业代码", "专业名称",
    "本科/专科", "计划人数", "专业组计划人数",
    "录取人数1", "最低分1", "最低位次1", "计划人数结果1",
]


def build_header_index(header) -> dict:
    """Map column name → index from the header row; fail fast if a required column is missing."""
    H = {name: i for i, name in enumerate(header) if name is not None}
    missing = [c for c in REQUIRED_COLS if c not in H]
    if missing:
        raise ValueError(f"主表缺少必需列: {missing}")
    return H


def row_to_dict(row, H) -> dict:
    """Project a row tuple to a {column-name: value} dict using the header index."""
    return {name: row[i] for name, i in H.items()}


# --- value helpers ---------------------------------------------------------

def _str(val):
    if val is None:
        return None
    s = str(val).strip()
    return s or None


def _int(val):
    if val is None:
        return None
    try:
        return int(float(str(val)))
    except (ValueError, TypeError):
        return None


def _float(val):
    if val is None:
        return None
    try:
        return float(str(val))
    except (ValueError, TypeError):
        return None


def _bool_from_str(val) -> bool:
    return _str(val) == "是"


def _any_not_none(*vals) -> bool:
    return any(v is not None for v in vals)


def _norm_level(val) -> str:
    """本/专科层级：职业本科 → 本科（下游本/专二分一致）。"""
    s = _str(val) or "本科"
    return "本科" if s == "职业本科" else s


# --- majors ----------------------------------------------------------------

def convert_majors(rows) -> list:
    """Deduplicate majors by (专业名称, level). See spec『majors_enriched』mapping."""
    seen, out = set(), []
    for r in rows:
        name = _str(r.get("专业名称"))
        if not name:
            continue
        level = _norm_level(r.get("本科/专科"))
        key = (name, level)
        if key in seen:
            continue
        seen.add(key)
        out.append({
            "name": name,
            "code": _str(r.get("专业代码")),
            "category": _str(r.get("门类")),
            "level": level,
            "discipline": _str(r.get("专业类")),
            "type": None,
            "notes": _str(r.get("专业备注")),
            "majorLevel": _str(r.get("专业水平")),
            "softRating": _str(r.get("软科评级")),
        })
    return out


# --- group name (directed / special groups only) ---------------------------

# 定向/专项类招生：组的可读标识(定向县/专项类型)在 专业备注 里；普通类组无 prose 名(= 生产 group_name 96% 空)
_DIRECTED_KEYWORDS = ("专项", "定向", "优师", "公费", "民族")


def extract_group_name(r) -> str | None:
    """定向/专项组：返回 专业备注 里的定向县/专项标识；普通类返回 None。best-effort，上线前对生产 group_name 样式校准。"""
    recruit = _str(r.get("招生类型")) or ""
    if not any(k in recruit for k in _DIRECTED_KEYWORDS):
        return None
    return _str(r.get("专业备注"))


# --- enrollment plans (2026 + historical 2025/2024/2023) -------------------

def convert_plans_row(r) -> list:
    """One master row → enrollment plan dicts: 2026 (always) + 2025/2024/2023 (if 计划人数结果N present).

    See spec『enrollment_plans』mapping. Historical years carry per-major planCount
    (计划人数结果1/2/3), fixing the old 2024/2023 group-grain defect.
    """
    shared = {
        "universityEnrollCode": _str(r.get("院校代码")),
        "majorName": _str(r.get("专业名称")),
        "majorCode": _str(r.get("专业代码")),
        "groupCode": _str(r.get("专业组代码")),
        "groupName": extract_group_name(r),
        "groupMajors": None,
        "subjects": _str(r.get("科类")),
        "batch": _str(r.get("批次")),
        "recruitType": _str(r.get("招生类型")),
        "province": "四川",
        "level": _norm_level(r.get("本科/专科")),
        "subjectRequirements": _str(r.get("选科要求")),
        "isNew": _bool_from_str(r.get("是否新增")),
        "oldBatch": _str(r.get("老批次1")),
        "disciplineEval": _str(r.get("学科评估")),
        "isNationalFeature": False,          # 新主表无 国家特色专业 列
        "majorRanking": None,                # 新主表无 专业排名 列
        "majorHonor": None,                  # 新主表无 专业荣誉 列
        "localMasterPoint": _str(r.get("本专业硕士点")),
        "localDoctoralPoint": _str(r.get("本专业博士点")),
        "softRating": _str(r.get("软科评级")),
        "planNotes": _str(r.get("专业备注")),
    }

    plans = [{
        **shared,
        "year": 2026,
        "planCount": _int(r.get("计划人数")),
        "groupPlanCount": _int(r.get("专业组计划人数")),
        "tuition": _int(r.get("学费")),
        "duration": _str(r.get("学制")),
        "groupPurityScore": _float(r.get("专业组干净度")),  # 仅 2026；历史年组成不同，复用此值有误
    }]

    # 历史计划：per-major planCount 来自 计划人数结果1/2/3；缺即跳过该年。groupPlanCount/tuition/duration 不可得。
    for year, col in ((2025, "计划人数结果1"), (2024, "计划人数结果2"), (2023, "计划人数结果3")):
        pc = _int(r.get(col))
        if pc is None:
            continue
        plans.append({
            **shared, "year": year, "planCount": pc,
            "groupPlanCount": None, "tuition": None, "duration": None,
        })

    return plans


# --- admission records (2025 full / 2024-2023 major-level) -----------------

def convert_admissions_row(r) -> list:
    """One master row → admission record dicts for 2025/2024/2023 (suffix 1/2/3).

    2025 carries group-level (专业组录取人数/最低分/位次1) + major-level; 2024/2023 are
    major-level only (matches prod: group_min_rank only filled from 2024 onward).
    A year is skipped if ALL its values are None. See spec『admission_records』.
    """
    shared = {
        "universityEnrollCode": _str(r.get("院校代码")),
        "majorName": _str(r.get("专业名称")),
        "majorCode": _str(r.get("专业代码")),
        "groupCode": _str(r.get("专业组代码")),
        "subjects": _str(r.get("科类")),
        "batch": _str(r.get("批次")),
        "recruitType": _str(r.get("招生类型")),
        "province": "四川",
        "level": _norm_level(r.get("本科/专科")),
    }

    records = []

    vals_25 = {
        "groupAdmissionCount": _int(r.get("专业组录取人数1")),
        "groupMinScore": _int(r.get("专业组最低分1")),
        "groupMinRank": _int(r.get("专业组最低位次1")),
        "majorAdmissionCount": _int(r.get("录取人数1")),
        "majorMinScore": _int(r.get("最低分1")),
        "majorMinRank": _int(r.get("最低位次1")),
        "majorAvgScore": _int(r.get("平均分1")),
        "majorAvgRank": _int(r.get("平均位次1")),
        "majorMaxScore": _int(r.get("最高分1")),
        "majorMaxRank": _int(r.get("最高位次1")),
    }
    if _any_not_none(*vals_25.values()):
        records.append({**shared, "year": 2025, **vals_25})

    # 2024：仅专业级（新主表后缀2 无 最高分/位次、无专业组级列）
    vals_24 = {
        "majorAdmissionCount": _int(r.get("录取人数2")),
        "majorMinScore": _int(r.get("最低分2")),
        "majorMinRank": _int(r.get("最低位次2")),
        "majorAvgScore": _int(r.get("平均分2")),
        "majorAvgRank": _int(r.get("平均位次2")),
    }
    if _any_not_none(*vals_24.values()):
        records.append({**shared, "year": 2024, **vals_24})

    # 2023：专业级（后缀3 有 最高分/位次）
    vals_23 = {
        "majorAdmissionCount": _int(r.get("录取人数3")),
        "majorMinScore": _int(r.get("最低分3")),
        "majorMinRank": _int(r.get("最低位次3")),
        "majorAvgScore": _int(r.get("平均分3")),
        "majorAvgRank": _int(r.get("平均位次3")),
        "majorMaxScore": _int(r.get("最高分3")),
        "majorMaxRank": _int(r.get("最高位次3")),
    }
    if _any_not_none(*vals_23.values()):
        records.append({**shared, "year": 2023, **vals_23})

    return records


# --- universities derived from master (for codes absent in 院校信息表) -------

def derive_universities_from_master(rows, existing_codes) -> list:
    """为 master 里存在、但 existing_codes(院校信息表) 缺失的院校, 从 master 自带院校列派生 university 记录。

    富集字段(满意度/章程细则/QS/logo)主表没有 → 缺省 None；核心字段(省市/层次/性质/硕博点/标签)从主表取。
    按 院校代码 去重。
    """
    seen, out = set(), []
    for r in rows:
        code = _str(r.get("院校代码"))
        name = _str(r.get("院校名称"))
        if not code or not name or code in existing_codes or code in seen:
            continue
        seen.add(code)
        tags = [t.strip() for t in (_str(r.get("院校标签")) or "").split("/") if t.strip()]
        master_cnt = _int(r.get("全校硕士专业数"))
        doctoral_cnt = _int(r.get("全校博士专业数"))
        out.append({
            "enrollCode": code,
            "code": None,                       # 主表无国标码
            "name": name,
            "province": _str(r.get("所在省")),
            "city": _str(r.get("城市")),
            "grade": _str(r.get("城市水平标签")),
            "type": _str(r.get("类型")),
            "level": _str(r.get("院校水平")),
            "runningNature": _str(r.get("公私性质")),
            "department": _str(r.get("隶属单位")),
            "tags": tags,
            "is985": "985" in tags,
            "is211": "211" in tags,
            "isDoubleFirstClass": any("双一流" in t for t in tags),
            "ranking": _str(r.get("院校排名")),
            "admissionGuide": _str(r.get("招生章程")),
            "renameHistory": _str(r.get("更名合并转设")),
            "transferDifficulty": _str(r.get("转专业情况")),
            "postgradRate": _str(r.get("保研率")),
            "disciplineEvaluationLevel": _str(r.get("学科评估")),
            "softRating": _str(r.get("软科评级")),
            "softRanking": _int(r.get("软科排名")),
            "masterProgramCount": master_cnt,
            "doctoralProgramCount": doctoral_cnt,
            "hasMasterProgram": (master_cnt or 0) > 0,
            "hasDoctoralProgram": (doctoral_cnt or 0) > 0,
            "masterPrograms": _str(r.get("全校硕士专业")),
            "doctoralPrograms": _str(r.get("全校博士专业")),
        })
    return out


# --- main ------------------------------------------------------------------

def _read_master_rows(xlsx_path: str):
    """Read the 86-col master; return (list[dict] rows, header index H)."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.active
    it = ws.iter_rows(values_only=True)
    header = next(it)
    H = build_header_index(header)  # fail-fast
    rows = []
    for row in it:
        # skip fully empty rows (first few identity cols all None)
        if not any(row[H[c]] is not None for c in ("院校代码", "专业组代码", "专业名称")):
            continue
        rows.append(row_to_dict(row, H))
    wb.close()
    return rows, H


def _write_json(data, out_path):
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"  written {len(data):>7} → {out_path}")


def main():
    parser = argparse.ArgumentParser(description="Convert 2026 86-col master xlsx to import JSON")
    parser.add_argument("--xlsx", default=None, help="86-col master xlsx path")
    parser.add_argument("--uni-xlsx", default=None, help="院校信息表.xlsx path (universities)")
    parser.add_argument("--out-dir", default=None, help="output dir for JSON files")
    args = parser.parse_args()

    root = Path(__file__).resolve().parent.parent.parent
    xlsx = Path(args.xlsx) if args.xlsx else root / "data" / "03_专家版主表" / "output" / "四川-2026-专家版数据_批次标准化.xlsx"
    uni_xlsx = Path(args.uni_xlsx) if args.uni_xlsx else root / "data" / "03_专家版主表" / "output" / "院校信息表.xlsx"
    out_dir = Path(args.out_dir) if args.out_dir else Path(__file__).resolve().parent / "output_2026"

    if not xlsx.exists():
        print(f"ERROR: {xlsx} not found", file=sys.stderr); sys.exit(1)
    if not uni_xlsx.exists():
        print(f"ERROR: {uni_xlsx} not found", file=sys.stderr); sys.exit(1)

    print("=" * 60)
    print("86-col master → JSON (2026)")
    print("=" * 60)
    print(f"Reading {xlsx} ...")
    rows, _H = _read_master_rows(str(xlsx))
    print(f"  → {len(rows)} data rows")

    majors = convert_majors(rows)
    plans = [p for r in rows for p in convert_plans_row(r)]
    records = [a for r in rows for a in convert_admissions_row(r)]

    # universities: reuse the existing converter on 院校信息表.xlsx (unchanged)
    sys.path.insert(0, str(Path(__file__).resolve().parent))
    from xlsx_to_json import convert_universities
    universities = convert_universities(str(uni_xlsx))

    # 主表里存在但院校信息表缺失的院校 → 从主表派生, 否则其专业行入库被 skip
    existing_codes = {str(u.get("enrollCode")).zfill(4) for u in universities}
    derived = derive_universities_from_master(rows, existing_codes)
    if derived:
        print(f"  + {len(derived)} universities derived from master (院校信息表 缺失)")
        universities = universities + derived

    _write_json(universities, str(out_dir / "universities_enriched.json"))
    _write_json(majors, str(out_dir / "majors_enriched.json"))
    _write_json(plans, str(out_dir / "enrollment_plans_enriched.json"))
    _write_json(records, str(out_dir / "admission_records_filled.json"))

    # --- validation / report ---
    def _by_year(items):
        c = {}
        for it in items:
            c[it["year"]] = c.get(it["year"], 0) + 1
        return dict(sorted(c.items()))

    uni_codes = {str(u.get("enrollCode")).zfill(4) for u in universities}
    master_codes = {str(r.get("院校代码")).zfill(4) for r in rows if r.get("院校代码") is not None}
    unmatched = sorted(master_codes - uni_codes)

    print("\n" + "=" * 60)
    print(f"majors:            {len(majors)}")
    print(f"enrollment_plans:  {len(plans)}  by year={_by_year(plans)}")
    print(f"admission_records: {len(records)}  by year={_by_year(records)}")
    print(f"universities:      {len(universities)}")
    print(f"院校代码 未命中 universities 的: {len(unmatched)}  {unmatched[:20]}")
    print("=" * 60)


if __name__ == "__main__":
    main()
