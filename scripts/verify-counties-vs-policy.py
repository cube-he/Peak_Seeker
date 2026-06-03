"""
对比 data/seed/batch-region-counties.json (从 PDF 招生考试报采集) 与
data/07_政策文件/ 下 4 个官方 xlsx (3/4/5/6 号文件)。

Plan A 校验: 用户要求"附件采集的数据,与07_政策文件 对比验证"。
"""
import json
import os
import re
import sys
from pathlib import Path

import openpyxl

# 强制 UTF-8 输出
sys.stdout.reconfigure(encoding="utf-8")  # type: ignore

REPO = Path(r"C:\Users\17697\Documents\VolunteerHelper")
POLICY_DIR = REPO / "data" / "07_政策文件"
JSON_PATH = REPO / "data" / "seed" / "batch-region-counties.json"


def parse_policy_xlsx(fp: Path) -> dict[str, list[str]]:
    """官方 xlsx 格式: r3 表头, r4+ 是 (地市, 县逗号串) 行"""
    wb = openpyxl.load_workbook(fp, data_only=True)
    ws = wb.active
    result: dict[str, list[str]] = {}
    for r in range(4, ws.max_row + 1):  # type: ignore
        city = ws.cell(r, 1).value  # type: ignore
        counties_raw = ws.cell(r, 2).value  # type: ignore
        if not city or not counties_raw:
            continue
        city = str(city).strip()
        # 县名分隔: 全角顿号、半角逗号、全角逗号
        counties = [
            c.strip()
            for c in re.split(r"[、,，]", str(counties_raw))
            if c.strip()
        ]
        result[city] = counties
    return result


def flatten(by_city: dict[str, list[str]]) -> set[str]:
    return {c for cs in by_city.values() for c in cs}


def main() -> None:
    # 读官方 4 个文件
    files = {
        "policy_2_师范": "2 2024省级公费师范生范围.xlsx",
        "policy_3_集中连片": "3 四川省原集中连片特殊困难县和原国家级扶贫开发重点县名单.xlsx",
        "policy_4_乡村振兴": "4 四川省乡村振兴实施范围县(市、区)名单.xlsx",
        "policy_5_深度贫困": "5 四川省原深度贫困县名单.xlsx",
        "policy_6_民族艰苦": "6 四川省民族地区、原集中连片特殊困难地区和革命老区、艰苦边远地区名单.xlsx",
    }
    policy: dict[str, dict[str, list[str]]] = {}
    for key, fn in files.items():
        fp = POLICY_DIR / fn
        policy[key] = parse_policy_xlsx(fp)
        n = sum(len(v) for v in policy[key].values())
        print(f"{key}: {fn}")
        print(f"  cities={len(policy[key])}  counties={n}")

    # 读我的 JSON
    seed = json.loads(JSON_PATH.read_text(encoding="utf-8"))
    seed_appendices = {
        "appendix_2_119": seed["appendix_2_119"],
        "appendix_4_143": seed["appendix_4_143"],
        "appendix_5_88": seed["appendix_5_88"],
        "appendix_6_45": seed["appendix_6_45"],
    }
    for k, v in seed_appendices.items():
        n = sum(len(c) for c in v["counties"].values())
        print(f"{k}: title={v['title'][:30]}...  cities={len(v['counties'])}  counties={n}")

    print("\n" + "=" * 60)
    print("CROSS-CHECK: 数量级匹配 (按 declared/实际)")
    print("=" * 60)
    # 数量对照
    counts = {k: sum(len(v) for v in pv.values()) for k, pv in policy.items()}
    for k, n in counts.items():
        print(f"{k}: {n}")

    # 求每个 appendix 跟哪个 policy 数量最接近
    print("\n" + "=" * 60)
    print("可能的对应关系 (基于数量相同/接近)")
    print("=" * 60)
    for ak, av in seed_appendices.items():
        ac = sum(len(c) for c in av["counties"].values())
        for pk, pc in counts.items():
            if pc == ac:
                print(f"  {ak} ({ac}) <- {pk} ({pc})  ★ 数量完全相同")

    # 对每对可能匹配,做集合 diff
    print("\n" + "=" * 60)
    print("逐对 diff (集合差异)")
    print("=" * 60)
    candidate_pairs = [
        ("appendix_2_119", "policy_6_民族艰苦"),
        ("appendix_4_143", "policy_2_师范"),  # 师范生 — 关键候选
        ("appendix_5_88", "policy_4_乡村振兴"),
        ("appendix_6_45", "policy_5_深度贫困"),
    ]
    for ak, pk in candidate_pairs:
        a_set = flatten(seed_appendices[ak]["counties"])
        p_set = flatten(policy[pk])
        if not a_set or not p_set:
            continue
        only_seed = a_set - p_set
        only_policy = p_set - a_set
        inter = a_set & p_set
        if len(inter) < min(len(a_set), len(p_set)) * 0.5:
            continue  # 跳过明显不相关的
        print(f"\n{ak} vs {pk}: |seed|={len(a_set)} |policy|={len(p_set)} |共有|={len(inter)}")
        if only_seed:
            print(f"  仅 seed 有 ({len(only_seed)}): {sorted(only_seed)[:10]}{'...' if len(only_seed) > 10 else ''}")
        if only_policy:
            print(f"  仅 policy 有 ({len(only_policy)}): {sorted(only_policy)[:10]}{'...' if len(only_policy) > 10 else ''}")
        if not only_seed and not only_policy:
            print(f"  ✓ 完全一致")

    # 并集分析: seed 大集合 = policy 多个文件的并集吗?
    print("\n" + "=" * 60)
    print("并集分析 (seed = policy_A ∪ policy_B 吗?)")
    print("=" * 60)
    a4 = flatten(seed_appendices["appendix_4_143"]["counties"])
    a5 = flatten(seed_appendices["appendix_5_88"]["counties"])
    p2 = flatten(policy["policy_2_师范"])
    p3 = flatten(policy["policy_3_集中连片"])
    p4 = flatten(policy["policy_4_乡村振兴"])
    p5 = flatten(policy["policy_5_深度贫困"])
    p6 = flatten(policy["policy_6_民族艰苦"])

    combos = [
        ("appendix_4_143", a4, "policy_2_师范",          p2),
        ("appendix_4_143", a4, "policy_2 ∪ policy_4",   p2 | p4),
        ("appendix_4_143", a4, "policy_2 ∪ policy_6",   p2 | p6),
        ("appendix_5_88",  a5, "policy_3 ∪ policy_4",   p3 | p4),
        ("appendix_5_88",  a5, "policy_4 ∪ policy_5",   p4 | p5),
    ]
    for name_a, a, name_p, p in combos:
        if not a or not p:
            continue
        only_a = a - p
        only_p = p - a
        inter = a & p
        if len(inter) < min(len(a), len(p)) * 0.5:
            continue
        print(f"\n{name_a} ({len(a)}) vs {name_p} ({len(p)}): 共有={len(inter)}")
        if only_a:
            print(f"  仅 seed 有 ({len(only_a)}): {sorted(only_a)[:15]}{'...' if len(only_a) > 15 else ''}")
        if only_p:
            print(f"  仅 policy 有 ({len(only_p)}): {sorted(only_p)[:15]}{'...' if len(only_p) > 15 else ''}")
        if not only_a and not only_p:
            print(f"  ✓ 完全一致")


if __name__ == "__main__":
    main()
