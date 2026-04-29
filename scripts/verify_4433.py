# -*- coding: utf-8 -*-
"""Batch 4433 verification script - creates verified xlsx output."""
import openpyxl
from openpyxl.styles import Font, PatternFill
import re
import sys

sys.stdout.reconfigure(encoding='utf-8')

src_path = 'C:/Users/Administrator/Documents/VolunteerHelper/data/13_征集志愿/普通高考/专科批次/4433_2025_历史类_专科批次_征集志愿_第二次/4433_2025_历史类_专科批次_征集志愿_第二次_mimo-v2-omni.xlsx'
out_path = 'C:/Users/Administrator/Documents/VolunteerHelper/data/13_征集志愿/普通高考/专科批次/4433_2025_历史类_专科批次_征集志愿_第二次/4433_2025_历史类_专科批次_征集志愿_第二次_已校验.xlsx'

wb_out = openpyxl.Workbook()
ws_out = wb_out.active
ws_out.title = '征集志愿'

# Style definitions
font_default = Font(name='微软雅黑', size=10)
font_bold = Font(name='微软雅黑', size=10, bold=True)
fill_yellow = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
fill_green = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')

# Headers
headers = ['科类', '招生类型', '院校代码', '院校名称', '办学性质', '院校地址', '院校备注',
           '专业组代码', '再选科目要求', '专业组计划数', '专业代码', '专业名称', '专业备注',
           '专业计划数', '收费标准', '页码', '校正备注']
for c, h in enumerate(headers, 1):
    cell = ws_out.cell(1, c, h)
    cell.font = font_bold

# Load source
wb_src = openpyxl.load_workbook(src_path)
ws_src = wb_src.active

COL_MAP = ['科类', '招生类型', '院校代码', '院校名称', '办学性质', '院校地址', '院校备注',
           '专业组代码', '再选科目要求', '专业组计划数', '专业代码', '专业名称', '专业备注',
           '专业计划数', '收费标准', '页码']


def split_brackets(name, existing_note):
    """Split bracket content from specialty name to specialty note."""
    name = str(name or '')
    idx = -1
    for i, ch in enumerate(name):
        if ch in ('(', '\uff08'):  # ASCII and fullwidth
            idx = i
            break
    if idx == -1:
        return name, str(existing_note or '')

    pure_name = name[:idx].strip()
    bracket_content = name[idx:]

    parts = []
    pattern = r'[(\uff08]([^)\uff09]*)[)\uff09]'
    matches = re.findall(pattern, bracket_content)
    for m in matches:
        parts.append(m)

    note_str = '\uff1b'.join(parts)  # Use fullwidth semicolon

    existing = str(existing_note or '').strip()
    if existing:
        note_str = existing + '\uff1b' + note_str

    return pure_name, note_str


def make_missing_row(inst_code, inst_name, banxue, addr, inst_note,
                     group_code, group_plan, major_code, major_name,
                     major_note, major_plan, fee, page):
    """Create a missing row dict."""
    return {
        '科类': '历史类',
        '招生类型': '普通类专科',
        '院校代码': inst_code,
        '院校名称': inst_name,
        '办学性质': banxue,
        '院校地址': addr,
        '院校备注': inst_note,
        '专业组代码': group_code,
        '再选科目要求': '不限',
        '专业组计划数': group_plan,
        '专业代码': major_code,
        '专业名称': major_name,
        '专业备注': major_note,
        '专业计划数': major_plan,
        '收费标准': fee,
        '页码': page,
    }


# Build missing rows registry
from collections import defaultdict
missing_by_inst = defaultdict(list)

# -- 2180 大连枫叶职业技术学院 (page 5 continuation) --
note_2180 = '执行部委属和外省属高校少数民族地区加分办法1。'
for mc, mn, mnote, mp, mf in [
    ('14', '计算机网络技术', '七顶山校区', 2, 16800),
    ('16', '数字媒体技术', '七顶山校区', 4, 17800),
    ('24', '密码技术应用', '七顶山校区', 2, 18800),
    ('26', '护理', '大黑石校区', 1, 15800),
    ('28', '中医康复技术', '大黑石校区', 1, 13800),
    ('30', '婴幼儿托育服务与管理', '大黑石校区', 2, 13800),
    ('31', '视觉训练与康复', '大黑石校区', 4, 13800),
    ('39', '环境艺术设计', '七顶山校区', 3, 16400),
    ('42', '智慧健康养老服务与管理', '大黑石校区', 2, 12800),
]:
    missing_by_inst['2180'].append(make_missing_row(
        '2180', '大连枫叶职业技术学院', '民办', '辽宁省大连市', note_2180,
        '101', 26, mc, mn, mnote, mp, mf, 5))

# -- 3561 福州软件职业技术学院 (page 8 continuation) --
note_3561 = '执行部委属和外省属高校少数民族地区加分办法1。未来技术产业学院产教融合基地就读。'
for mc, mn, mp, mf in [
    ('39', '工业互联网应用', 8, 13000),
    ('40', '智能光电技术应用', 7, 13000),
]:
    missing_by_inst['3561'].append(make_missing_row(
        '3561', '福州软件职业技术学院', '民办', '福建省福州市', note_3561,
        '101', 22, mc, mn, '', mp, mf, 8))

# -- 4354 湖南应用技术学院 (page 14 continuation) --
note_gen = '执行部委属和外省属高校少数民族地区加分办法1。'
for mc, mn, mp, mf in [
    ('0U', '软件技术', 1, 14800),
    ('0X', '酒店管理与数字化运营', 1, 14200),
]:
    missing_by_inst['4354'].append(make_missing_row(
        '4354', '湖南应用技术学院', '民办', '湖南省常德市', note_gen,
        '101', 3, mc, mn, '', mp, mf, 14))

# -- 4472 广东文理职业学院 (page 15 continuation) --
for mc, mn, mp, mf in [
    ('2F', '电子商务', 3, 17800),
    ('30', '市场营销', 3, 17800),
    ('32', '工商企业管理', 2, 17800),
]:
    missing_by_inst['4472'].append(make_missing_row(
        '4472', '广东文理职业学院', '民办', '广东省湛江市', note_gen,
        '101', 23, mc, mn, '', mp, mf, 15))

# -- 4783 广州华立科技职业学院 (page 16 continuation) --
for mc, mn, mnote, mp, mf in [
    ('YC', '无人机应用技术', '云浮校区', 2, 16800),
    ('YE', '汽车检测与维修技术', '云浮校区', 1, 16800),
    ('YM', '计算机网络技术', '云浮校区', 2, 16800),
]:
    missing_by_inst['4783'].append(make_missing_row(
        '4783', '广州华立科技职业学院', '民办', '广东省广州市', note_gen,
        '101', 13, mc, mn, mnote, mp, mf, 16))

# -- 4913 广东创新科技职业学院 (page 17 continuation) --
for mc, mn, mp, mf in [
    ('48', '数字媒体技术', 4, 19800),
    ('49', '建筑消防技术', 5, 19800),
    ('63', '电子竞技运动与管理', 5, 19800),
]:
    missing_by_inst['4913'].append(make_missing_row(
        '4913', '广东创新科技职业学院', '民办', '广东省东莞市', note_gen,
        '101', 22, mc, mn, '', mp, mf, 17))

# -- 4634 海南东方新丝路职业学院 (page 18 continuation, multiple groups) --
for gc, mc, mn, mp, mf in [
    ('103', '03', '工业机器人技术', 2, 16800),
    ('104', '04', '新能源汽车技术', 2, 16800),
    ('105', '08', '大数据与会计', 2, 16800),
    ('106', '09', '电子商务', 2, 19800),
    ('107', '10', '视觉传达设计', 2, 16800),
    ('108', '11', '数字媒体艺术设计', 2, 16800),
]:
    missing_by_inst['4634'].append(make_missing_row(
        '4634', '海南东方新丝路职业学院', '民办', '海南省东方市', note_gen,
        gc, 2, mc, mn, '', mp, mf, 18))

# -- 5073 重庆科创职业学院 (page 21 continuation) --
note_sc = '执行省属高校少数民族地区加分办法1。'
for gc, gp, mc, mn, mnote, mp, mf in [
    ('103', 3, '22', '建筑室内设计', '', 1, 11000),
    ('104', 1, '28', '网络营销与直播电商', '', 1, 10800),
    ('106', 1, '39', '电子竞技运动与管理', '', 1, 11000),
    ('107', 3, '00', '计算机应用技术', '中外合作办学', 3, 20000),
]:
    missing_by_inst['5073'].append(make_missing_row(
        '5073', '重庆科创职业学院', '民办', '重庆市', note_gen,
        gc, gp, mc, mn, mnote, mp, mf, 21))

# -- 5080 重庆艺术工程职业学院 (page 22 continuation) --
for mc, mn, mp, mf in [
    ('D4', '数字化设计与制造技术', 5, 12000),
    ('D5', '工业设计', 2, 12000),
    ('D6', '智能物流技术', 5, 12000),
    ('D7', '空中乘务', 7, 12000),
    ('D8', '民航安全技术管理', 7, 12000),
]:
    missing_by_inst['5080'].append(make_missing_row(
        '5080', '重庆艺术工程职业学院', '民办', '重庆市', note_gen,
        '101', 59, mc, mn, '', mp, mf, 22))
# 5080 group 102
missing_by_inst['5080'].append(make_missing_row(
    '5080', '重庆艺术工程职业学院', '民办', '重庆市', note_gen,
    '102', 5, '85', '建筑设计', '中外合作办学', 5, 16000, 22))

# -- 5160 四川国际标榜职业学院 (page 24 continuation - massive) --
note_sc2 = '执行省属高校少数民族地区加分办法2。'
for mc, mn, mp, mf in [
    ('26', '建筑室内设计', 14, 15900),
    ('27', '建筑设计', 10, 15900),
    ('29', '园林工程技术', 23, 15900),
    ('30', '电子商务', 8, 12720),
    ('33', '现代文秘', 11, 12000),
    ('34', '大数据与财务管理', 11, 12720),
    ('35', '工程造价', 13, 15900),
    ('36', '工商企业管理', 4, 12720),
    ('37', '建设工程管理', 11, 15900),
    ('38', '市场营销', 8, 12720),
    ('43', '中文', 22, 12720),
    ('44', '商务英语', 3, 15900),
    ('46', '旅游管理', 2, 14840),
    ('47', '酒店管理与数字化运营', 14, 15900),
    ('BQ', '婴幼儿托育服务与管理', 7, 15900),
    ('KN', '烹饪工艺与营养', 22, 15000),
    ('T2', '体育艺术表演', 1, 15900),
    ('TB', '国际标准舞', 1, 15900),
    ('W1', '表演艺术', 1, 15900),
    ('YG', '网络直播与运营', 8, 15900),
    ('YH', '人工智能技术应用', 10, 15900),
    ('YJ', '商务数据分析与应用', 12, 15900),
    ('YL', '数字媒体技术', 7, 15900),
]:
    mnote = ''
    if mc == '46':
        mnote = '涉外旅游'
    missing_by_inst['5160'].append(make_missing_row(
        '5160', '四川国际标榜职业学院', '民办', '四川省成都市', note_sc2,
        '101', 308, mc, mn, mnote, mp, mf, 24))

# -- 5164 四川托普信息技术职业学院 (page 25 continuation) --
for mc, mn, mp, mf in [
    ('13', '电子信息工程技术', 4, 13100),
    ('14', '智能机器人技术', 6, 13100),
    ('15', '光伏工程技术', 3, 13100),
    ('17', '机电一体化技术', 388, 13100),
    ('18', '大数据与会计', 2, 11800),
    ('19', '电子商务', 316, 11800),
    ('20', '市场营销', 5, 11800),
    ('21', '现代物流管理', 1, 11800),
    ('22', '网络营销与直播电商', 3, 11800),
    ('23', '工程造价', 7, 13100),
    ('24', '建筑工程技术', 7, 13100),
    ('25', '建筑消防技术', 5, 13100),
    ('26', '婴幼儿托育服务与管理', 8, 11800),
    ('27', '学前教育', 32, 10900),
    ('28', '商务英语', 3, 11800),
    ('30', '建筑室内设计', 6, 13100),
    ('32', '影视多媒体技术', 1, 12360),
    ('33', '园林工程技术', 8, 13100),
]:
    missing_by_inst['5164'].append(make_missing_row(
        '5164', '四川托普信息技术职业学院', '民办', '四川省成都市', note_sc2,
        '101', 1010, mc, mn, '', mp, mf, 25))

# -- 5169 雅安职业技术学院 (page 26 continuation) --
for mc, mn, mnote, mp, mf in [
    ('G2', '旅游管理', '', 20, 5280),
    ('G4', '酒店管理与数字化运营', '', 18, 5280),
]:
    missing_by_inst['5169'].append(make_missing_row(
        '5169', '雅安职业技术学院', '公办', '四川省雅安市', note_sc2,
        '102', 94, mc, mn, mnote, mp, mf, 26))

# -- 5186 四川科技职业学院 (page 28 continuation - massive) --
for mc, mn, mnote, mp, mf in [
    ('T6', '民航运输服务', '', 1, 15900),
    ('V2', '空中乘务', '', 2, 15900),
    ('V5', '烹饪工艺与营养', '', 19, 14840),
    ('W1', '中西面点工艺', '', 12, 14840),
    ('W5', '酒店管理与数字化运营', '', 12, 14840),
    ('W8', '酒店管理与数字化运营', '民宿产品开发', 9, 14840),
    ('X4', '学前教育', '', 22, 14840),
    ('X6', '早期教育', '', 12, 14840),
]:
    missing_by_inst['5186'].append(make_missing_row(
        '5186', '四川科技职业学院', '民办', '四川省眉山市', note_sc2,
        '101', 965, mc, mn, mnote, mp, mf, 28))

# -- 5189 四川华新现代职业学院 (page 29 continuation) --
for mc, mn, mnote, mp, mf in [
    ('12', '大数据技术', '', 9, 14000),
    ('14', '电气自动化技术', '', 42, 12500),
    ('15', '新能源汽车技术', '', 49, 12500),
    ('17', '服装设计与工艺', '', 3, 12500),
    ('18', '动漫制作技术', '', 4, 12500),
    ('19', '旅游管理', '', 2, 12500),
    ('20', '烹饪工艺与营养', '', 14, 12500),
]:
    missing_by_inst['5189'].append(make_missing_row(
        '5189', '四川华新现代职业学院', '民办', '四川省成都市', note_sc2,
        '101', 302, mc, mn, mnote, mp, mf, 29))

# -- 5195 四川三河职业学院 (page 33 continuation - massive) --
for mc, mn, mp, mf in [
    ('09', '汽车检测与维修技术', 7, 10600),
    ('10', '新能源汽车技术', 5, 10600),
    ('11', '智能机电技术', 8, 10600),
    ('12', '大数据与会计', 7, 10600),
    ('13', '行政管理', 6, 10600),
    ('14', '旅游管理', 14, 10600),
    ('15', '现代物流管理', 16, 10600),
    ('16', '市场营销', 10, 10600),
    ('17', '护理', 11, 10600),
    ('18', '康复治疗技术', 28, 10600),
    ('19', '药品经营与管理', 27, 10600),
    ('20', '中药学', 22, 10600),
    ('21', '助产', 37, 10600),
    ('22', '学前教育', 18, 10600),
    ('23', '动物医学', 20, 10600),
    ('24', '现代农业技术', 14, 10600),
    ('25', '智能建造技术', 16, 10600),
    ('26', '无人机应用技术', 10, 10600),
    ('27', '婴幼儿托育服务与管理', 24, 10600),
    ('28', '网络营销与直播电商', 28, 10600),
]:
    missing_by_inst['5195'].append(make_missing_row(
        '5195', '四川三河职业学院', '民办', '四川省泸州市', note_sc2,
        '101', 400, mc, mn, '', mp, mf, 33))

# -- 5651 巴中职业技术学院 (page 35 continuation) --
for mc, mn, mp, mf in [
    ('16', '法律事务', 10, 11800),
    ('17', '工程造价', 2, 11800),
    ('21', '旅游管理', 6, 11800),
    ('23', '动物医学', 20, 11800),
    ('24', '计算机网络技术', 10, 11800),
    ('28', '风景园林设计', 3, 11800),
    ('30', '大数据与财务管理', 3, 11800),
    ('31', '大数据与会计', 4, 11800),
    ('33', '畜牧兽医', 20, 11800),
]:
    missing_by_inst['5651'].append(make_missing_row(
        '5651', '巴中职业技术学院', '民办', '四川省巴中市', note_sc2,
        '101', 295, mc, mn, '', mp, mf, 35))

# -- 5653 四川电子机械职业技术学院 (page 36 continuation - massive) --
for mc, mn, mnote, mp, mf in [
    ('05', '现代通信技术', '', 3, 11600),
    ('06', '现代移动通信技术', '', 2, 12290),
    ('07', '计算机应用技术', '', 9, 12290),
    ('08', '软件技术', '', 9, 12290),
    ('09', '移动应用开发', '', 2, 12290),
    ('10', '应用电子技术', '国产芯片应用与开发', 2, 12290),
    ('11', '智能产品开发与应用', '系统适配与测试', 2, 12290),
    ('12', '计算机应用技术', '国产软件与数据库应用', 3, 12290),
    ('13', '机械设计与制造', '', 27, 12290),
    ('14', '机械制造及自动化', '', 38, 12290),
    ('15', '数控技术', '', 18, 12290),
    ('16', '机电一体化技术', '', 10, 12290),
    ('17', '工业机器人技术', '', 12, 12290),
    ('18', '智能控制技术', '', 4, 12290),
    ('19', '智能机电技术', '', 3, 12290),
    ('20', '人工智能技术应用', '', 3, 12290),
    ('21', '智能机器人技术', '', 2, 12290),
    ('22', '智能产品开发与应用', '', 2, 12290),
    ('23', '数字化设计与制造技术', '', 2, 12290),
    ('24', '大数据与会计', '', 11, 11600),
    ('25', '工商企业管理', '', 2, 11600),
    ('26', '现代物流管理', '', 2, 11600),
    ('27', '市场营销', '', 2, 11600),
    ('28', '电子商务', '', 2, 11600),
    ('29', '婴幼儿托育服务与管理', '', 2, 11600),
    ('30', '融媒体技术与运营', '', 2, 12290),
    ('31', '市场营销', '网络营销与策划', 2, 11600),
    ('32', '电子商务', '新媒体电商', 2, 11600),
    ('33', '建筑工程技术', '', 3, 11600),
    ('34', '建设工程管理', '', 2, 11600),
    ('35', '工程造价', '', 2, 11600),
    ('36', '建筑设计', '', 2, 12290),
    ('39', '数字媒体技术', '', 4, 11600),
]:
    missing_by_inst['5653'].append(make_missing_row(
        '5653', '四川电子机械职业技术学院', '民办', '四川省绵阳市', note_sc2,
        '101', 234, mc, mn, mnote, mp, mf, 36))

# -- 5655 川南幼儿师范高等专科学校 (page 37 continuation) --
for mc, mn, mnote, mp, mf in [
    ('06', '旅游管理', '', 10, 4800),
    ('07', '民宿管理与运营', '', 15, 4800),
    ('08', '电子商务', '', 47, 4800),
    ('09', '大数据与会计', '', 8, 4800),
    ('10', '现代物流管理', '', 19, 4800),
    ('11', '学前教育', '师范', 6, 4800),
    ('12', '计算机应用技术', '', 13, 5200),
    ('13', '环境监测技术', '', 21, 5200),
]:
    missing_by_inst['5655'].append(make_missing_row(
        '5655', '川南幼儿师范高等专科学校', '公办', '四川省内江市', note_sc2,
        '101', 396, mc, mn, mnote, mp, mf, 37))

# -- 5698 甘孜职业学院 (page 38 continuation) --
for mc, mn, mp, mf in [
    ('06', '旅游管理', 10, 4800),
    ('07', '民宿管理与运营', 15, 4800),
    ('08', '电子商务', 47, 4800),
    ('09', '大数据与会计', 8, 4800),
    ('10', '现代物流管理', 19, 4800),
    ('11', '学前教育', 6, 4800),
    ('12', '计算机应用技术', 13, 5200),
    ('13', '环境监测技术', 21, 5200),
]:
    # Actually 5698 is 甘孜职业学院 - need to verify from image
    pass
# Skip 5698 as the image showed its data correctly

# -- 5781 天府新区通用航空职业学院 (page 40 continuation) --
for mc, mn, mp, mf in [
    ('16', '现代物流管理', 4, 12720),
    ('18', '大数据与会计', 3, 10490),
    ('19', '护理', 23, 14300),
    ('20', '中药学', 6, 14300),
    ('21', '医学美容技术', 6, 14300),
    ('22', '物联网应用技术', 4, 11660),
    ('23', '数字媒体技术', 1, 12720),
    ('25', '软件技术', 3, 11660),
    ('26', '汽车制造与试验技术', 10, 11660),
    ('27', '大数据技术', 4, 12720),
    ('30', '建筑工程技术', 16, 14300),
    ('31', '智慧健康养老服务与管理', 15, 14300),
]:
    missing_by_inst['5781'].append(make_missing_row(
        '5781', '天府新区通用航空职业学院', '民办', '四川省眉山市', note_sc2,
        '101', 170, mc, mn, '', mp, mf, 40))

# -- 5783 天府新区航空旅游职业学院 (page 41 continuation - massive) --
for mc, mn, mp, mf in [
    ('0E', '宠物医疗技术', 15, 13800),
    ('0F', '高速铁路客运服务', 7, 13990),
    ('0G', '城市轨道交通运营管理', 9, 13990),
    ('0H', '旅游管理', 25, 12720),
    ('0J', '会展策划与管理', 25, 12720),
    ('0K', '飞机机电设备维修', 11, 17170),
    ('0L', '飞机电子设备维修', 17, 17170),
    ('0M', '无人机应用技术', 7, 13990),
    ('0N', '电子商务', 24, 13990),
    ('0P', '软件技术', 22, 13990),
    ('0Q', '计算机应用技术', 25, 13990),
    ('0R', '法律事务', 19, 12720),
    ('0S', '数字媒体技术', 18, 13990),
]:
    missing_by_inst['5783'].append(make_missing_row(
        '5783', '天府新区航空旅游职业学院', '民办', '四川省眉山市', note_sc2,
        '101', 489, mc, mn, '', mp, mf, 41))

# -- 5784 德阳科贸职业学院 (page 41 already captured) --
# Actually 5784 starts on page 41 - check if continuation on page 42 is missing
# No, 5784 has all data on page 41 based on image

# -- 5785 天府新区信息职业学院 (page 43 continuation - massive) --
for mc, mn, mnote, mp, mf in [
    ('45', '电子竞技运动与管理', '', 8, 14620),
    ('46', '休闲体育', '', 7, 14620),
    ('47', '休闲体育', '武术', 5, 14620),
    ('48', '智慧健康养老服务与管理', '营养膳食', 4, 13560),
    ('49', '智慧健康养老服务与管理', '', 3, 13560),
    ('50', '婴幼儿托育服务与管理', '', 4, 14620),
    ('51', '金融科技应用', '互联网金融', 5, 13560),
    ('KM', '机电一体化技术', '', 7, 14620),
    ('KQ', '智慧旅游技术应用', '', 8, 13560),
    ('LH', '网络直播与运营', '', 8, 14620),
]:
    missing_by_inst['5785'].append(make_missing_row(
        '5785', '天府新区信息职业学院', '民办', '四川省眉山市', note_sc2,
        '101', 363, mc, mn, mnote, mp, mf, 43))

# -- 5787 德阳城市轨道交通职业学院 (page 44 continuation) --
for mc, mn, mnote, mp, mf in [
    ('05', '城市轨道交通供配电技术', '', 15, 13990),
    ('06', '铁道车辆技术', '', 11, 13990),
    ('07', '电气自动化技术', '', 2, 13990),
    ('08', '城市轨道交通车辆制造与维护', '', 10, 13990),
    ('11', '城市轨道交通工程技术', '', 11, 13990),
    ('13', '高速铁路综合维修技术', '', 16, 13990),
    ('16', '铁道工程技术', '', 10, 13990),
    ('19', '环境艺术设计', '', 10, '待定'),
    ('20', '城市轨道交通运营管理', '', 9, 12720),
    ('25', '旅游管理', '', 10, '待定'),
    ('26', '城市轨道交通通信信号技术', '', 15, 13990),
    ('27', '智能控制技术', '', 11, 13990),
    ('28', '人工智能技术应用', '', 11, 13990),
    ('30', '融媒体技术与运营', '', 10, '待定'),
]:
    missing_by_inst['5787'].append(make_missing_row(
        '5787', '德阳城市轨道交通职业学院', '民办', '四川省德阳市', note_sc2,
        '101', 193, mc, mn, mnote, mp, mf, 44))

# -- 5793 南充科技职业学院 (page 45 continuation) --
for mc, mn, mnote, mp, mf in [
    ('09', '医学美容技术', '', 12, 10600),
    ('10', '康复治疗技术', '', 51, 10600),
    ('11', '应急救援技术', '', 7, 11660),
    ('12', '应急救援技术', '生产事故救援方向', 5, 11660),
    ('13', '消防救援技术', '', 6, 11000),
    ('14', '消防救援技术', '城市消防救援方向', 4, 11000),
    ('15', '建筑工程技术', '', 3, 11660),
    ('16', '建筑工程技术', '数字孪生工程修复技术方向', 5, 11660),
    ('17', '建筑消防技术', '', 4, 11660),
    ('18', '建筑消防技术', '智慧消防方向', 4, 11660),
    ('19', '建筑消防技术', '消防工程管理方向', 5, 11660),
    ('20', '无人机应用技术', '', 8, 12800),
    ('21', '无人机应用技术', '无人机救援方向', 6, 12800),
    ('22', '机械装备制造技术', '', 5, 12800),
    ('23', '智能机器人技术', '', 9, 11660),
    ('24', '新能源汽车技术', '', 5, 12800),
    ('25', '智能网联汽车技术', '', 4, 12800),
    ('26', '物联网应用技术', '', 4, 11660),
    ('27', '大数据技术', '', 3, 11660),
    ('29', '人工智能技术应用', '', 2, 11660),
    ('30', '人工智能技术应用', '生成式人工智能方向', 5, 11660),
    ('31', '人工智能技术应用', '元宇宙方向', 5, 11660),
    ('32', '现代移动通信技术', '', 3, 11000),
    ('33', '现代移动通信技术', '国产通信设备维护方向', 5, 11000),
    ('34', '电子商务', '', 4, 12800),
    ('35', '农产品加工与质量检测', '', 3, 12800),
    ('36', '高速铁路客运服务', '', 5, 10600),
    ('37', '大数据与财务管理', '', 9, 11800),
    ('38', '旅游管理', '', 9, 11800),
    ('39', '烹饪工艺与营养', '', 3, 11800),
    ('40', '畜牧兽医', '', 14, 12800),
]:
    missing_by_inst['5793'].append(make_missing_row(
        '5793', '南充科技职业学院', '民办', '四川省南充市', note_sc2,
        '101', 407, mc, mn, mnote, mp, mf, 45))

# -- 5794 攀枝花攀西职业学院 (page 46 continuation) --
for mc, mn, mp, mf in [
    ('0D', '网络营销与直播电商', 13, 12000),
    ('0E', '数字媒体艺术设计', 20, 12000),
    ('0F', '环境艺术设计', 19, 12000),
    ('0G', '社会体育', 13, 12000),
    ('0H', '行政管理', 37, 10000),
]:
    missing_by_inst['5794'].append(make_missing_row(
        '5794', '攀枝花攀西职业学院', '民办', '四川省攀枝花市', note_sc2,
        '101', 281, mc, mn, '', mp, mf, 46))

# -- 5803 绵阳飞行职业学院 (page 47 continuation) --
for mc, mn, mp, mf in [
    ('03', '无人机应用技术', 38, 14700),
    ('07', '空中乘务', 28, 15500),
    ('08', '民航空中安全保卫', 7, 15500),
    ('09', '环境艺术设计', 10, 13300),
    ('10', '休闲体育', 7, 11000),
    ('11', '社区康复', 6, 11000),
    ('12', '民航安全技术管理', 12, 13400),
    ('13', '安全智能监测技术', 8, 13400),
    ('17', '跨境电子商务', 4, 13400),
    ('19', '数字媒体技术', 2, 13400),
    ('21', '消防救援技术', 1, 15500),
    ('22', '建筑消防技术', 10, 13400),
]:
    missing_by_inst['5803'].append(make_missing_row(
        '5803', '绵阳飞行职业学院', '民办', '四川省绵阳市', note_sc2,
        '101', 133, mc, mn, '', mp, mf, 47))

# -- 5882 遂宁职业学院 (page 48 continuation) --
for mc, mn, mp, mf in [
    ('05', '建筑消防技术', 14, 13500),
    ('06', '高速铁路客运服务', 33, 13500),
    ('07', '城市轨道交通管理', 9, 13500),
    ('08', '智能机电技术', 25, 13500),
    ('09', '网络营销与直播电商', 9, 13500),
    ('10', '艺术设计', 16, 13500),
    ('11', '电子信息工程技术', 13, 13500),
    ('12', '计算机应用技术', 9, 13500),
    ('13', '软件技术', 27, 13500),
    ('15', '人工智能技术应用', 16, 13500),
]:
    missing_by_inst['5882'].append(make_missing_row(
        '5882', '遂宁职业学院', '民办', '四川省遂宁市', note_sc2,
        '101', 301, mc, mn, '', mp, mf, 48))

# -- 5376 云南工程职业学院 (page 51 continuation) --
for mc, mn, mp, mf in [
    ('01', '电气自动化技术', 3, 14000),
    ('03', '新能源汽车技术', 2, 14000),
    ('05', '机电一体化技术', 1, 14000),
    ('07', '护理', 2, 14000),
    ('08', '医学影像技术', 1, 14000),
]:
    missing_by_inst['5376'].append(make_missing_row(
        '5376', '云南工程职业学院', '民办', '云南省昆明市', note_gen,
        '101', 9, mc, mn, '', mp, mf, 51))

# -- 6164 西安思源学院 (page 52 continuation) --
for mc, mn, mp, mf in [
    ('68', '护理', 2, 17670),
    ('69', '医学检验技术', 2, 17670),
    ('70', '口腔医学技术', 2, 17670),
    ('73', '大数据与会计', 2, 16720),
]:
    missing_by_inst['6164'].append(make_missing_row(
        '6164', '西安思源学院', '民办', '陕西省西安市', note_gen,
        '101', 8, mc, mn, '', mp, mf, 52))

# ---- Known specific field corrections ----
# Row 209: 4354 专业代码 "05" -> "0S"
field_corrections = {
    209: [('专业代码', '0S', '专业代码05→0S')]
}

# ---- Write output ----
out_row = 2
stats = {'matched': 0, 'corrected': 0, 'missing': 0, 'bracket_split': 0}
inserted = set()

for src_row in range(2, ws_src.max_row + 1):
    vals = {}
    for c, key in enumerate(COL_MAP, 1):
        vals[key] = ws_src.cell(src_row, c).value
    # Ensure string types
    for k in ('院校代码', '专业组代码', '专业代码'):
        vals[k] = str(vals[k] or '')
    for k in ('科类', '招生类型', '院校名称', '办学性质', '院校地址', '院校备注',
              '再选科目要求', '专业名称', '专业备注'):
        vals[k] = str(vals[k] or '')

    notes = []

    # 1. Normalize 招生类型
    if vals['招生类型'] != '普通类专科':
        old = vals['招生类型']
        vals['招生类型'] = '普通类专科'
        if old not in ('普通类', ''):
            notes.append(f'招生类型{old}→普通类专科')

    # 2. Apply field corrections
    if src_row in field_corrections:
        for field, new_val, note in field_corrections[src_row]:
            vals[field] = new_val
            notes.append(note)

    # 3. Bracket splitting
    orig_name = vals['专业名称']
    if '(' in orig_name or '\uff08' in orig_name:
        new_name, new_note = split_brackets(orig_name, vals['专业备注'])
        if new_name != orig_name:
            vals['专业名称'] = new_name
            vals['专业备注'] = new_note
            notes.append('括号拆分')
            stats['bracket_split'] += 1

    # Write row
    for c, key in enumerate(COL_MAP, 1):
        cell = ws_out.cell(out_row, c, vals[key])
        cell.font = font_default
        # Yellow fill for corrected cells
        if src_row in field_corrections:
            for field, _, _ in field_corrections[src_row]:
                if key == field:
                    cell.fill = fill_yellow

    note_text = '；'.join(notes) if notes else ''
    cell_note = ws_out.cell(out_row, 17, note_text)
    cell_note.font = font_default

    if notes:
        stats['corrected'] += 1
    else:
        stats['matched'] += 1

    out_row += 1

    # Insert missing rows after last entry of institution
    inst_code = vals['院校代码']
    if inst_code in missing_by_inst and inst_code not in inserted:
        next_code = None
        if src_row < ws_src.max_row:
            next_code = str(ws_src.cell(src_row + 1, 3).value or '')
        if next_code != inst_code:
            for mr in missing_by_inst[inst_code]:
                for c, key in enumerate(COL_MAP, 1):
                    cell = ws_out.cell(out_row, c, mr[key])
                    cell.font = font_default
                    cell.fill = fill_green
                ws_out.cell(out_row, 17, '缺漏行:跨页续写丢失').font = font_default
                ws_out.cell(out_row, 17).fill = fill_green
                out_row += 1
                stats['missing'] += 1
            inserted.add(inst_code)

# Save
wb_out.save(out_path)

print(f'\n=== 校验完成 ===')
print(f'输出文件: {out_path}')
print(f'总行数: {out_row - 2}')
print(f'  匹配行: {stats["matched"]}')
print(f'  修正行: {stats["corrected"]}')
print(f'  缺漏补充行: {stats["missing"]}')
print(f'  括号拆分: {stats["bracket_split"]}')
