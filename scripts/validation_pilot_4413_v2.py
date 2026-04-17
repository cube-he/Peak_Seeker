# -*- coding: utf-8 -*-
"""
4413 Validation Pilot v2: fix two systematic issues from v1.

Fixes:
  (1) Bracket attribution — location/campus/training-location brackets belong
      to 专业备注, NOT 专业名称 (v1 merged them into 专业名称 wrongly).
  (2) Forward-fill 招生类型 — when a section title is not reprinted after a
      page break, downstream rows must inherit the last-seen section title
      (v1 / both engines mislabeled 5125/5130/5137/5142 rows as "普通类",
      but the image has NO "普通类" subsection — those rows belong to
      "区域教育均衡发展专项计划").

Authority: images > corrected > mimo-v2-omni.

Output colors:
  - YELLOW: v1 already corrected, v2 retains same value
  - ORANGE: v2 re-corrected (delta vs v1)
  - RED:    image unclear, manual review needed (none for 4413)
"""
import os
import sys
import math
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font

sys.stdout.reconfigure(encoding='utf-8')

BASE = r'C:\Users\Administrator\Documents\VolunteerHelper\data\13_征集志愿\普通高考\少数民族\4413_征集志愿_第一次_2025_本科批'
MIMO = os.path.join(BASE, '4413_征集志愿_第一次_2025_mimo-v2-omni.xlsx')
CORR = os.path.join(BASE, '4413_征集志愿_第一次_2025_mimo-v2-omni_corrected.xlsx')
V1   = os.path.join(BASE, '4413_征集志愿_第一次_2025_mimo-v2-omni_verified.xlsx')
OUT  = os.path.join(BASE, '4413_征集志愿_第一次_2025_mimo-v2-omni_verified_v2.xlsx')

COLS = ['科类', '招生类型', '院校代码', '院校名称', '办学性质', '院校地址', '院校备注',
        '专业组代码', '再选科目要求', '专业组计划数', '专业代码', '专业名称', '专业备注',
        '专业计划数', '收费标准', '页码']

# -------- Ground truth rows (v2, strictly following image 003 + 004) --------
# Rule applied:
#   - Location brackets (雅安市/凉山州/乐山市/阿坝州 ...) → 专业备注
#   - Campus brackets (宜宾校区/江北校区A区 ...) → 专业备注
#   - Training-location brackets (预科培养地点为XXX) → 专业备注
#   - Type brackets like (师范) → 专业备注 (image lists it separately)
#   - Sub-discipline brackets (应用心理学/护理学/旅游管理) that are part of
#     the parent program name (少数民族预科) → 专业名称
#
# 招生类型 forward-fill (verified against image):
#   Historical section: 省属高校少数民族预科 (Y5, Y8 on page 4)
#   Physical §1: 区域教育均衡发展专项计划 (page 4 + continued on page 5 until §2 heading)
#   Physical §2: 省属高校少数民族预科 (Y9, Y8, Y1 on page 5)
#   NO "普通类" subsection exists in image for 4413.
GT_V2 = [
    # --- 历史类 省属高校少数民族预科 (page 4) ---
    {'科类': '历史类', '招生类型': '省属高校少数民族预科', '院校代码': 5137, '院校名称': '成都大学', '办学性质': '公办',
     '院校地址': '四川省成都市', '院校备注': None, '专业组代码': 108, '再选科目要求': '不限', '专业组计划数': 1,
     '专业代码': 'Y5', '专业名称': '少数民族预科(旅游管理)', '专业备注': '预科培养地点为西昌学院。',
     '专业计划数': 1, '收费标准': 5500, '页码': 4},
    {'科类': '历史类', '招生类型': '省属高校少数民族预科', '院校代码': 5137, '院校名称': '成都大学', '办学性质': '公办',
     '院校地址': '四川省成都市', '院校备注': None, '专业组代码': 109, '再选科目要求': '化学', '专业组计划数': 2,
     '专业代码': 'Y8', '专业名称': '少数民族预科(护理学)', '专业备注': '预科培养地点为西昌学院。',
     '专业计划数': 2, '收费标准': 5500, '页码': 4},

    # --- 物理类 · §1 区域教育均衡发展专项计划 (page 4) ---
    {'科类': '物理类', '招生类型': '区域教育均衡发展专项计划', '院校代码': 5108, '院校名称': '四川轻化工大学', '办学性质': '公办',
     '院校地址': '四川省自贡市', '院校备注': None, '专业组代码': 812, '再选科目要求': '化学', '专业组计划数': 1,
     '专业代码': '8T', '专业名称': '食品科学与工程', '专业备注': '雅安市',
     '专业计划数': 1, '收费标准': 5980, '页码': 4},
    {'科类': '物理类', '招生类型': '区域教育均衡发展专项计划', '院校代码': 5119, '院校名称': '成都医学院', '办学性质': '公办',
     '院校地址': '四川省成都市', '院校备注': '色盲、色弱考生不予录取。新都校区就读。', '专业组代码': 103, '再选科目要求': '化学', '专业组计划数': 2,
     '专业代码': 'J8', '专业名称': '护理学', '专业备注': '凉山州',
     '专业计划数': 1, '收费标准': 6670, '页码': 4},
    {'科类': '物理类', '招生类型': '区域教育均衡发展专项计划', '院校代码': 5119, '院校名称': '成都医学院', '办学性质': '公办',
     '院校地址': '四川省成都市', '院校备注': '色盲、色弱考生不予录取。新都校区就读。', '专业组代码': 103, '再选科目要求': '化学', '专业组计划数': 2,
     '专业代码': 'JD', '专业名称': '智能医学工程', '专业备注': '凉山州',
     '专业计划数': 1, '收费标准': 6670, '页码': 4},
    {'科类': '物理类', '招生类型': '区域教育均衡发展专项计划', '院校代码': 5123, '院校名称': '绵阳师范学院', '办学性质': '公办',
     '院校地址': '四川省绵阳市', '院校备注': None, '专业组代码': 108, '再选科目要求': '不限', '专业组计划数': 1,
     '专业代码': 'QH', '专业名称': '物流管理', '专业备注': '雅安市',
     '专业计划数': 1, '收费标准': 5520, '页码': 4},

    # --- 物理类 · §1 区域教育均衡发展专项计划 (continued on page 5, no heading reprinted) ---
    {'科类': '物理类', '招生类型': '区域教育均衡发展专项计划', '院校代码': 5125, '院校名称': '宜宾学院', '办学性质': '公办',
     '院校地址': '四川省宜宾市', '院校备注': None, '专业组代码': 107, '再选科目要求': '化学', '专业组计划数': 1,
     '专业代码': '5D', '专业名称': '制药工程', '专业备注': '雅安市；江北校区A区',
     '专业计划数': 1, '收费标准': 5980, '页码': 5},
    {'科类': '物理类', '招生类型': '区域教育均衡发展专项计划', '院校代码': 5130, '院校名称': '阿坝师范学院', '办学性质': '公办',
     '院校地址': '四川省阿坝州', '院校备注': None, '专业组代码': 104, '再选科目要求': '不限', '专业组计划数': 1,
     '专业代码': '2G', '专业名称': '学前教育', '专业备注': '雅安市；师范',
     '专业计划数': 1, '收费标准': 4800, '页码': 5},
    {'科类': '物理类', '招生类型': '区域教育均衡发展专项计划', '院校代码': 5137, '院校名称': '成都大学', '办学性质': '公办',
     '院校地址': '四川省成都市', '院校备注': None, '专业组代码': 105, '再选科目要求': '化学', '专业组计划数': 1,
     '专业代码': '37', '专业名称': '环境工程', '专业备注': '凉山州',
     '专业计划数': 1, '收费标准': 5980, '页码': 5},
    {'科类': '物理类', '招生类型': '区域教育均衡发展专项计划', '院校代码': 5142, '院校名称': '成都工业学院', '办学性质': '公办',
     '院校地址': '四川省成都市', '院校备注': None, '专业组代码': 146, '再选科目要求': '化学', '专业组计划数': 1,
     '专业代码': '47', '专业名称': '汽车服务工程', '专业备注': '乐山市；宜宾校区；入学后不允许跨校区转专业。',
     '专业计划数': 1, '收费标准': 5200, '页码': 5},
    {'科类': '物理类', '招生类型': '区域教育均衡发展专项计划', '院校代码': 5142, '院校名称': '成都工业学院', '办学性质': '公办',
     '院校地址': '四川省成都市', '院校备注': None, '专业组代码': 147, '再选科目要求': '化学', '专业组计划数': 3,
     '专业代码': '48', '专业名称': '汽车服务工程', '专业备注': '雅安市；宜宾校区；入学后不允许跨校区转专业。',
     '专业计划数': 3, '收费标准': 5200, '页码': 5},

    # --- 物理类 · §2 省属高校少数民族预科 (page 5) ---
    {'科类': '物理类', '招生类型': '省属高校少数民族预科', '院校代码': 5119, '院校名称': '成都医学院', '办学性质': '公办',
     '院校地址': '四川省成都市', '院校备注': None, '专业组代码': 109, '再选科目要求': '不限', '专业组计划数': 2,
     '专业代码': 'Y9', '专业名称': '少数民族预科(应用心理学)',
     '专业备注': '预科培养地点为四川民族学院。色盲、色弱考生不予录取。',
     '专业计划数': 2, '收费标准': 5980, '页码': 5},
    {'科类': '物理类', '招生类型': '省属高校少数民族预科', '院校代码': 5119, '院校名称': '成都医学院', '办学性质': '公办',
     '院校地址': '四川省成都市', '院校备注': None, '专业组代码': 110, '再选科目要求': '化学', '专业组计划数': 2,
     '专业代码': 'Y8', '专业名称': '少数民族预科(护理学)',
     '专业备注': '预科培养地点为四川民族学院。色盲、色弱考生不予录取。',
     '专业计划数': 2, '收费标准': 6670, '页码': 5},
    {'科类': '物理类', '招生类型': '省属高校少数民族预科', '院校代码': 5142, '院校名称': '成都工业学院', '办学性质': '公办',
     '院校地址': '四川省成都市', '院校备注': None, '专业组代码': 130, '再选科目要求': '化学', '专业组计划数': 34,
     '专业代码': 'Y1', '专业名称': '少数民族预科',
     '专业备注': '预科培养地点为阿坝师范学院。宜宾校区就读。',
     '专业计划数': 34, '收费标准': 5500, '页码': 5},
]

# ---------- Helpers ----------
def norm(v):
    if v is None:
        return None
    if isinstance(v, float) and math.isnan(v):
        return None
    return str(v).strip()

def load_rows(path):
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    header = [c.value for c in ws[1]]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in r):
            continue
        rows.append(dict(zip(header, r)))
    return rows

def key(r):
    return (norm(r.get('科类')), norm(r.get('院校代码')),
            norm(r.get('专业组代码')), norm(r.get('专业代码')))

# ---------- Load sources ----------
mimo_rows = load_rows(MIMO)
corr_rows = load_rows(CORR)
v1_rows   = load_rows(V1)

mimo_map = {key(r): r for r in mimo_rows}
corr_map = {key(r): r for r in corr_rows}
v1_map   = {key(r): r for r in v1_rows}

# ---------- Build workbook ----------
wb = Workbook()
ws1 = wb.active
ws1.title = '真值数据'

yellow  = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # v1-already-corrected
orange  = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')  # v2 new delta
red     = PatternFill(start_color='FF6666', end_color='FF6666', fill_type='solid')  # needs human review
gray    = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')
header_font = Font(bold=True)

for ci, col in enumerate(COLS, start=1):
    c = ws1.cell(row=1, column=ci, value=col)
    c.font = header_font
    c.fill = gray

change_log = []
rule_hits = {'rule1_bracket': 0, 'rule2_forward_fill_type': 0, 'rule_other': 0}

for ri, gt in enumerate(GT_V2, start=2):
    k = key(gt)
    m  = mimo_map.get(k)
    cr = corr_map.get(k)
    v1 = v1_map.get(k)
    src_img = '4413_003.jpg' if gt['页码'] == 4 else '4413_004.jpg'

    for ci, col in enumerate(COLS, start=1):
        gt_val = gt.get(col)
        cell = ws1.cell(row=ri, column=ci, value=gt_val)

        m_val  = m.get(col)  if m  else None
        c_val  = cr.get(col) if cr else None
        v1_val = v1.get(col) if v1 else None

        gt_n = norm(gt_val)
        m_n  = norm(m_val)
        c_n  = norm(c_val)
        v1_n = norm(v1_val)

        # Determine fill color and change reason
        row_missing_in_mimo = (m is None)
        v1_matches_gt = (v1_n == gt_n)

        if row_missing_in_mimo and v1_matches_gt:
            # v1 added this row (historical Y5/Y8), v2 keeps the same
            cell.fill = yellow
            change_log.append({
                '行号': ri,
                '院校代码_专业代码': f"{gt.get('院校代码')}_{gt.get('专业代码')}",
                '列名': col,
                'mimo-v2-omni原值': '(整行缺失)',
                'corrected原值': c_val,
                'v1_verified值': v1_val,
                'v2最终真值': gt_val,
                '图片来源': src_img,
                '修改依据': 'mimo 漏此行，corrected 补入；v1/v2 均沿用',
                'v1_vs_v2差异说明': '无 — v1 已采纳 corrected 值，v2 沿用',
                '置信度': '高',
            })
            continue

        # Row exists in both engines. If all four agree (including v1), no mark.
        if gt_n == m_n and gt_n == c_n and v1_matches_gt:
            continue

        if v1_matches_gt and gt_n != m_n:
            # v1 already corrected it, v2 retains → YELLOW
            cell.fill = yellow
            if gt_n == c_n:
                reason = '采用 corrected 值（与图片一致）；v1 已修、v2 沿用'
            else:
                reason = '两版引擎均错，v1 已修、v2 沿用'
            change_log.append({
                '行号': ri,
                '院校代码_专业代码': f"{gt.get('院校代码')}_{gt.get('专业代码')}",
                '列名': col,
                'mimo-v2-omni原值': m_val,
                'corrected原值': c_val,
                'v1_verified值': v1_val,
                'v2最终真值': gt_val,
                '图片来源': src_img,
                '修改依据': reason,
                'v1_vs_v2差异说明': '无 — v1/v2 值相同',
                '置信度': '高',
            })
        else:
            # v1 != gt → v2 made a NEW change vs v1 → ORANGE
            cell.fill = orange
            # Classify by column
            if col == '招生类型':
                rule = 'Rule 2 (招生类型前向填充)'
                rule_hits['rule2_forward_fill_type'] += 1
                reason = ('图片中本节无"普通类"标题，应前向继承上节"区域教育均衡发展专项计划"；'
                          'v1 误采纳引擎的"普通类"默认值')
                delta_note = f'v1={v1_val!r} → v2={gt_val!r}（Rule 2 前向填充）'
            elif col in ('专业名称', '专业备注'):
                rule = 'Rule 1 (括号归属)'
                rule_hits['rule1_bracket'] += 1
                reason = ('图片原文中地名/校区/培养地点括号独立于专业名称列出，应归属专业备注；'
                          'v1 错误地将其合并进专业名称')
                delta_note = f'v1={v1_val!r} → v2={gt_val!r}（Rule 1 括号归属修正）'
            else:
                rule = '其他'
                rule_hits['rule_other'] += 1
                reason = '按图片核对'
                delta_note = f'v1={v1_val!r} → v2={gt_val!r}'

            change_log.append({
                '行号': ri,
                '院校代码_专业代码': f"{gt.get('院校代码')}_{gt.get('专业代码')}",
                '列名': col,
                'mimo-v2-omni原值': m_val,
                'corrected原值': c_val,
                'v1_verified值': v1_val,
                'v2最终真值': gt_val,
                '图片来源': src_img,
                '修改依据': f'[{rule}] {reason}（见 {src_img}）',
                'v1_vs_v2差异说明': delta_note,
                '置信度': '高',
            })

# Column widths
widths = {
    '科类': 8, '招生类型': 24, '院校代码': 10, '院校名称': 16, '办学性质': 10,
    '院校地址': 14, '院校备注': 30, '专业组代码': 10, '再选科目要求': 12,
    '专业组计划数': 12, '专业代码': 10, '专业名称': 34, '专业备注': 38,
    '专业计划数': 10, '收费标准': 10, '页码': 6,
}
for ci, col in enumerate(COLS, start=1):
    ws1.column_dimensions[ws1.cell(row=1, column=ci).column_letter].width = widths.get(col, 12)

# ====== Sheet 2: 修改记录 ======
ws2 = wb.create_sheet('修改记录')
cols2 = ['行号', '院校代码_专业代码', '列名', 'mimo-v2-omni原值', 'corrected原值',
         'v1_verified值', 'v2最终真值', '图片来源', '修改依据', 'v1_vs_v2差异说明', '置信度']
for ci, col in enumerate(cols2, start=1):
    c = ws2.cell(row=1, column=ci, value=col)
    c.font = header_font
    c.fill = gray
for ri, rec in enumerate(change_log, start=2):
    for ci, col in enumerate(cols2, start=1):
        val = rec.get(col)
        ws2.cell(row=ri, column=ci, value=val)
    # Color-code the row based on whether it's a v2 delta
    delta = rec.get('v1_vs_v2差异说明', '')
    if 'Rule' in delta or '→' in delta and '无' not in delta:
        for ci in range(1, len(cols2) + 1):
            ws2.cell(row=ri, column=ci).fill = orange
    else:
        for ci in range(1, len(cols2) + 1):
            ws2.cell(row=ri, column=ci).fill = yellow
for ci, col in enumerate(cols2, start=1):
    w = 28 if col in ('修改依据', 'v1_vs_v2差异说明') else 18
    ws2.column_dimensions[ws2.cell(row=1, column=ci).column_letter].width = w

# ====== Sheet 3: 引擎差异 ======
ws3 = wb.create_sheet('引擎差异')
cols3 = ['行号', '列名', 'mimo-v2-omni值', 'corrected值', 'v1_verified值', 'v2最终真值', '备注']
for ci, col in enumerate(cols3, start=1):
    c = ws3.cell(row=1, column=ci, value=col)
    c.font = header_font
    c.fill = gray

diff_rows = []
for ri, gt in enumerate(GT_V2, start=2):
    k = key(gt)
    m  = mimo_map.get(k)
    cr = corr_map.get(k)
    v1 = v1_map.get(k)
    for col in COLS:
        gt_val = gt.get(col)
        m_val  = m.get(col)  if m  else '(整行缺失)'
        c_val  = cr.get(col) if cr else '(整行缺失)'
        v1_val = v1.get(col) if v1 else '(整行缺失)'
        vals = [norm(m_val), norm(c_val), norm(v1_val), norm(gt_val)]
        if len(set(vals)) > 1:
            diff_rows.append({
                '行号': ri, '列名': col,
                'mimo-v2-omni值': m_val, 'corrected值': c_val,
                'v1_verified值': v1_val, 'v2最终真值': gt_val,
                '备注': '四者不完全一致',
            })

for ri, rec in enumerate(diff_rows, start=2):
    for ci, col in enumerate(cols3, start=1):
        ws3.cell(row=ri, column=ci, value=rec.get(col))
for ci, col in enumerate(cols3, start=1):
    ws3.column_dimensions[ws3.cell(row=1, column=ci).column_letter].width = 28

wb.save(OUT)

# ---- Print summary ----
yellow_cnt = sum(1 for r in change_log if '无' in r.get('v1_vs_v2差异说明', ''))
orange_cnt = sum(1 for r in change_log if 'Rule' in r.get('v1_vs_v2差异说明', ''))
print(f'Saved: {OUT}')
print(f'GT rows: {len(GT_V2)}')
print(f'Change log entries: {len(change_log)}')
print(f'  Yellow (v1 retained): {yellow_cnt}')
print(f'  Orange (v2 new delta): {orange_cnt}')
print(f'Rule hits:')
print(f'  Rule 1 (括号归属): {rule_hits["rule1_bracket"]}')
print(f'  Rule 2 (招生类型前向填充): {rule_hits["rule2_forward_fill_type"]}')
print(f'  Other: {rule_hits["rule_other"]}')
print(f'Engine-vs-verified diffs: {len(diff_rows)}')
