# -*- coding: utf-8 -*-
"""Verification script for batch 4428 (2025 历史类 专科批次 征集志愿 第一次)"""
import openpyxl, sys, re
sys.stdout.reconfigure(encoding='utf-8')

src_path = 'data/13_征集志愿/普通高考/专科批次/4428_2025_历史类_专科批次_征集志愿_第一次/4428_2025_历史类_专科批次_征集志愿_第一次_mimo-v2-omni.xlsx'
out_path = 'data/13_征集志愿/普通高考/专科批次/4428_2025_历史类_专科批次_征集志愿_第一次/4428_2025_历史类_专科批次_征集志愿_第一次_已校验.xlsx'

wb_src = openpyxl.load_workbook(src_path)
ws_src = wb_src.active

COL = {
    '科类': 1, '招生类型': 2, '院校代码': 3, '院校名称': 4, '办学性质': 5,
    '院校地址': 6, '院校备注': 7, '专业组代码': 8, '再选科目要求': 9,
    '专业组计划数': 10, '专业代码': 11, '专业名称': 12, '专业备注': 13,
    '专业计划数': 14, '收费标准': 15, '页码': 16
}
headers = list(COL.keys())

all_rows = []
for r in range(2, ws_src.max_row + 1):
    row = {}
    for h, c in COL.items():
        row[h] = ws_src.cell(r, c).value
    row['_src_row'] = r
    row['校正备注'] = ''
    all_rows.append(row)

corrections = {'部委属': 0, '空格': 0, '括号拆分': 0, '跨页补全': 0, '误归属': 0, '缺漏行': 0, '备注补全': 0}

# CORRECTION 1: 部委和 -> 部委属和
for row in all_rows:
    note = row.get('院校备注') or ''
    if '部委和外省属' in note:
        row['院校备注'] = note.replace('部委和外省属', '部委属和外省属')
        row['校正备注'] += '院校备注:部委->部委属；'
        corrections['部委属'] += 1

# CORRECTION 2: Remove extra spaces in 院校备注
for row in all_rows:
    note = row.get('院校备注') or ''
    if re.search(r'加分办法\s+\d', note):
        new_note = re.sub(r'加分办法\s+(\d)', r'加分办法\1', note)
        if new_note != note:
            row['院校备注'] = new_note
            row['校正备注'] += '院校备注:去空格；'
            corrections['空格'] += 1

# CORRECTION 3: Split parentheticals from 专业名称
for row in all_rows:
    name = row.get('专业名称') or ''
    if '(' not in name:
        continue
    idx = name.index('(')
    pure_name = name[:idx]
    paren_content = name[idx:]
    parts = re.findall(r'[\(\uff08]([^)\uff09]*)[\)\uff09]', paren_content)
    if parts:
        paren_text = '\uff1b'.join(parts)  # use Chinese semicolon
        existing_note = row.get('专业备注') or ''
        if existing_note:
            row['专业备注'] = paren_text + '\uff1b' + existing_note
        else:
            row['专业备注'] = paren_text
        row['专业名称'] = pure_name
        row['校正备注'] += '专业名称拆分括号；'
        corrections['括号拆分'] += 1

# CORRECTION 4a: 1343 rows 28-30 missing grp
for row in all_rows:
    r = row['_src_row']
    if 28 <= r <= 30 and row.get('专业组代码') is None:
        row['专业组代码'] = '101'
        row['再选科目要求'] = '不限'
        row['专业组计划数'] = 3
        row['校正备注'] += '跨页补全专业组(101)；'
        corrections['跨页补全'] += 1

# CORRECTION 4b: Page 29 rows 435-440 -> 4652
for row in all_rows:
    r = row['_src_row']
    if 435 <= r <= 440 and row.get('院校代码') is None:
        row['院校代码'] = '4652'
        row['院校名称'] = '海口经济学院'
        row['办学性质'] = '民办'
        row['院校地址'] = '海南省海口市'
        row['院校备注'] = '执行部委属和外省属高校少数民族地区加分办法1。'
        row['专业组代码'] = '101'
        row['再选科目要求'] = '不限'
        row['校正备注'] += '跨页补全(4652海口经济学院)；'
        corrections['跨页补全'] += 1

# CORRECTION 4c: Page 40 rows 696-741 -> 5186
for row in all_rows:
    r = row['_src_row']
    if 696 <= r <= 741 and row.get('院校代码') is None:
        row['院校代码'] = '5186'
        row['院校名称'] = '四川科技职业学院'
        row['办学性质'] = '民办'
        row['院校地址'] = '四川省眉山市'
        row['院校备注'] = '执行省属高校少数民族地区加分办法2。'
        row['专业组代码'] = '101'
        row['再选科目要求'] = '不限'
        row['专业组计划数'] = 981
        row['校正备注'] += '跨页补全(5186四川科技职业学院)；'
        corrections['跨页补全'] += 1

# CORRECTION 4d: Page 56 rows 1187-1214 -> 5786
for row in all_rows:
    r = row['_src_row']
    if 1187 <= r <= 1214 and row.get('院校代码') is None:
        row['院校代码'] = '5786'
        row['院校名称'] = '眉山药科职业学院'
        row['办学性质'] = '民办'
        row['院校地址'] = '四川省眉山市'
        row['院校备注'] = '执行省属高校少数民族地区加分办法2。'
        row['专业组代码'] = '101'
        row['再选科目要求'] = '不限'
        row['专业组计划数'] = 477
        row['校正备注'] += '跨页补全(5786眉山药科职业学院)；'
        corrections['跨页补全'] += 1

# CORRECTION 5: Fix 5785/5786 misattribution (rows 1143-1186)
for row in all_rows:
    r = row['_src_row']
    if 1143 <= r <= 1186 and row.get('院校代码') == '5786':
        row['院校代码'] = '5785'
        row['院校名称'] = '天府新区信息职业学院'
        row['院校地址'] = '四川省眉山市'
        row['院校备注'] = '执行省属高校少数民族地区加分办法2。'
        row['专业组代码'] = '101'
        row['再选科目要求'] = '不限'
        row['专业组计划数'] = 312
        row['校正备注'] += '院校代码5786->5785(跨页误归属)；'
        corrections['误归属'] += 1

# CORRECTION 6: Add missing rows for 2171 (page 8)
insert_idx = None
for i, row in enumerate(all_rows):
    if row['_src_row'] == 69:
        insert_idx = i + 1
        break

missing_2171 = [
    {
        '科类': '历史类', '招生类型': '普通类', '院校代码': '2171',
        '院校名称': '大连装备制造职业技术学院', '办学性质': '民办',
        '院校地址': '辽宁省大连市',
        '院校备注': '执行部委属和外省属高校少数民族地区加分办法1。',
        '专业组代码': '101', '再选科目要求': '不限', '专业组计划数': 4,
        '专业代码': '14', '专业名称': '汽车技术服务与营销', '专业备注': None,
        '专业计划数': 1, '收费标准': '9800', '页码': 8,
        '_src_row': -1, '校正备注': '缺漏行(图片可见,OCR遗漏)；'
    },
    {
        '科类': '历史类', '招生类型': '普通类', '院校代码': '2171',
        '院校名称': '大连装备制造职业技术学院', '办学性质': '民办',
        '院校地址': '辽宁省大连市',
        '院校备注': '执行部委属和外省属高校少数民族地区加分办法1。',
        '专业组代码': '101', '再选科目要求': '不限', '专业组计划数': 4,
        '专业代码': '20', '专业名称': '港口物流管理', '专业备注': None,
        '专业计划数': 1, '收费标准': '9800', '页码': 8,
        '_src_row': -1, '校正备注': '缺漏行(图片可见,OCR遗漏)；'
    },
    {
        '科类': '历史类', '招生类型': '普通类', '院校代码': '2171',
        '院校名称': '大连装备制造职业技术学院', '办学性质': '民办',
        '院校地址': '辽宁省大连市',
        '院校备注': '执行部委属和外省属高校少数民族地区加分办法1。',
        '专业组代码': '101', '再选科目要求': '不限', '专业组计划数': 4,
        '专业代码': '22', '专业名称': '旅游管理', '专业备注': None,
        '专业计划数': 1, '收费标准': '9800', '页码': 8,
        '_src_row': -1, '校正备注': '缺漏行(图片可见,OCR遗漏)；'
    },
]
if insert_idx is not None:
    for m in reversed(missing_2171):
        all_rows.insert(insert_idx, m)
        corrections['缺漏行'] += 1

# CORRECTION 7: Fill missing 院校备注 by same-school propagation
school_notes = {}
for row in all_rows:
    code = row.get('院校代码')
    note = row.get('院校备注')
    if code and note:
        school_notes[code] = note

for row in all_rows:
    code = row.get('院校代码')
    note = row.get('院校备注')
    if code and not note and code in school_notes:
        row['院校备注'] = school_notes[code]
        row['校正备注'] += '院校备注:同校补全；'
        corrections['备注补全'] += 1

# CORRECTION 8: Fill remaining None 院校备注 based on school code and image verification
# Pattern: Sichuan provincial schools (5xxx codes) use 省属 variant,
# out-of-province schools use 部委属和外省属 variant
# Determined by reviewing all relevant images
remaining_notes = {
    '1429': '执行部委属和外省属高校少数民族地区加分办法1。',
    '1449': '执行部委属和外省属高校少数民族地区加分办法1。',
    '1475': '执行部委属和外省属高校少数民族地区加分办法1。',
    '1558': '执行部委属和外省属高校少数民族地区加分办法1。',
    '1566': '执行部委属和外省属高校少数民族地区加分办法1。',
    '1846': '执行部委属和外省属高校少数民族地区加分办法1。',
    '1849': '执行部委属和外省属高校少数民族地区加分办法1。',
    '1850': '执行部委属和外省属高校少数民族地区加分办法1。',
    '1859': '执行部委属和外省属高校少数民族地区加分办法1。',
    '1908': '执行部委属和外省属高校少数民族地区加分办法1。',
    '2130': '执行部委属和外省属高校少数民族地区加分办法1。',
    '2137': '执行部委属和外省属高校少数民族地区加分办法1。',
    '2146': '执行部委属和外省属高校少数民族地区加分办法1。',
    '2152': '执行部委属和外省属高校少数民族地区加分办法1。',
    '2153': '执行部委属和外省属高校少数民族地区加分办法1。',
    '2161': '执行部委属和外省属高校少数民族地区加分办法1。',
    '2163': '执行部委属和外省属高校少数民族地区加分办法1。',
    '2164': '执行部委属和外省属高校少数民族地区加分办法1。',
    '3622': '执行部委属和外省属高校少数民族地区加分办法1。',
    '3683': '执行部委属和外省属高校少数民族地区加分办法1。',
    '3768': '执行部委属和外省属高校少数民族地区加分办法1。',
    '4173': '执行部委属和外省属高校少数民族地区加分办法1。',
    '4270': '执行部委属和外省属高校少数民族地区加分办法1。',
    '4589': '执行部委属和外省属高校少数民族地区加分办法1。',
    '4694': '执行部委属和外省属高校少数民族地区加分办法1。',
    '4783': '执行部委属和外省属高校少数民族地区加分办法1。',
    '4986': '执行部委属和外省属高校少数民族地区加分办法1。',
    '5067': '执行部委属和外省属高校少数民族地区加分办法1。',
    '5084': '执行部委属和外省属高校少数民族地区加分办法1。',
    '5140': '执行省属高校少数民族地区加分办法2。',
    '5164': '执行省属高校少数民族地区加分办法2。',
    '5192': '执行省属高校少数民族地区加分办法2。',
    '5253': '执行部委属和外省属高校少数民族地区加分办法1。',
    '5655': '执行省属高校少数民族地区加分办法2。',
    '5656': '执行省属高校少数民族地区加分办法2。',
    '5657': '执行省属高校少数民族地区加分办法2。',
    '5658': '执行省属高校少数民族地区加分办法2。',
    '5659': '执行省属高校少数民族地区加分办法2。',
    '5779': '执行省属高校少数民族地区加分办法2。',
    '5787': '执行省属高校少数民族地区加分办法2。',
    '5805': '执行省属高校少数民族地区加分办法2。',
}
for row in all_rows:
    code = row.get('院校代码')
    note = row.get('院校备注')
    if code and not note and code in remaining_notes:
        row['院校备注'] = remaining_notes[code]
        row['校正备注'] += '院校备注:根据图片补全；'
        corrections['备注补全'] += 1

# CORRECTION 9: Ensure all rows have 科类 and 招生类型
for row in all_rows:
    if not row.get('科类'):
        row['科类'] = '历史类'
    if not row.get('招生类型'):
        row['招生类型'] = '普通类'

# Print summary
print('=== Correction Summary ===')
for k, v in corrections.items():
    print(f'  {k}: {v}')
print(f'  Total rows: {len(all_rows)} (original: 1464, added: {len(all_rows) - 1464})')

# WRITE OUTPUT
from openpyxl.styles import Font, PatternFill

wb_out = openpyxl.Workbook()
ws_out = wb_out.active
ws_out.title = '征集志愿'

font_header = Font(name='微软雅黑', size=10, bold=True)
font_normal = Font(name='微软雅黑', size=10)
fill_yellow = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
fill_green = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')

out_headers = headers + ['校正备注']
for c, h in enumerate(out_headers, 1):
    cell = ws_out.cell(1, c, value=h)
    cell.font = font_header

for i, row in enumerate(all_rows, 2):
    is_added = row['_src_row'] == -1
    has_correction = bool(row['校正备注']) and not is_added

    for c, h in enumerate(out_headers, 1):
        val = row.get(h)
        cell = ws_out.cell(i, c, value=val)
        cell.font = font_normal

        if is_added:
            cell.fill = fill_green
        elif has_correction:
            notes = row['校正备注']
            if h == '院校备注' and '院校备注' in notes:
                cell.fill = fill_yellow
            elif h == '专业名称' and '专业名称拆分' in notes:
                cell.fill = fill_yellow
            elif h == '专业备注' and '专业名称拆分' in notes:
                cell.fill = fill_yellow
            elif h == '院校代码' and ('跨页补全' in notes or '误归属' in notes):
                cell.fill = fill_yellow
            elif h == '院校名称' and ('跨页补全' in notes or '误归属' in notes):
                cell.fill = fill_yellow
            elif h == '院校地址' and '跨页补全' in notes:
                cell.fill = fill_yellow
            elif h == '办学性质' and '跨页补全' in notes:
                cell.fill = fill_yellow
            elif h == '专业组代码' and ('跨页补全' in notes or '误归属' in notes):
                cell.fill = fill_yellow
            elif h == '再选科目要求' and ('跨页补全' in notes or '误归属' in notes):
                cell.fill = fill_yellow
            elif h == '专业组计划数' and ('跨页补全' in notes or '误归属' in notes):
                cell.fill = fill_yellow

# Set column widths
col_widths = {1: 8, 2: 10, 3: 10, 4: 30, 5: 10, 6: 25, 7: 45,
              8: 12, 9: 15, 10: 12, 11: 10, 12: 25, 13: 40,
              14: 10, 15: 10, 16: 6, 17: 40}
for c, w in col_widths.items():
    ws_out.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w

wb_out.save(out_path)
print(f'\nOutput saved: {out_path}')
print(f'Total rows: {len(all_rows)} (original: 1464, added: {len(all_rows) - 1464})')
